import csv
import base64
import html
import mimetypes
import io
import hashlib
import json
import os
import re
import smtplib
import ssl
import sqlite3
import subprocess
import sys
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
import uuid
import xml.etree.ElementTree as ET
from copy import deepcopy
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from dataclasses import dataclass
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from urllib.parse import quote
from zipfile import BadZipFile, ZipFile, ZIP_DEFLATED
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st
import httpx
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openai import APIConnectionError, APIStatusError, AuthenticationError, OpenAI, RateLimitError

try:
    import fitz
except ImportError:
    fitz = None

try:
    import numpy as np
except ImportError:
    np = None

try:
    from PIL import Image, ImageDraw, ImageEnhance, ImageFilter, ImageFont
except ImportError:
    Image = None
    ImageDraw = None
    ImageEnhance = None
    ImageFilter = None
    ImageFont = None

try:
    import truststore
except ImportError:
    truststore = None

try:
    from langsmith.wrappers import wrap_openai
except ImportError:
    wrap_openai = None


BASE_DIR = Path(__file__).parent
ASSETS_DIR = BASE_DIR / "assets"
SIDEBAR_LOGO_PATH = ASSETS_DIR / "controls-logo.png"
SIDEBAR_FOOTER_PATH = ASSETS_DIR / "pe-battery-footer.png"
APP_STARTED_AT = datetime.now()
ENV_PATH = BASE_DIR / "app.env"
DEFAULT_GLOSSARY_PATHS = [
    BASE_DIR / "glossary.xlsx",
    BASE_DIR / "glossary.csv",
]
DEFAULT_PLC_RULE_PATHS = [
    BASE_DIR / "plc_abbreviation_rules.xlsx",
    BASE_DIR / "plc_abbreviation_rules.csv",
]
DEFAULT_MODEL = "gpt-4.1-mini-2025-04-14"
DOCUMENT_BATCH_SIZE = 25
MAX_PARALLEL_BATCHES = 12
MAX_TRANSLATION_RETRIES = 2
PPT_CONTEXT_TRANSLATION_VERSION = "ppt-textbox-unit-v2"
PDF_EXTRACTION_VERSION = "pdf-visual-line-unit-v1"
OPENAI_TIMEOUT_SECONDS = 60
HMI_MAX_VISION_BOXES = 80
HMI_BOX_VISION_BATCH_SIZE = 20
MAX_UPLOAD_BYTES = 100 * 1024 * 1024
MAX_EMAIL_ATTACHMENT_BYTES = 20 * 1024 * 1024
TRANSLATION_USAGE_SINCE_LABEL = "Jul 17, 2026"
TEXT_TRANSLATION_USAGE_KEY = "text_translation_started"
DOCUMENT_TRANSLATION_USAGE_KEY = "document_translation_started"
IMAGE_TRANSLATION_USAGE_KEY = "image_translation_started"
PROGRESS_DIR = BASE_DIR / ".term1_progress"
USAGE_COUNT_PATH = BASE_DIR / ".term1_usage_count.json"
JOB_DB_PATH = BASE_DIR / ".term1_jobs.db"
TRANSLATION_MEMORY_DB_PATH = BASE_DIR / "translation_memory.sqlite"
JOB_STORAGE_DIR = BASE_DIR / ".term1_job_storage"
JOB_UPLOAD_DIR = JOB_STORAGE_DIR / "uploads"
JOB_RESULT_DIR = JOB_STORAGE_DIR / "results"
LARGE_PDF_WORK_DIR = JOB_STORAGE_DIR / "large-pdf-jobs"
LARGE_PDF_PAGE_THRESHOLD = 300
LARGE_PDF_BLOCK_THRESHOLD = 10_000
GENERAL_TRANSLATION_MODE = "General Plant Document"
PLC_TRANSLATION_MODE = "PLC/SPLC Comment Style"
SUPPLIER_EMAIL_TRANSLATION_MODE = "Business Email Style"
PRODUCT_CATALOG_TRANSLATION_MODE = "Catalog Document Style"
POWERPOINT_TRANSLATION_MODE = "PowerPoint Presentation Style"
ROBOT_PROGRAM_TRANSLATION_MODE = "Kawasaki Robot .as file"
HMI_SCREEN_TRANSLATION_MODE = "HMI Screen Translation"
IMAGE_MODE_HMI = "HMI Screen"
IMAGE_MODE_ENGINEERING = "Other Image / CAD / Drawing"
IMAGE_MODE_GENERAL = "General Image"
IMAGE_TRANSLATION_MODES = [
    IMAGE_MODE_HMI,
    IMAGE_MODE_ENGINEERING,
    IMAGE_MODE_GENERAL,
]
HMI_REVIEW_IMPORTANT = "Key review labels"
HMI_REVIEW_ALL = "All detected labels"
HMI_REVIEW_DETAIL_MODES = [
    HMI_REVIEW_IMPORTANT,
    HMI_REVIEW_ALL,
]
TRANSLATION_MODES = [
    PLC_TRANSLATION_MODE,
    GENERAL_TRANSLATION_MODE,
    SUPPLIER_EMAIL_TRANSLATION_MODE,
    PRODUCT_CATALOG_TRANSLATION_MODE,
    POWERPOINT_TRANSLATION_MODE,
    ROBOT_PROGRAM_TRANSLATION_MODE,
]
TRANSLATION_DIRECTION_JP_EN = "JP → EN"
TRANSLATION_DIRECTION_EN_JP = "EN → JP"
TRANSLATION_DIRECTIONS = [
    TRANSLATION_DIRECTION_JP_EN,
    TRANSLATION_DIRECTION_EN_JP,
]
TEXT_TRANSLATION_MODES = [
    PLC_TRANSLATION_MODE,
    GENERAL_TRANSLATION_MODE,
    SUPPLIER_EMAIL_TRANSLATION_MODE,
    PRODUCT_CATALOG_TRANSLATION_MODE,
]
DOCUMENT_TRANSLATION_MODES = [
    PLC_TRANSLATION_MODE,
    GENERAL_TRANSLATION_MODE,
    SUPPLIER_EMAIL_TRANSLATION_MODE,
    PRODUCT_CATALOG_TRANSLATION_MODE,
    POWERPOINT_TRANSLATION_MODE,
]
DOCUMENT_TRANSLATION_MODE_CAPTIONS = {
    PLC_TRANSLATION_MODE: "Short standardized output. Example: 自動運転中 -> Automatic Operation Active",
    GENERAL_TRANSLATION_MODE: "Clear manufacturing document translation.",
    SUPPLIER_EMAIL_TRANSLATION_MODE: "Natural business English nuance.",
    PRODUCT_CATALOG_TRANSLATION_MODE: "Concise product/spec wording.",
    POWERPOINT_TRANSLATION_MODE: "Concise slide-ready wording for titles, bullets, tables, and callouts.",
}
DOCUMENT_TRANSLATION_MODE_LABELS = {
    PLC_TRANSLATION_MODE: "MES Files & PLC Comments (short standardized output)",
    GENERAL_TRANSLATION_MODE: "General Plant (clear manufacturing wording)",
    SUPPLIER_EMAIL_TRANSLATION_MODE: "Business Email (natural business English)",
    PRODUCT_CATALOG_TRANSLATION_MODE: "Catalog / Specs (concise product/spec wording)",
    POWERPOINT_TRANSLATION_MODE: "PowerPoint Presentation (concise slide wording)",
}
TEXT_TRANSLATION_MODE_LABELS = {
    PLC_TRANSLATION_MODE: "PLC/SPLC Comment",
    GENERAL_TRANSLATION_MODE: "General Plant Text",
    SUPPLIER_EMAIL_TRANSLATION_MODE: "Business Email",
    PRODUCT_CATALOG_TRANSLATION_MODE: "Catalog Text",
}
PLC_DUPLICATE_STATUS_WORDS = [
    "ON",
    "OFF",
    "OK",
    "NG",
    "Mode",
    "Error",
    "Complete",
    "Confirm",
    "Request",
    "Command",
    "Present",
    "Absent",
]
PLC_SYNONYM_CLEANUPS = [
    (re.compile(r"\bpoor,\s*defective,\s*NG,\s*inoperative\b", re.IGNORECASE), "NG"),
    (re.compile(r"\bdefective,\s*NG\b", re.IGNORECASE), "NG"),
]
PROTECTED_PATTERN = re.compile(
    r"\b(?:[A-Z]{1,6}[-_]?\d{1,6}[A-Z]?|\d+[A-Z]{1,4}|[XYMDSZR][0-9]{1,5}|[A-Z]{2,}-[A-Z0-9-]+)\b"
)
ENCLOSED_ALNUM_PATTERN = re.compile(r"[\u2460-\u24ff\u3200-\u32ff]")
LEADING_CODE_PATTERN = re.compile(r"^([A-Z]{1,6}[-_]?\d{1,6}[A-Z]?)(.*)$")

WORD_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
PPT_NS = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
EXCEL_NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
EXCEL_SERIALIZE_NAMESPACES = {
    "": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "mc": "http://schemas.openxmlformats.org/markup-compatibility/2006",
    "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
    "xr": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
    "xr2": "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2",
    "xr3": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3",
}


@dataclass(frozen=True)
class TermHit:
    jp: str
    en: str
    count: int


@dataclass(frozen=True)
class TextBlock:
    location: str
    text: str
    context: str = ""
    context_group: str = ""


@dataclass(frozen=True)
class HmiTextRegion:
    location: str
    jp: str
    x: int
    y: int
    width: int
    height: int
    confidence: float
    note: str = ""
    kind: str = ""


@dataclass(frozen=True)
class HmiDetectedBox:
    no: int
    x: int
    y: int
    width: int
    height: int


@dataclass
class TokenUsage:
    input_tokens: int = 0
    output_tokens: int = 0
    total_tokens: int = 0

    def add(self, other: "TokenUsage") -> None:
        self.input_tokens += other.input_tokens
        self.output_tokens += other.output_tokens
        self.total_tokens += other.total_tokens

    def display(self) -> str:
        if self.total_tokens <= 0:
            return "Token usage unavailable."
        return (
            f"Tokens: input {self.input_tokens:,}, "
            f"output {self.output_tokens:,}, total {self.total_tokens:,}."
        )


def load_env() -> None:
    if ENV_PATH.exists():
        load_dotenv(ENV_PATH)
    else:
        load_dotenv()
    if truststore is not None:
        truststore.inject_into_ssl()


def clean_text(value: str) -> str:
    text = str(value)
    protected = {}

    def stash(match: re.Match) -> str:
        token = f"__TERM1_ENCLOSED_{len(protected)}__"
        protected[token] = match.group(0)
        return token

    normalized = unicodedata.normalize("NFKC", ENCLOSED_ALNUM_PATTERN.sub(stash, text)).strip()
    for token, original in protected.items():
        normalized = normalized.replace(token, original)
    return normalized


def has_japanese_text(value: str) -> bool:
    return bool(re.search(r"[\u3040-\u30ff\u3400-\u9fff]", value))


def has_english_text(value: str) -> bool:
    text = clean_text(value)
    if not text or text.startswith("=") or PROTECTED_PATTERN.fullmatch(text):
        return False
    return bool(re.search(r"[A-Za-z]{2,}", text))


def direction_language_names(translation_direction: str) -> tuple[str, str]:
    if translation_direction == TRANSLATION_DIRECTION_EN_JP:
        return "English", "Japanese"
    return "Japanese", "English"


def translation_direction_key(translation_direction: str) -> str:
    return "en_to_jp" if translation_direction == TRANSLATION_DIRECTION_EN_JP else "jp_to_en"


def translation_cache_mode(translation_mode: str, translation_direction: str) -> str:
    return f"{translation_mode}::{translation_direction_key(translation_direction)}"


def should_translate(
    text: str,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> bool:
    cleaned = clean_text(text)
    if translation_direction == TRANSLATION_DIRECTION_EN_JP:
        return has_english_text(cleaned)
    return has_japanese_text(cleaned)


def decode_document_text_with_encoding(raw: bytes) -> tuple[str, str]:
    candidates = []
    has_utf16_bom = raw.startswith((b"\xff\xfe", b"\xfe\xff"))
    null_ratio = raw.count(b"\x00") / max(len(raw), 1)
    encodings = ["utf-8-sig", "utf-8", "cp932", "shift_jis", "euc_jp", "iso2022_jp"]
    if has_utf16_bom or null_ratio > 0.1:
        encodings = ["utf-16", "utf-16-le", "utf-16-be"] + encodings
    encodings.append("cp1252")

    for encoding in encodings:
        try:
            text = raw.decode(encoding)
        except UnicodeDecodeError:
            continue

        japanese_count = len(re.findall(r"[\u3040-\u30ff\u3400-\u9fff]", text))
        replacement_count = text.count("ï¿½")
        mojibake_count = sum(text.count(marker) for marker in ("Ã£", "Ã‚", "Ãƒ", "\x00"))
        control_count = sum(1 for char in text if ord(char) < 32 and char not in "\r\n\t")
        c1_count = sum(1 for char in text if 0x80 <= ord(char) <= 0x9F)
        ascii_count = sum(1 for char in text if ord(char) < 128)
        score = (
            japanese_count * 50
            + min(ascii_count, 200) * 0.01
            - replacement_count * 200
            - mojibake_count * 80
            - control_count * 60
            - c1_count * 40
        )
        candidates.append((score, encoding, text))

    if candidates:
        _score, encoding, text = max(candidates, key=lambda candidate: candidate[0])
        return text, encoding

    return raw.decode("utf-8", errors="replace"), "utf-8"


def decode_document_text(raw: bytes) -> str:
    return decode_document_text_with_encoding(raw)[0]


def robot_program_decode_score(text: str) -> float:
    comment_text = "\n".join(value for _start, _end, value in robot_comment_segments(text))
    japanese_comment_count = len(re.findall(r"[\u3040-\u30ff\u3400-\u9fff]", comment_text))
    total_japanese_count = len(re.findall(r"[\u3040-\u30ff\u3400-\u9fff]", text))
    replacement_count = text.count("\ufffd")
    mojibake_count = sum(text.count(marker) for marker in ("ÃƒÂ£", "Ãƒâ€š", "ÃƒÆ’", "éƒç·’ç”³", "éƒ", "\x00"))
    control_count = sum(1 for char in text if ord(char) < 32 and char not in "\r\n\t")
    c1_count = sum(1 for char in text if 0x80 <= ord(char) <= 0x9F)
    return (
        japanese_comment_count * 200
        + total_japanese_count * 10
        - replacement_count * 500
        - mojibake_count * 100
        - control_count * 80
        - c1_count * 40
    )


def decode_robot_program_text_with_encoding(raw: bytes) -> tuple[str, str]:
    encodings = [
        "euc_jp",
        "cp932",
        "shift_jis",
        "utf-8-sig",
        "utf-8",
        "iso2022_jp",
    ]
    strict_candidates = []
    for encoding in encodings:
        try:
            text = raw.decode(encoding)
        except UnicodeDecodeError:
            continue
        strict_candidates.append((robot_program_decode_score(text), encoding, text))

    if strict_candidates:
        _score, encoding, text = max(strict_candidates, key=lambda candidate: candidate[0])
        return text, encoding

    replacement_candidates = []
    for encoding in encodings:
        text = raw.decode(encoding, errors="replace")
        replacement_candidates.append((robot_program_decode_score(text), encoding, text))
    _score, encoding, text = max(replacement_candidates, key=lambda candidate: candidate[0])
    return text, encoding


def looks_like_mojibake(value: str) -> bool:
    markers = ("ï¿½", "éƒ", "éƒç·’ç”³", "ÃƒÂ£", "Ãƒâ€š", "ÃƒÆ’")
    return any(marker in value for marker in markers)


def has_robot_mojibake_marker(value: str) -> bool:
    markers = (
        "\ufffd",
        chr(0x9403),
        chr(0x9403) + chr(0x7DD2) + chr(0x7533),
        "ÃƒÂ£",
        "Ãƒâ€š",
        "ÃƒÆ’",
    )
    return any(marker in value for marker in markers) or looks_like_mojibake(value)


def compact_warning_line(line: str, limit: int = 180) -> str:
    compact = re.sub(r"\s+", " ", line).strip()
    if len(compact) <= limit:
        return compact
    return compact[: limit - 3].rstrip() + "..."


def robot_block_ranges(raw: bytes) -> list[tuple[int, int]]:
    ranges = []
    try:
        blocks = extract_robot_program_blocks(raw)
    except Exception:
        return ranges
    for block in blocks:
        parts = block.location.split(":")
        if len(parts) != 3 or not parts[1].isdigit() or not parts[2].isdigit():
            continue
        ranges.append((int(parts[1]), int(parts[2])))
    return ranges


def robot_encoding_warning(raw: bytes, file_name: str) -> str:
    if not file_name.lower().endswith((".as", ".ad")):
        return ""

    replacement_bytes = raw.count(b"\xef\xbf\xbd")
    warning_parts = []
    if replacement_bytes:
        warning_parts.append(
            f"{replacement_bytes:,} replacement-character byte sequence(s) were found. "
            "This usually means Japanese text was already damaged by a wrong encoding conversion."
        )

    try:
        text, encoding = decode_robot_program_text_with_encoding(raw)
    except Exception:
        return (
            "Encoding warning: this robot program could not be decoded reliably. "
            "Please request the original exported .as/.ad file."
        )

    bad_lines = []
    suspicious_segments = 0
    unsupported_japanese_lines = []
    translated_ranges = robot_block_ranges(raw)
    line_start = 0
    for line_number, raw_line in enumerate(text.splitlines(keepends=True), start=1):
        line = raw_line.rstrip("\r\n")
        line_end = line_start + len(line)
        has_bad_replacement = "\ufffd" in line
        comment = line.split(";", 1)[1] if ";" in line else ""
        has_bad_comment = bool(comment and has_robot_mojibake_marker(comment))
        has_japanese_outside_supported_fields = (
            has_japanese_text(line)
            and not has_bad_replacement
            and not has_bad_comment
            and not any(start < line_end and end > line_start for start, end in translated_ranges)
        )
        if has_bad_comment:
            suspicious_segments += 1
        if has_japanese_outside_supported_fields:
            unsupported_japanese_lines.append(line_number)
        if has_bad_replacement or has_bad_comment or has_japanese_outside_supported_fields:
            bad_lines.append((line_number, compact_warning_line(line)))
        line_start += len(raw_line)

    if suspicious_segments:
        warning_parts.append(
            f"{suspicious_segments:,} robot comment line(s) look like mojibake/fake Japanese after decoding as {encoding}."
        )
    if unsupported_japanese_lines:
        warning_parts.append(
            f"{len(unsupported_japanese_lines):,} line(s) contain Japanese outside the supported Kawasaki comment/label fields, so they were not translated."
        )

    if not warning_parts:
        return ""

    line_details = ""
    if bad_lines:
        line_details = (
            "\n\nProblem line(s) to send to the robot/program team:\n"
            + "\n".join(f"Line {line_number}: {line}" for line_number, line in bad_lines[:20])
        )
        if len(bad_lines) > 20:
            line_details += f"\n...and {len(bad_lines) - 20:,} more suspicious line(s)."

    return (
        "AS File Warning: This Kawasaki AS/AD file has Japanese text that may not be safely translated or written back. "
        + " ".join(warning_parts)
        + " Review the listed line(s) before translation. If the lines contain corrupted Japanese, do not open and re-save the robot file with Notepad, Excel, or browser preview; ask for the original exported file or a file confirmed to display Japanese correctly."
        + line_details
    )


def document_fingerprint(file_name: str, raw: bytes) -> str:
    digest = hashlib.sha256(raw).hexdigest()[:16]
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(file_name).stem).strip("_") or "document"
    return f"{safe_name}_{len(raw)}_{digest}"


def checkpoint_path_for(
    file_name: str,
    raw: bytes,
    translation_mode: str = GENERAL_TRANSLATION_MODE,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
    owner_session_id: str = "",
) -> Path:
    cache_mode = translation_cache_mode(translation_mode, translation_direction)
    mode_key = re.sub(r"[^A-Za-z0-9_.-]+", "_", cache_mode).strip("_").lower()
    if file_name.lower().endswith(".pptx"):
        mode_key = f"{mode_key}_{PPT_CONTEXT_TRANSLATION_VERSION}"
    elif file_name.lower().endswith(".pdf"):
        mode_key = f"{mode_key}_{PDF_EXTRACTION_VERSION}"
    owner_key = hashlib.sha256((owner_session_id or "shared").encode("utf-8")).hexdigest()[:12]
    return PROGRESS_DIR / f"{document_fingerprint(file_name, raw)}_{mode_key}_{owner_key}.json"


def load_checkpoint(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    translations = data.get("translations", {})
    if not isinstance(translations, dict):
        return {}
    return {str(key): str(value) for key, value in translations.items()}


def save_checkpoint(path: Path, translations: dict[str, str]) -> None:
    PROGRESS_DIR.mkdir(exist_ok=True)
    payload = {
        "updated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "translations": translations,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def format_duration(seconds: float) -> str:
    seconds = max(int(seconds), 0)
    minutes, second = divmod(seconds, 60)
    hours, minute = divmod(minutes, 60)
    if hours:
        return f"{hours}h {minute}m"
    if minute:
        return f"{minute}m {second}s"
    return f"{second}s"


def format_file_size(size_bytes: int) -> str:
    if size_bytes >= 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.1f} MB"
    if size_bytes >= 1024:
        return f"{size_bytes / 1024:.0f} KB"
    return f"{size_bytes} B"


def estimate_remaining_time(done: int, total: int, elapsed_seconds: float) -> str:
    if done <= 0 or total <= 0 or done >= total or elapsed_seconds <= 0:
        return ""
    progress_ratio = done / total
    if progress_ratio < 0.25:
        return ""
    remaining_seconds = (elapsed_seconds / done) * max(total - done, 0)
    return format_duration(remaining_seconds)


def progress_text(done: int, total: int, elapsed_seconds: float | None = None) -> str:
    if total <= 0:
        return "Preparing file"
    if done >= total:
        return "Complete"
    if done <= 0:
        return "Starting translation"
    return "Translating"


def progress_percent(done: int, total: int) -> str:
    if total <= 0:
        return "Preparing"
    return f"{min(max(done / total, 0), 1) * 100:.2f}%"


def elapsed_since_timestamp(timestamp_text: str) -> float | None:
    try:
        started_at = datetime.strptime(timestamp_text, "%Y-%m-%d %H:%M:%S")
    except (TypeError, ValueError):
        return None
    return max((datetime.now() - started_at).total_seconds(), 0)


def parse_timestamp(timestamp_text: str) -> datetime | None:
    try:
        return datetime.strptime(timestamp_text, "%Y-%m-%d %H:%M:%S")
    except (TypeError, ValueError):
        return None


def render_download_ready(data: bytes, file_name: str, mime: str, key: str = "download_ready") -> None:
    st.success("Complete | Download ready")
    st.download_button(
        "Download Translated File",
        data=data,
        file_name=file_name,
        mime=mime,
        type="primary",
        key=key,
    )


def translation_pairs_preview(
    blocks: list[TextBlock],
    translations: dict[str, str],
    limit: int = 200,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
    glossary: pd.DataFrame | None = None,
) -> pd.DataFrame:
    rows = []
    translated_pair_count = 0
    source_language, target_language = direction_language_names(translation_direction)
    for block in blocks:
        if block.location not in translations or not should_translate(block.text, translation_direction):
            continue
        target_text = clean_text(translations[block.location])
        source_text = clean_text(block.text)
        if not target_text or target_text == source_text:
            continue
        base_row = {
                "Location": block.location,
            source_language: source_text,
            target_language: target_text,
        }
        glossary_rows = []
        if glossary is not None and not glossary.empty:
            _, hits = apply_glossary_to_source(
                block.text,
                glossary,
                replace_source=False,
                translation_direction=translation_direction,
            )
            for hit in hits:
                matched_records = glossary[
                    (glossary["JP"].astype(str).map(clean_text) == hit.jp)
                    & (glossary["EN"].astype(str).map(clean_text) == hit.en)
                ]
                glossary_record = matched_records.iloc[0].to_dict() if not matched_records.empty else {
                    "JP": hit.jp,
                    "EN": hit.en,
                }
                trace_row = {
                    "Glossary Match": f"{hit.jp} → {hit.en}",
                    "Match Count": hit.count,
                }
                trace_row.update(
                    {
                        f"Glossary {column}": glossary_record.get(column, "")
                        for column in glossary.columns
                    }
                )
                glossary_rows.append(trace_row)

        if glossary_rows:
            rows.extend({**base_row, **glossary_row} for glossary_row in glossary_rows)
        else:
            empty_trace = {"Glossary Match": "", "Match Count": ""}
            if glossary is not None:
                empty_trace.update({f"Glossary {column}": "" for column in glossary.columns})
            rows.append({**base_row, **empty_trace})

        translated_pair_count += 1
        if translated_pair_count >= limit:
            break
    return pd.DataFrame(rows)


def render_translation_pairs_preview(
    raw_document: bytes,
    file_name: str,
    translation_mode: str,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
    owner_session_id: str = "",
    max_rows: int = 200,
    glossary: pd.DataFrame | None = None,
) -> None:
    try:
        blocks = extract_text_blocks(raw_document, file_name)
        translations = load_checkpoint(
            checkpoint_path_for(
                file_name,
                raw_document,
                translation_mode,
                translation_direction,
                owner_session_id,
            )
        )
        preview = translation_pairs_preview(
            blocks,
            translations,
            max_rows,
            translation_direction,
            glossary,
        )
    except Exception as exc:
        st.caption(f"Translation preview unavailable: {exc}")
        return

    if preview.empty:
        st.caption("Translation preview: no translated source text was available for this file.")
        return

    st.subheader(f"{translation_direction} Preview")
    st.caption(
        "Glossary matches include every available glossary column for traceability, "
        "including collector, verifier, and date fields when present in the source glossary."
    )
    st.dataframe(preview, use_container_width=True, hide_index=True)
    total_pairs = sum(
        1
        for block in blocks
        if block.location in translations and should_translate(block.text, translation_direction)
    )
    shown_pairs = min(total_pairs, max_rows)
    if total_pairs > shown_pairs:
        st.caption(f"Showing first {shown_pairs:,} of {total_pairs:,} translated pairs.")


def start_background_translation_job(
    raw_document: bytes,
    file_name: str,
    blocks: list[TextBlock],
    glossary: pd.DataFrame,
    translation_mode: str,
    translation_direction: str,
    keep_source_with_translation: bool,
    notify_email: str,
    batch_count: int,
    progress_path: Path,
    owner_session_id: str,
) -> str:
    translatable_blocks = [block for block in blocks if should_translate(block.text, translation_direction)]
    job_id = create_translation_job(
        file_name,
        len(raw_document),
        len(blocks),
        len(translatable_blocks),
        batch_count,
        translation_mode,
        translation_direction,
        owner_session_id=owner_session_id,
        notify_email=notify_email,
        status="pending",
    )
    source_path = job_upload_path(job_id, file_name)
    source_path.write_bytes(raw_document)
    update_translation_job(job_id, source_file_path=str(source_path))
    background_job_executor().submit(
        run_document_translation_job,
        job_id,
        raw_document,
        file_name,
        blocks,
        glossary,
        translation_mode,
        translation_direction,
        keep_source_with_translation,
        progress_path,
        batch_count,
    )
    return job_id


def start_queued_document_translation_job(
    raw_document: bytes,
    file_name: str,
    glossary: pd.DataFrame,
    translation_mode: str,
    translation_direction: str,
    keep_source_with_translation: bool,
    notify_email: str,
    owner_session_id: str,
) -> str:
    stop_active_translation_jobs_for_file(
        file_name,
        "Stopped automatically because a newer job was started for this file.",
        owner_session_id,
    )
    job_id = create_translation_job(
        file_name,
        len(raw_document),
        0,
        0,
        0,
        translation_mode,
        translation_direction,
        owner_session_id=owner_session_id,
        notify_email=notify_email,
        status="pending",
    )
    source_path = job_upload_path(job_id, file_name)
    source_path.write_bytes(raw_document)
    update_translation_job(
        job_id,
        source_file_path=str(source_path),
        progress_message="Queued. Preparing file.",
    )
    background_job_executor().submit(
        prepare_and_run_document_translation_job,
        job_id,
        raw_document,
        file_name,
        glossary,
        translation_mode,
        translation_direction,
        keep_source_with_translation,
        owner_session_id,
    )
    return job_id


def clean_office_xml_text(value: str) -> str:
    # Office XML files cannot contain most control characters.
    return "".join(
        char
        for char in str(value)
        if char in "\t\n\r" or ord(char) >= 32
    )


def rerun_app() -> None:
    rerun = getattr(st, "rerun", None) or getattr(st, "experimental_rerun", None)
    if rerun:
        rerun()


def current_client_session_id() -> str:
    existing = str(st.session_state.get("client_session_id") or "")
    if re.fullmatch(r"client_[a-f0-9]{32}", existing):
        return existing

    query_value = st.query_params.get("client", "")
    if isinstance(query_value, list):
        query_value = query_value[0] if query_value else ""
    candidate = str(query_value or "")
    if not re.fullmatch(r"client_[a-f0-9]{32}", candidate):
        candidate = f"client_{uuid.uuid4().hex}"
        st.query_params["client"] = candidate
    st.session_state["client_session_id"] = candidate
    return candidate


def read_usage_count() -> int:
    return int(read_usage_stats().get("count", 0))


def read_usage_stats() -> dict[str, int | str]:
    if not USAGE_COUNT_PATH.exists():
        return {"count": 0, "sessions": 0, "actions": 0, "updated_at": ""}
    try:
        data = json.loads(USAGE_COUNT_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {"count": 0, "sessions": 0, "actions": 0, "updated_at": ""}
    count = int(data.get("count", 0))
    return {
        "count": count,
        "sessions": int(data.get("sessions", count)),
        "actions": int(data.get("actions", 0)),
        "updated_at": str(data.get("updated_at", "")),
    }


def write_usage_stats(stats: dict[str, int | str]) -> int:
    count = int(stats.get("count", 0))
    payload = {
        "count": count,
        "sessions": int(stats.get("sessions", 0)),
        "actions": int(stats.get("actions", 0)),
        "updated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    USAGE_COUNT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return count


def increment_usage_session_once() -> int:
    if st.session_state.get("usage_session_counted"):
        return read_usage_count()
    stats = read_usage_stats()
    stats["count"] = int(stats.get("count", 0)) + 1
    stats["sessions"] = int(stats.get("sessions", 0)) + 1
    count = write_usage_stats(stats)
    st.session_state["usage_session_counted"] = True
    return count


def increment_usage_action(action_name: str = "action") -> int:
    stats = read_usage_stats()
    stats["count"] = int(stats.get("count", 0)) + 1
    stats["actions"] = int(stats.get("actions", 0)) + 1
    stats["last_action"] = action_name
    return write_usage_stats(stats)


def init_translation_usage_store() -> None:
    with sqlite3.connect(JOB_DB_PATH, timeout=30) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS translation_usage_metrics (
                metric_key TEXT PRIMARY KEY,
                use_count INTEGER NOT NULL DEFAULT 0,
                first_recorded_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )


def read_translation_usage_count(metric_key: str) -> int:
    init_translation_usage_store()
    with sqlite3.connect(JOB_DB_PATH, timeout=30) as conn:
        row = conn.execute(
            "SELECT use_count FROM translation_usage_metrics WHERE metric_key = ?",
            (metric_key,),
        ).fetchone()
    return int(row[0] if row else 0)


def increment_translation_usage_count(metric_key: str) -> int:
    init_translation_usage_store()
    now = time.strftime("%Y-%m-%d %H:%M:%S")
    with sqlite3.connect(JOB_DB_PATH, timeout=30) as conn:
        conn.execute(
            """
            INSERT INTO translation_usage_metrics (
                metric_key, use_count, first_recorded_at, updated_at
            )
            VALUES (?, 1, ?, ?)
            ON CONFLICT(metric_key) DO UPDATE SET
                use_count = translation_usage_metrics.use_count + 1,
                updated_at = excluded.updated_at
            """,
            (metric_key, now, now),
        )
        row = conn.execute(
            "SELECT use_count FROM translation_usage_metrics WHERE metric_key = ?",
            (metric_key,),
        ).fetchone()
    return int(row[0] if row else 0)


def init_job_store() -> None:
    with sqlite3.connect(JOB_DB_PATH) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS translation_jobs (
                job_id TEXT PRIMARY KEY,
                file_name TEXT NOT NULL,
                file_size_bytes INTEGER NOT NULL,
                status TEXT NOT NULL,
                total_blocks INTEGER DEFAULT 0,
                translatable_blocks INTEGER DEFAULT 0,
                completed_blocks INTEGER DEFAULT 0,
                total_batches INTEGER DEFAULT 0,
                completed_batches INTEGER DEFAULT 0,
                input_tokens INTEGER DEFAULT 0,
                output_tokens INTEGER DEFAULT 0,
                total_tokens INTEGER DEFAULT 0,
                result_file_name TEXT DEFAULT '',
                error_message TEXT DEFAULT '',
                owner_session_id TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                finished_at TEXT DEFAULT ''
            )
            """
        )
        columns = {
            row[1]
            for row in conn.execute("PRAGMA table_info(translation_jobs)").fetchall()
        }
        if "translation_mode" not in columns:
            conn.execute(
                "ALTER TABLE translation_jobs ADD COLUMN translation_mode TEXT DEFAULT ''"
            )
        if "translation_direction" not in columns:
            conn.execute(
                f"ALTER TABLE translation_jobs ADD COLUMN translation_direction TEXT DEFAULT '{TRANSLATION_DIRECTION_JP_EN}'"
            )
        if "owner_session_id" not in columns:
            conn.execute(
                "ALTER TABLE translation_jobs ADD COLUMN owner_session_id TEXT DEFAULT ''"
            )
        for column_name in (
            "source_file_path",
            "result_file_path",
            "result_mime",
            "progress_message",
            "notify_email",
            "notification_status",
        ):
            if column_name not in columns:
                conn.execute(
                    f"ALTER TABLE translation_jobs ADD COLUMN {column_name} TEXT DEFAULT ''"
                )
        for column_name in ("generated_pages", "total_pages"):
            if column_name not in columns:
                conn.execute(
                    f"ALTER TABLE translation_jobs ADD COLUMN {column_name} INTEGER DEFAULT 0"
                )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_translation_jobs_owner_status "
            "ON translation_jobs(owner_session_id, status, updated_at)"
        )


def create_translation_job(
    file_name: str,
    file_size_bytes: int,
    total_blocks: int,
    translatable_blocks: int,
    total_batches: int,
    translation_mode: str = GENERAL_TRANSLATION_MODE,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
    owner_session_id: str = "",
    source_file_path: str = "",
    result_file_path: str = "",
    result_mime: str = "",
    notify_email: str = "",
    status: str = "running",
) -> str:
    init_job_store()
    job_id = f"job_{int(time.time())}_{hashlib.sha256(f'{file_name}:{file_size_bytes}:{time.time()}'.encode()).hexdigest()[:8]}"
    now = time.strftime("%Y-%m-%d %H:%M:%S")
    with sqlite3.connect(JOB_DB_PATH) as conn:
        conn.execute(
            """
            INSERT INTO translation_jobs (
                job_id, file_name, file_size_bytes, translation_mode, translation_direction, owner_session_id, status,
                source_file_path, result_file_path, result_mime, total_blocks,
                translatable_blocks, completed_blocks, total_batches,
                completed_batches, notify_email, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                job_id,
                file_name,
                file_size_bytes,
                translation_mode,
                translation_direction,
                owner_session_id,
                status,
                source_file_path,
                result_file_path,
                result_mime,
                total_blocks,
                translatable_blocks,
                0,
                total_batches,
                0,
                notify_email,
                now,
                now,
            ),
        )
    return job_id


def update_translation_job(job_id: str, **fields) -> None:
    if not job_id or not fields:
        return
    init_job_store()
    allowed_fields = {
        "status",
        "total_blocks",
        "translatable_blocks",
        "completed_blocks",
        "total_batches",
        "completed_batches",
        "input_tokens",
        "output_tokens",
        "total_tokens",
        "result_file_name",
        "source_file_path",
        "result_file_path",
        "result_mime",
        "notify_email",
        "notification_status",
        "progress_message",
        "generated_pages",
        "total_pages",
        "error_message",
        "finished_at",
    }
    updates = {key: value for key, value in fields.items() if key in allowed_fields}
    if not updates:
        return
    updates["updated_at"] = time.strftime("%Y-%m-%d %H:%M:%S")
    assignments = ", ".join(f"{key} = ?" for key in updates)
    values = list(updates.values()) + [job_id]
    with sqlite3.connect(JOB_DB_PATH) as conn:
        conn.execute(f"UPDATE translation_jobs SET {assignments} WHERE job_id = ?", values)


def recent_translation_jobs(owner_session_id: str, limit: int = 10) -> pd.DataFrame:
    init_job_store()
    with sqlite3.connect(JOB_DB_PATH) as conn:
        return pd.read_sql_query(
            """
            SELECT
                job_id AS "Job ID",
                file_name AS "File",
                translation_mode AS "Mode",
                translation_direction AS "Direction",
                status AS "Status",
                completed_blocks || '/' || translatable_blocks AS "Blocks",
                completed_batches || '/' || total_batches AS "Batches",
                result_file_name AS "Result",
                updated_at AS "Updated"
            FROM translation_jobs
            WHERE owner_session_id = ?
            ORDER BY created_at DESC
            LIMIT ?
            """,
            conn,
            params=(owner_session_id, limit),
        )


def recent_translation_job_details(owner_session_id: str, limit: int = 8) -> pd.DataFrame:
    init_job_store()
    with sqlite3.connect(JOB_DB_PATH) as conn:
        return pd.read_sql_query(
            """
            SELECT
                job_id,
                file_name,
                translation_mode,
                translation_direction,
                status,
                translatable_blocks,
                completed_blocks,
                total_batches,
                completed_batches,
                result_file_name,
                created_at,
                updated_at
            FROM translation_jobs
            WHERE owner_session_id = ?
            ORDER BY updated_at DESC, created_at DESC
            LIMIT ?
            """,
            conn,
            params=(owner_session_id, limit),
        )


def translation_job_detail(job_id: str, owner_session_id: str) -> pd.DataFrame:
    init_job_store()
    with sqlite3.connect(JOB_DB_PATH) as conn:
        return pd.read_sql_query(
            """
            SELECT
                job_id,
                file_name,
                file_size_bytes,
                translation_mode,
                translation_direction,
                owner_session_id,
                source_file_path,
                result_file_path,
                result_mime,
                notify_email,
                notification_status,
                progress_message,
                generated_pages,
                total_pages,
                status,
                total_blocks,
                translatable_blocks,
                completed_blocks,
                total_batches,
                completed_batches,
                input_tokens,
                output_tokens,
                total_tokens,
                result_file_name,
                error_message,
                created_at,
                updated_at,
                finished_at
            FROM translation_jobs
            WHERE job_id = ? AND owner_session_id = ?
            """,
            conn,
            params=(job_id, owner_session_id),
        )


def latest_running_translation_job_id(owner_session_id: str) -> str:
    init_job_store()
    with sqlite3.connect(JOB_DB_PATH) as conn:
        row = conn.execute(
            """
            SELECT job_id
            FROM translation_jobs
            WHERE owner_session_id = ?
              AND status IN ('pending', 'running')
            ORDER BY
                updated_at DESC,
                created_at DESC
            LIMIT 1
            """,
            (owner_session_id,),
        ).fetchone()
    return row[0] if row else ""


def active_translation_job_count(owner_session_id: str) -> int:
    init_job_store()
    with sqlite3.connect(JOB_DB_PATH) as conn:
        row = conn.execute(
            "SELECT COUNT(*) FROM translation_jobs WHERE owner_session_id = ? AND status IN ('pending', 'running')",
            (owner_session_id,),
        ).fetchone()
    return int(row[0] if row else 0)


def translation_job_is_active(job_id: str) -> bool:
    init_job_store()
    with sqlite3.connect(JOB_DB_PATH) as conn:
        row = conn.execute(
            "SELECT status FROM translation_jobs WHERE job_id = ?",
            (job_id,),
        ).fetchone()
    return bool(row and row[0] in {"pending", "running"})


def stop_translation_job(job_id: str, owner_session_id: str, reason: str = "Stopped by user.") -> None:
    detail = translation_job_detail(job_id, owner_session_id)
    if detail.empty:
        return
    update_translation_job(
        job_id,
        status="failed",
        error_message=f"{reason} Saved progress is preserved and can be resumed later.",
        progress_message=reason,
        finished_at=time.strftime("%Y-%m-%d %H:%M:%S"),
    )


def stop_all_active_translation_jobs(owner_session_id: str, reason: str = "Stopped by user.") -> int:
    init_job_store()
    now = time.strftime("%Y-%m-%d %H:%M:%S")
    message = f"{reason} Saved progress is preserved and can be resumed later."
    with sqlite3.connect(JOB_DB_PATH) as conn:
        cursor = conn.execute(
            """
            UPDATE translation_jobs
            SET status = 'failed',
                error_message = ?,
                progress_message = ?,
                finished_at = ?,
                updated_at = ?
            WHERE owner_session_id = ?
              AND status IN ('pending', 'running')
            """,
            (message, reason, now, now, owner_session_id),
        )
        return cursor.rowcount


def stop_active_translation_jobs_for_file(
    file_name: str,
    reason: str,
    owner_session_id: str,
    keep_job_id: str = "",
) -> None:
    init_job_store()
    now = time.strftime("%Y-%m-%d %H:%M:%S")
    with sqlite3.connect(JOB_DB_PATH) as conn:
        if keep_job_id:
            conn.execute(
                """
                UPDATE translation_jobs
                SET status = 'failed',
                    error_message = ?,
                    progress_message = 'Stopped.',
                    finished_at = ?,
                    updated_at = ?
                WHERE file_name = ?
                  AND owner_session_id = ?
                  AND job_id != ?
                  AND status IN ('pending', 'running')
                """,
                (reason, now, now, file_name, owner_session_id, keep_job_id),
            )
        else:
            conn.execute(
                """
                UPDATE translation_jobs
                SET status = 'failed',
                    error_message = ?,
                    progress_message = 'Stopped.',
                    finished_at = ?,
                    updated_at = ?
                WHERE file_name = ?
                  AND owner_session_id = ?
                  AND status IN ('pending', 'running')
                """,
                (reason, now, now, file_name, owner_session_id),
            )


def init_translation_memory() -> None:
    with sqlite3.connect(TRANSLATION_MEMORY_DB_PATH) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS translation_memory (
                source_key TEXT NOT NULL,
                source_text TEXT NOT NULL,
                translation_mode TEXT NOT NULL,
                translated_text TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY (source_key, translation_mode)
            )
            """
        )


def translation_memory_lookup(source_texts: list[str], translation_mode: str) -> dict[str, str]:
    if not source_texts:
        return {}
    init_translation_memory()
    keys = sorted({translation_memory_key(text) for text in source_texts if clean_text(text)})
    if not keys:
        return {}
    found = {}
    with sqlite3.connect(TRANSLATION_MEMORY_DB_PATH) as conn:
        for start in range(0, len(keys), 900):
            key_chunk = keys[start:start + 900]
            placeholders = ",".join("?" for _ in key_chunk)
            rows = conn.execute(
                f"""
                SELECT source_key, translated_text
                FROM translation_memory
                WHERE translation_mode = ? AND source_key IN ({placeholders})
                """,
                [translation_mode, *key_chunk],
            ).fetchall()
            found.update({str(key): str(value) for key, value in rows})
    return found


def save_translation_memory_pairs(pairs: list[tuple[str, str]], translation_mode: str) -> None:
    cleaned_pairs = [
        (translation_memory_key(source), clean_text(source), clean_text(translated))
        for source, translated in pairs
        if clean_text(source) and clean_text(translated)
    ]
    if not cleaned_pairs:
        return
    init_translation_memory()
    now = time.strftime("%Y-%m-%d %H:%M:%S")
    with sqlite3.connect(TRANSLATION_MEMORY_DB_PATH) as conn:
        conn.executemany(
            """
            INSERT INTO translation_memory (
                source_key, source_text, translation_mode, translated_text, updated_at
            )
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(source_key, translation_mode) DO UPDATE SET
                translated_text = excluded.translated_text,
                updated_at = excluded.updated_at
            """,
            [
                (source_key, source_text, translation_mode, translated_text, now)
                for source_key, source_text, translated_text in cleaned_pairs
            ],
        )


def hydrate_translation_memory_from_checkpoint(
    blocks: list[TextBlock],
    checkpoint_translations: dict[str, str],
    translation_mode: str,
) -> None:
    pairs = [
        (block.text, checkpoint_translations[block.location])
        for block in blocks
        if block.location in checkpoint_translations
    ]
    save_translation_memory_pairs(pairs, translation_mode)


def is_safe_glossary_term(jp: str) -> bool:
    jp = clean_text(jp)
    if len(jp) < 2:
        return False
    if not has_japanese_text(jp):
        return False
    if PROTECTED_PATTERN.fullmatch(jp):
        return False
    return True


def empty_terms_dataframe() -> pd.DataFrame:
    return pd.DataFrame(columns=["JP", "EN", "Note", "Category", "Owner"])


def xlsx_to_dataframe(raw: bytes, sheet_index: int = 0) -> pd.DataFrame:
    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

    def column_index(cell_ref: str) -> int:
        letters = re.sub(r"[^A-Z]", "", cell_ref.upper())
        index = 0
        for letter in letters:
            index = index * 26 + (ord(letter) - ord("A") + 1)
        return max(index - 1, 0)

    with ZipFile(io.BytesIO(raw)) as archive:
        names = archive.namelist()
        shared_strings = []

        if "xl/sharedStrings.xml" in names:
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in root.findall("a:si", ns):
                shared_strings.append("".join(text.text or "" for text in item.findall(".//a:t", ns)))

        sheet_names = sorted(
            name for name in names if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )
        if not sheet_names:
            raise ValueError("No worksheet found in Excel file.")

        sheet_name = sheet_names[min(sheet_index, len(sheet_names) - 1)]
        root = ET.fromstring(archive.read(sheet_name))
        rows = []

        for row in root.findall("a:sheetData/a:row", ns):
            indexed_values = {}
            for cell in row.findall("a:c", ns):
                if cell.attrib.get("t") == "inlineStr":
                    value = "".join(text.text or "" for text in cell.findall(".//a:t", ns))
                else:
                    value_node = cell.find("a:v", ns)
                    value = "" if value_node is None else value_node.text or ""
                    if cell.attrib.get("t") == "s" and value.isdigit():
                        value = shared_strings[int(value)]
                ref = cell.attrib.get("r", "")
                indexed_values[column_index(ref)] = value
            values = [indexed_values.get(index, "") for index in range(max(indexed_values.keys(), default=-1) + 1)]
            if any(str(value).strip() for value in values):
                rows.append(values)

    if not rows:
        return pd.DataFrame(columns=["JP", "EN"])

    header = [str(value).strip() for value in rows[0]]
    width = len(header)
    normalized_rows = [(row + [""] * width)[:width] for row in rows[1:]]
    return pd.DataFrame(normalized_rows, columns=header)


def read_glossary(uploaded_file) -> pd.DataFrame:
    if uploaded_file is None:
        glossary_path = next((path for path in DEFAULT_GLOSSARY_PATHS if path.exists()), None)
        if glossary_path is None:
            expected = ", ".join(path.name for path in DEFAULT_GLOSSARY_PATHS)
            raise FileNotFoundError(f"No glossary file found. Expected one of: {expected}")
        raw = glossary_path.read_bytes()
        name = glossary_path.name
    else:
        raw = uploaded_file.getvalue()
        name = uploaded_file.name

    is_excel = raw[:2] == b"PK" or name.lower().endswith((".xlsx", ".xlsm", ".xls"))
    if is_excel:
        try:
            return pd.read_excel(io.BytesIO(raw), sheet_name=0)
        except ImportError:
            return xlsx_to_dataframe(raw)
        except BadZipFile as exc:
            raise ValueError("The glossary Excel file could not be opened.") from exc

    for encoding in ("utf-8-sig", "utf-8", "cp932", "shift_jis", "cp1252"):
        try:
            return pd.read_csv(io.BytesIO(raw), encoding=encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Could not read glossary file. Please use CSV UTF-8 or Excel format.")


def read_rules_file(path: Path) -> pd.DataFrame:
    raw = path.read_bytes()
    is_excel = raw[:2] == b"PK" or path.name.lower().endswith((".xlsx", ".xlsm", ".xls"))
    if is_excel:
        try:
            return pd.read_excel(io.BytesIO(raw), sheet_name=0)
        except ImportError:
            return xlsx_to_dataframe(raw)
        except BadZipFile as exc:
            raise ValueError(f"The rule Excel file could not be opened: {path.name}") from exc

    for encoding in ("utf-8-sig", "utf-8", "cp932", "shift_jis", "cp1252"):
        try:
            return pd.read_csv(io.BytesIO(raw), encoding=encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Could not read rule file: {path.name}")


def read_plc_rules() -> pd.DataFrame:
    rule_path = next((path for path in DEFAULT_PLC_RULE_PATHS if path.exists()), None)
    if rule_path is None:
        return empty_terms_dataframe()
    return read_rules_file(rule_path)


def normalize_glossary(df: pd.DataFrame) -> pd.DataFrame:
    aliases = {
        "jp": "JP",
        "japanese": "JP",
        "japanese term": "JP",
        "æ—¥æœ¬èªž": "JP",
        "en": "EN",
        "english": "EN",
        "english term": "EN",
        "è‹±èªž": "EN",
        "note": "Note",
        "notes": "Note",
        "comment": "Note",
    }
    aliases.update(
        {
            "source japanese": "JP",
            "source": "JP",
            "japanese source": "JP",
            "japanese comment": "JP",
            "source term": "JP",
            "source text": "JP",
            "preferred english": "EN",
            "preferred abbreviation": "EN",
            "preferred en": "EN",
            "approved english": "EN",
            "approved abbreviation": "EN",
            "abbreviation": "EN",
            "standard english": "EN",
            "standard wording": "EN",
            "target": "EN",
            "target english": "EN",
            "do not use": "Note",
        }
    )

    renamed = {}
    for column in df.columns:
        key = str(column).strip().lower()
        renamed[column] = aliases.get(key, str(column).strip())

    glossary = df.rename(columns=renamed).fillna("")
    if "JP" not in glossary.columns or "EN" not in glossary.columns:
        raise ValueError("Glossary must include JP/Japanese and EN/English columns.")

    glossary = glossary.copy()
    glossary["JP"] = glossary["JP"].astype(str).map(clean_text)
    glossary["EN"] = glossary["EN"].astype(str).map(clean_text)
    glossary = glossary[(glossary["JP"] != "") & (glossary["EN"] != "")]
    glossary = glossary[glossary["JP"].map(is_safe_glossary_term)]
    glossary = glossary.drop_duplicates(subset=["JP"], keep="first")
    glossary["term_length"] = glossary["JP"].str.len()
    return glossary.sort_values("term_length", ascending=False).drop(columns=["term_length"]).reset_index(drop=True)


def normalize_plc_rules(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty and not set(df.columns).intersection({"JP", "EN", "Japanese", "English"}):
        return empty_terms_dataframe()
    try:
        rules = normalize_glossary(df)
    except Exception:
        return empty_terms_dataframe()
    if rules.empty:
        return empty_terms_dataframe()
    rules = rules.copy()
    rules["Category"] = "PLC/SPLC Rule"
    return rules


def glossary_for_mode(glossary: pd.DataFrame, plc_rules: pd.DataFrame, translation_mode: str) -> pd.DataFrame:
    if translation_mode != PLC_TRANSLATION_MODE or plc_rules.empty:
        return glossary

    combined = pd.concat([plc_rules, glossary], ignore_index=True, sort=False).fillna("")
    combined = combined.drop_duplicates(subset=["JP"], keep="first")
    combined["term_length"] = combined["JP"].str.len()
    return combined.sort_values("term_length", ascending=False).drop(columns=["term_length"]).reset_index(drop=True)


def controlled_term_pairs(
    glossary: pd.DataFrame,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> list[tuple[str, str, str, str]]:
    rows = []
    if translation_direction == TRANSLATION_DIRECTION_EN_JP:
        grouped: dict[str, list[tuple[str, str]]] = {}
        for _, row in glossary.iterrows():
            jp = clean_text(row["JP"])
            en = clean_text(row["EN"])
            if jp and en:
                grouped.setdefault(en.casefold(), []).append((jp, en))
        for values in grouped.values():
            unique_japanese = {jp for jp, _ in values}
            if len(unique_japanese) != 1:
                continue
            jp, en = values[0]
            rows.append((en, jp, jp, en))
    else:
        for _, row in glossary.iterrows():
            jp = clean_text(row["JP"])
            en = clean_text(row["EN"])
            if jp and en:
                rows.append((jp, en, jp, en))
    return sorted(rows, key=lambda item: len(item[0]), reverse=True)


def controlled_term_pattern(source_term: str, ignore_case: bool = False) -> re.Pattern:
    escaped = re.escape(source_term)
    if source_term and source_term[0].isascii() and source_term[0].isalnum() and source_term[-1].isascii() and source_term[-1].isalnum():
        escaped = rf"(?<![A-Za-z0-9]){escaped}(?![A-Za-z0-9])"
    return re.compile(escaped, re.IGNORECASE if ignore_case else 0)


def apply_glossary_to_source(
    text: str,
    glossary: pd.DataFrame,
    replace_source: bool = True,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> tuple[str, list[TermHit]]:
    translated_source = clean_text(text)
    hits = []

    for source_term, target_term, jp, en in controlled_term_pairs(glossary, translation_direction):
        pattern = controlled_term_pattern(
            source_term,
            ignore_case=translation_direction == TRANSLATION_DIRECTION_EN_JP,
        )
        count = len(pattern.findall(translated_source))
        if count:
            hits.append(TermHit(jp=jp, en=en, count=count))
            if replace_source:
                translated_source = pattern.sub(target_term, translated_source)

    return translated_source, hits


def exact_controlled_term_match(
    text: str,
    glossary: pd.DataFrame,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> tuple[str | None, list[TermHit]]:
    source = clean_text(text)
    if not source:
        return None, []

    for source_term, target_term, jp, en in controlled_term_pairs(glossary, translation_direction):
        matches = source.casefold() == source_term.casefold() if translation_direction == TRANSLATION_DIRECTION_EN_JP else source == source_term
        if matches:
            return target_term, [TermHit(jp=jp, en=en, count=1)]
    return None, []


def find_protected_codes(text: str) -> list[str]:
    return sorted(set(PROTECTED_PATTERN.findall(text)) | set(ENCLOSED_ALNUM_PATTERN.findall(text)))


def plc_mode_rules() -> str:
    return """
PLC/SPLC comment mode:
1. Treat the source as PLC/SPLC device comments, HMI labels, alarm labels, or control logic comments, not normal prose.
2. Meaning accuracy is the highest priority. First preserve the exact meaning, then make the wording concise.
3. Output short engineering labels only when doing so does not change the meaning. If compression would change the meaning, keep the translation slightly longer but accurate.
4. Even if the source is a full sentence or question, convert it into a concise PLC/HMI-style label only when the original intent is still preserved.
5. Do not convert terminology/wording questions into process-method labels. For example, wording like "冷やすって言い方" means cooling terms/expressions, not cooling methods.
6. Do not output conversational questions such as "How many..." unless the source explicitly must remain a user-facing question.
7. Do not string together multiple synonyms. Never output lists like "poor, defective, NG, inoperative".
8. Choose one stable plant-control term for each Japanese concept. Prefer concise PLC terms such as ON, OFF, OK, NG, Present, Absent, Complete, Confirm, Request, Command, Auto, Manual, Standby, Error.
9. Preserve PLC addresses, device IDs, robot names, station names, prefixes, symbols, arrows, brackets, and separators exactly.
10. Preserve enclosed/circled markers such as â“, â“‘, â’¸, and â‘  exactly. Do not change them to plain letters or numbers.
11. Keep repeated Japanese patterns translated with repeated English patterns.
12. If a company glossary term is provided, it overrides the default PLC wording.
""".strip()


def general_mode_rules() -> str:
    return """
General plant translation mode:
1. Translate into clear, natural plant-floor engineering English for manufacturing users.
2. Use concise plant-floor English suitable for controls, seibi, production, and engineering users.
3. Preserve line breaks and list structure when useful.
""".strip()


def supplier_email_mode_rules() -> str:
    return """
Supplier email translation mode:
1. Translate into natural, professional business English for a manufacturing technical email.
2. Keep names, company names, signal names, PLC addresses, HMI terms, ladder terms, and quoted alarm names accurate.
3. Use normal business email phrasing for standard Japanese email expressions:
   - inquiry-opening phrases should become "Regarding your inquiry"
   - relationship/polite opening phrases should become "Thank you for your continued support"
   - attachment-reference phrases should become "as shown in the attached document"
   - "also check/confirm together with" phrases should become "please also confirm" or "please confirm together with"
   - soft opinion phrases should become "it would likely be appropriate" or "I believe"
4. Do not over-translate polite Japanese. Keep the tone clear, respectful, and practical.
5. Prefer natural phrases such as "a specified period of time" for fixed-duration wording and "waiting for aging completion" for aging-completion-wait wording when context fits.
""".strip()


def product_catalog_mode_rules() -> str:
    return """
Product catalog translation mode:
1. Translate into concise, polished technical catalog English for product catalogs, specifications, case studies, and marketing-technical brochures.
2. Preserve product names, model names, belt types, pulley types, page references, catalog numbers, dimensions, symbols, registered marks, and part numbers exactly unless the source clearly requires translation.
3. Keep headings, table labels, captions, index entries, and callout labels short. Do not expand short catalog labels into long sentences.
4. Preserve line breaks and compact list/table structure where possible. Avoid merging unrelated headings, page references, and descriptions.
5. Use natural product English, but do not add marketing claims, features, applications, or recommendations that are not in the source.
6. If a glossary term has multiple English choices, choose one context-appropriate term. Never output multiple alternatives separated by commas or slashes.
7. Prefer concise phrases such as "Application", "Features", "Technical Data", "Dimension Tolerance", "Surface Roughness", and "Case Study" when context fits.
""".strip()


def powerpoint_mode_rules() -> str:
    return """
PowerPoint presentation translation mode:
1. Translate into concise, slide-ready manufacturing English for titles, bullets, tables, diagrams, and callouts.
2. Keep titles short and action-oriented. Do not expand bullets or labels into explanatory paragraphs.
3. Preserve paragraph boundaries, list hierarchy, numbers, units, model names, equipment IDs, symbols, and page/section references.
4. Translate each source block only; do not add speaker notes, commentary, background explanation, or new claims.
5. Prefer wording that fits the original text box while preserving the complete engineering meaning.
6. Before returning, verify that every explicit actor, condition, action, object, negation, number, unit, code, and required glossary term remains present. Never remove meaning merely to shorten the slide text.
""".strip()


def robot_program_mode_rules() -> str:
    return """
Robot program comment translation mode:
1. Treat the source as Kawasaki/industrial robot program comments or operator messages, not normal prose.
2. Translate only the Japanese comment meaning into concise engineering English.
3. Preserve robot commands, variables, positions, labels, numbers, symbols, punctuation, and code-style wording exactly when they appear.
4. Keep the translation short enough to fit inside a robot program comment when possible.
5. Do not add explanations, troubleshooting advice, or programming changes.
6. Use stable plant-floor terms such as Home Position, Workpiece, Clamp, Unclamp, Pick, Place, Start, Stop, Complete, Error, and Check when context fits.
""".strip()


def hmi_screen_mode_rules() -> str:
    return """
HMI screen translation mode:
1. Treat the source as short HMI labels, button text, parameter names, alarm labels, or screen navigation text.
2. Output concise engineering UI text that can fit inside the original HMI area.
3. Preserve numbers, units, axis names, PLC/HMI codes, model names, arrows, symbols, and separators exactly.
4. Prefer stable manufacturing UI terms such as Main Screen, Selection Screen, Register, Cancel, Speed, Limit, Spare, Home Return, Workpiece, Alarm History, Monitor Screen, and Takt Time.
5. If a Toyota/company glossary term is provided, it overrides the default HMI wording.
6. Do not add explanations, notes, or extra words that were not present in the source.
""".strip()


def japanese_target_mode_rules(translation_mode: str) -> str:
    if translation_mode == PLC_TRANSLATION_MODE:
        return """
PLC/SPLC comment mode (English to Japanese):
1. Translate into concise, standard Japanese controls terminology suitable for PLC/SPLC comments.
2. Preserve PLC addresses, device IDs, signal names, robot names, station IDs, prefixes, symbols, arrows, brackets, and separators exactly.
3. Prefer one stable Japanese control term for each concept; do not provide synonym lists or explanations.
4. Keep short labels short, while preserving the complete engineering meaning.
5. Approved glossary mappings override general wording when the reverse English mapping is unambiguous.
""".strip()
    if translation_mode == SUPPLIER_EMAIL_TRANSLATION_MODE:
        return """
Business email mode (English to Japanese):
1. Translate into natural, professional Japanese for a manufacturing technical email.
2. Preserve names, company names, signal names, PLC addresses, quoted alarms, and model numbers exactly.
3. Use respectful business Japanese without adding promises, causes, or technical details.
""".strip()
    if translation_mode == PRODUCT_CATALOG_TRANSLATION_MODE:
        return """
Catalog/specification mode (English to Japanese):
1. Translate into concise, polished Japanese suitable for technical catalogs and specifications.
2. Preserve model names, dimensions, symbols, page references, registered marks, catalog numbers, and part numbers exactly.
3. Keep headings, table labels, captions, and callouts compact; do not add marketing claims.
""".strip()
    if translation_mode == POWERPOINT_TRANSLATION_MODE:
        return """
PowerPoint presentation mode (English to Japanese):
1. Translate into concise, slide-ready manufacturing Japanese for titles, bullets, tables, diagrams, and callouts.
2. Keep titles, bullets, and labels compact; do not expand them into explanatory paragraphs.
3. Preserve paragraph boundaries, list hierarchy, numbers, units, model names, equipment IDs, symbols, and references.
4. Translate only the supplied slide text and do not add speaker notes, commentary, explanations, or new claims.
""".strip()
    if translation_mode == ROBOT_PROGRAM_TRANSLATION_MODE:
        return """
Robot program comment mode (English to Japanese):
1. Translate only eligible English comments or operator messages into concise engineering Japanese.
2. Preserve robot instructions, commands, variables, positions, labels, numbers, symbols, punctuation, and program structure exactly.
3. Do not add explanations, troubleshooting advice, or programming changes.
""".strip()
    if translation_mode == HMI_SCREEN_TRANSLATION_MODE:
        return """
HMI screen mode (English to Japanese):
1. Translate into concise Japanese UI wording that can fit the original HMI area.
2. Preserve numbers, units, axis names, PLC/HMI codes, model names, arrows, symbols, and separators exactly.
3. Do not add explanations or words not present in the source.
""".strip()
    return """
General plant translation mode (English to Japanese):
1. Translate into clear, natural Japanese for manufacturing, controls, production, and engineering users.
2. Preserve the complete technical meaning and useful line/list structure.
3. Keep equipment identifiers, signal names, part numbers, and measurements unchanged.
""".strip()


def mode_rules_for(
    translation_mode: str,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> str:
    if translation_direction == TRANSLATION_DIRECTION_EN_JP:
        return japanese_target_mode_rules(translation_mode)
    if translation_mode == PLC_TRANSLATION_MODE:
        return plc_mode_rules()
    if translation_mode == SUPPLIER_EMAIL_TRANSLATION_MODE:
        return supplier_email_mode_rules()
    if translation_mode == PRODUCT_CATALOG_TRANSLATION_MODE:
        return product_catalog_mode_rules()
    if translation_mode == POWERPOINT_TRANSLATION_MODE:
        return powerpoint_mode_rules()
    if translation_mode == ROBOT_PROGRAM_TRANSLATION_MODE:
        return robot_program_mode_rules()
    if translation_mode == HMI_SCREEN_TRANSLATION_MODE:
        return hmi_screen_mode_rules()
    return general_mode_rules()


def normalize_plc_translation_line(line: str) -> str:
    normalized = re.sub(r"\s+", " ", line).strip()
    for pattern, replacement in PLC_SYNONYM_CLEANUPS:
        normalized = pattern.sub(replacement, normalized)

    changed = True
    while changed:
        changed = False
        for word in PLC_DUPLICATE_STATUS_WORDS:
            pattern = re.compile(rf"\b({re.escape(word)})\s+\1\b", re.IGNORECASE)
            normalized, count = pattern.subn(r"\1", normalized)
            if count:
                changed = True

    return normalized


def post_process_translation(
    output_text: str,
    translation_mode: str,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> str:
    cleaned = output_text.strip()
    if translation_mode != PLC_TRANSLATION_MODE or translation_direction == TRANSLATION_DIRECTION_EN_JP:
        return cleaned
    return "\n".join(
        normalize_plc_translation_line(line)
        for line in cleaned.splitlines()
    ).strip()


def restore_missing_enclosed_markers(source_text: str, translated_text: str) -> str:
    source_markers = list(dict.fromkeys(ENCLOSED_ALNUM_PATTERN.findall(source_text)))
    if not source_markers:
        return translated_text

    restored = translated_text
    for marker in source_markers:
        if marker in restored:
            continue
        if source_text.startswith(marker):
            restored = marker + restored
        elif source_text.rstrip().endswith(marker):
            restored = restored.rstrip() + marker
        else:
            restored = restored.rstrip() + marker
    return restored


POWERPOINT_JP_EN_COVERAGE_RULES = [
    ("actor: 作業者", ("作業者",), r"\b(operator|worker|personnel|technician)\b"),
    ("condition: 場合", ("場合",), r"\b(if|when|whenever|in case)\b"),
    ("condition: 前に", ("前に",), r"\b(before|prior to)\b"),
    ("condition: 後", ("後",), r"\b(after|following|once|before)\b"),
    ("condition: まで", ("まで",), r"\b(until|before|by the time|up to)\b"),
    ("action: 確認", ("確認",), r"\b(confirm|check|verify|ensure)\w*\b"),
    ("action: 起動/開始", ("起動", "開始"), r"\b(restart|start|startup|launch|activate)\w*\b"),
    ("action: 停止", ("停止",), r"\b(stop|stoppage|halt|shutdown|shut down)\w*\b"),
    ("action: 連絡", ("連絡",), r"\b(contact|notify|inform|report)\w*\b"),
    ("action: 押す", ("押し", "押さ", "押す", "押された"), r"\b(press|pressed|push)\w*\b"),
    ("action: 開く", ("開いて", "開く", "開け"), r"\b(open|opened|opening)\b"),
    ("action: 閉じる", ("閉じ", "閉め"), r"\b(close|closed|closing|shut)\b"),
    ("action: 翻訳", ("翻訳",), r"\b(translate|translated|translating|translation)\b"),
]


def powerpoint_translation_quality_issues(
    source_text: str,
    translated_text: str,
    hits: list[TermHit] | None = None,
) -> list[str]:
    source = clean_text(source_text)
    target = clean_text(translated_text)
    target_folded = target.casefold()
    issues = []

    for label, source_markers, target_pattern in POWERPOINT_JP_EN_COVERAGE_RULES:
        if any(marker in source for marker in source_markers) and not re.search(target_pattern, target, re.IGNORECASE):
            issues.append(f"Missing {label}")

    if re.search(r"(?:ない|ません|禁止|不可|無効)", source) and not re.search(
        r"\b(no|not|never|without|prohibit|forbid|disable|disabled|invalid|unavailable|cannot|can't|do not|don't)\b",
        target,
        re.IGNORECASE,
    ):
        issues.append("Missing negation or prohibition")

    for code in find_protected_codes(source_text):
        if code not in translated_text:
            issues.append(f"Missing protected code: {code}")

    for number in re.findall(r"(?<![A-Za-z])\d+(?:\.\d+)?(?:%|℃|°C|mm|cm|m|kg|V|A|Hz|s|min)?", source):
        if number and number not in target:
            issues.append(f"Missing number or unit: {number}")

    for hit in hits or []:
        if hit.en and hit.en.casefold() not in target_folded:
            issues.append(f"Missing required glossary term: {hit.jp} → {hit.en}")

    source_lines = [line for line in str(source_text).splitlines() if line.strip()]
    target_lines = [line for line in str(translated_text).splitlines() if line.strip()]
    if len(source_lines) > 1 and all(line.lstrip().startswith(("・", "•", "-", "*")) for line in source_lines):
        if len(target_lines) != len(source_lines):
            issues.append("List item count changed")

    return list(dict.fromkeys(issues))


INSTRUCTION_LINE_PATTERNS = [
    re.compile(r"^\s*å·¥å ´ã§åƒãäºº(?:ã®ãŸã‚|å‘ã‘)?ã«è¨³ã—ã¦[ã€‚ï½¡.!ï¼]?\s*$"),
    re.compile(r"^\s*ç¾å ´(?:ã®äºº|ä½œæ¥­è€…)?(?:ã®ãŸã‚|å‘ã‘)?ã«è¨³ã—ã¦[ã€‚ï½¡.!ï¼]?\s*$"),
    re.compile(r"^\s*è£½é€ ç¾å ´(?:ã®äºº|ä½œæ¥­è€…)?(?:ã®ãŸã‚|å‘ã‘)?ã«è¨³ã—ã¦[ã€‚ï½¡.!ï¼]?\s*$"),
]
INSTRUCTION_SUFFIX_PATTERNS = [
    re.compile(r"\s*å·¥å ´ã§åƒãäºº(?:ã®ãŸã‚|å‘ã‘)?ã«è¨³ã—ã¦[ã€‚ï½¡.!ï¼]?\s*$"),
    re.compile(r"\s*ç¾å ´(?:ã®äºº|ä½œæ¥­è€…)?(?:ã®ãŸã‚|å‘ã‘)?ã«è¨³ã—ã¦[ã€‚ï½¡.!ï¼]?\s*$"),
    re.compile(r"\s*è£½é€ ç¾å ´(?:ã®äºº|ä½œæ¥­è€…)?(?:ã®ãŸã‚|å‘ã‘)?ã«è¨³ã—ã¦[ã€‚ï½¡.!ï¼]?\s*$"),
]


def split_text_translation_input(text: str) -> tuple[str, str]:
    source_lines = []
    guidance = []
    for line in str(text).splitlines():
        stripped = line.strip()
        if any(pattern.match(stripped) for pattern in INSTRUCTION_LINE_PATTERNS):
            guidance.append("Translate for people working on the manufacturing floor.")
            continue
        for pattern in INSTRUCTION_SUFFIX_PATTERNS:
            line, count = pattern.subn("", line)
            if count:
                guidance.append("Translate for people working on the manufacturing floor.")
                break
        source_lines.append(line)

    source_text = "\n".join(source_lines).strip()
    return source_text, " ".join(dict.fromkeys(guidance))


def build_prompt(
    source_text: str,
    hits: list[TermHit],
    protected_codes: list[str],
    translation_mode: str,
    user_guidance: str = "",
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> str:
    terms = "\n".join(f"{hit.jp} = {hit.en}" for hit in hits)
    codes = ", ".join(protected_codes) if protected_codes else "None detected"
    mode_rules = mode_rules_for(translation_mode, translation_direction)
    guidance = user_guidance.strip() or "None"
    source_language, target_language = direction_language_names(translation_direction)

    return f"""
You are a professional {source_language}-to-{target_language} translator for a battery manufacturing plant.

Translation mode: {translation_mode}
Translation direction: {translation_direction}

{mode_rules}

Mandatory rules:
1. Use approved company glossary terms when they are precise technical terms.
2. Apply glossary mappings in the selected direction only when the mapping is precise and unambiguous. Choose one context-appropriate target term; never output a list of alternatives.
3. Do not force a glossary term when it would mistranslate a normal business phrase. For example, standard inquiry-opening phrases should become "Regarding your inquiry".
4. Preserve PLC addresses, device IDs, model names, station IDs, alarm codes, part numbers, and equipment codes exactly.
5. Do not invent missing information, causes, actions, measurements, or context that is not in the source.
6. If the source is ambiguous, translate only the meaning that is present and keep the wording neutral.
7. Output only the {target_language} translation. Do not add explanations, notes, or commentary.
8. Preserve text already written in {target_language}; translate only the {source_language} content.

Company terminology detected in the source:
{terms if terms else "None"}

Protected codes detected:
{codes}

Additional translation guidance:
{guidance}

Source {source_language} text:
{source_text}
""".strip()


def openai_client() -> OpenAI:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError(
            "OPENAI_API_KEY was not found. Add it to app.env for local development "
            "or configure it as a secret/environment variable in the deployment environment."
        )
    base_url = os.getenv("OPENAI_BASE_URL")
    verify_ssl = os.getenv("OPENAI_SSL_VERIFY", "true").strip().lower() not in {"0", "false", "no"}
    ssl_verify: bool | ssl.SSLContext = verify_ssl
    if verify_ssl and truststore is not None:
        # Company VPN/proxy certificates are normally installed in the Windows
        # trust store rather than certifi's bundled CA file.
        ssl_verify = truststore.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
    http_client = httpx.Client(http2=False, timeout=60, verify=ssl_verify, trust_env=True)
    if base_url:
        client = OpenAI(api_key=api_key, base_url=base_url, http_client=http_client)
    else:
        client = OpenAI(api_key=api_key, http_client=http_client)

    tracing_enabled = os.getenv("LANGSMITH_TRACING", "").lower() == "force"
    if tracing_enabled and wrap_openai is not None:
        return wrap_openai(client)
    return client


def openai_model() -> str:
    return os.getenv("OPENAI_MODEL", DEFAULT_MODEL)


def openai_timeout_seconds() -> float:
    try:
        return float(os.getenv("OPENAI_TIMEOUT_SECONDS", OPENAI_TIMEOUT_SECONDS))
    except ValueError:
        return float(OPENAI_TIMEOUT_SECONDS)


def azure_translator_configured() -> bool:
    return bool(os.getenv("AZURE_TRANSLATOR_KEY"))


def azure_translate_texts(
    texts: list[str],
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> list[str]:
    key = os.getenv("AZURE_TRANSLATOR_KEY", "")
    endpoint = os.getenv("AZURE_TRANSLATOR_ENDPOINT", "https://api.cognitive.microsofttranslator.com")
    region = os.getenv("AZURE_TRANSLATOR_REGION", "")
    if not key:
        raise RuntimeError("Azure Translator is not configured.")

    source_code, target_code = ("en", "ja") if translation_direction == TRANSLATION_DIRECTION_EN_JP else ("ja", "en")
    route = "/translate?" + urllib.parse.urlencode({"api-version": "3.0", "from": source_code, "to": target_code})
    request = urllib.request.Request(
        endpoint.rstrip("/") + route,
        data=json.dumps([{"Text": text} for text in texts], ensure_ascii=False).encode("utf-8"),
        headers={
            "Ocp-Apim-Subscription-Key": key,
            "Ocp-Apim-Subscription-Region": region,
            "Content-Type": "application/json",
            "X-ClientTraceId": str(uuid.uuid4()),
        },
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=openai_timeout_seconds()) as response:
        payload = json.loads(response.read().decode("utf-8"))
    return [item["translations"][0]["text"] for item in payload]


def machine_translate_texts(
    texts: list[str],
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> list[str] | None:
    if not azure_translator_configured():
        return None
    translated = []
    for start in range(0, len(texts), 100):
        translated.extend(azure_translate_texts(texts[start:start + 100], translation_direction))
    return translated


def max_parallel_batches() -> int:
    try:
        value = int(os.getenv("MAX_PARALLEL_BATCHES", MAX_PARALLEL_BATCHES))
    except ValueError:
        value = MAX_PARALLEL_BATCHES
    return max(value, 1)


def response_token_usage(response) -> TokenUsage:
    usage = getattr(response, "usage", None)
    if usage is None:
        return TokenUsage()

    def usage_value(*names: str) -> int:
        for name in names:
            value = getattr(usage, name, None)
            if value is not None:
                return int(value)
        if isinstance(usage, dict):
            for name in names:
                value = usage.get(name)
                if value is not None:
                    return int(value)
        return 0

    input_tokens = usage_value("input_tokens", "prompt_tokens")
    output_tokens = usage_value("output_tokens", "completion_tokens")
    total_tokens = usage_value("total_tokens")
    if total_tokens == 0:
        total_tokens = input_tokens + output_tokens
    return TokenUsage(
        input_tokens=input_tokens,
        output_tokens=output_tokens,
        total_tokens=total_tokens,
    )


def exception_chain_text(exc: BaseException) -> str:
    details: list[str] = []
    current: BaseException | None = exc
    seen: set[int] = set()
    while current is not None and id(current) not in seen:
        seen.add(id(current))
        details.append(f"{type(current).__name__}: {current}")
        current = current.__cause__ or current.__context__
    return " | ".join(details)


def format_translation_error(exc: Exception) -> str:
    if isinstance(exc, APIConnectionError):
        detail = exception_chain_text(exc)
        if "CERTIFICATE_VERIFY_FAILED" in detail or "certificate verify failed" in detail.lower():
            return (
                "SSL certificate verification failed while calling the OpenAI API. "
                "Install truststore in the app virtual environment and restart the server so Python uses "
                "the Windows trusted certificate store."
            )
        if "timed out" in detail.lower() or "timeout" in detail.lower():
            return (
                "The GPT API connection timed out. Confirm the server VPN/proxy path and retry. "
                "If this repeats, check the server network logs and OPENAI_TIMEOUT_SECONDS."
            )
        return (
            "Connection error while calling the GPT API. Check whether OPENAI_BASE_URL is required "
            "for the company API, and confirm VPN/proxy/firewall access from this computer. "
            f"Technical detail: {detail[:500]}"
        )
    if isinstance(exc, AuthenticationError):
        return "Authentication failed. Check that OPENAI_API_KEY is correct and approved for this endpoint."
    if isinstance(exc, RateLimitError):
        return "Rate limit or quota exceeded. Check the API quota, rate limit, or billing/usage policy."
    if isinstance(exc, APIStatusError):
        return f"GPT API returned HTTP {exc.status_code}. {exc.message}"
    return str(exc)


def translate_text(
    source_text: str,
    hits: list[TermHit],
    protected_codes: list[str],
    translation_mode: str,
    user_guidance: str = "",
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> tuple[str, TokenUsage]:
    client = openai_client()
    response = client.responses.create(
        model=openai_model(),
        input=build_prompt(
            source_text,
            hits,
            protected_codes,
            translation_mode,
            user_guidance,
            translation_direction,
        ),
        temperature=0,
        timeout=openai_timeout_seconds(),
    )
    return post_process_translation(response.output_text, translation_mode, translation_direction), response_token_usage(response)


def translate_block(
    text: str,
    glossary: pd.DataFrame,
    translation_mode: str,
    user_guidance: str = "",
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> tuple[str, list[TermHit], TokenUsage]:
    exact_translation, exact_hits = exact_controlled_term_match(text, glossary, translation_direction)
    if exact_translation is not None:
        translation = post_process_translation(exact_translation, translation_mode, translation_direction)
        return restore_missing_enclosed_markers(text, translation), exact_hits, TokenUsage()

    glossary_applied_text, hits = apply_glossary_to_source(
        text,
        glossary,
        replace_source=translation_mode == PLC_TRANSLATION_MODE,
        translation_direction=translation_direction,
    )
    protected_codes = find_protected_codes(text)
    translation, token_usage = translate_text(
        glossary_applied_text,
        hits,
        protected_codes,
        translation_mode,
        user_guidance,
        translation_direction,
    )
    return restore_missing_enclosed_markers(text, translation), hits, token_usage


def build_batch_prompt(
    items: list[tuple[int, str, list[TermHit], list[str], str]],
    translation_mode: str,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> str:
    item_sections = []
    for item_id, source_text, _, _, context in items:
        lines = []
        if context and clean_text(context) != clean_text(source_text):
            lines.extend(
                [
                    f"[CONTEXT {item_id}]",
                    context,
                    f"[/CONTEXT {item_id}]",
                ]
            )
        lines.extend(
            [
                f"[BLOCK {item_id}]",
                source_text,
                f"[/BLOCK {item_id}]",
            ]
        )
        item_sections.append("\n".join(lines))
    item_text = "\n\n".join(item_sections)
    terms = []
    codes = []
    for _, _, hits, protected_codes, _ in items:
        terms.extend(f"{hit.jp} = {hit.en}" for hit in hits)
        codes.extend(protected_codes)

    unique_terms = "\n".join(sorted(set(terms))) or "None"
    unique_codes = ", ".join(sorted(set(codes))) if codes else "None detected"
    mode_rules = mode_rules_for(translation_mode, translation_direction)
    source_language, target_language = direction_language_names(translation_direction)

    return f"""
You are a professional {source_language}-to-{target_language} translator for a battery manufacturing plant.

Translation mode: {translation_mode}
Translation direction: {translation_direction}

{mode_rules}

Mandatory rules:
1. Use approved company glossary terms when they are precise technical terms.
2. Apply glossary mappings in the selected direction only when the mapping is precise and unambiguous. Choose one context-appropriate target term; never output a list of alternatives.
3. Do not force a glossary term when it would mistranslate a normal business phrase. For example, standard inquiry-opening phrases should become "Regarding your inquiry".
4. Preserve PLC addresses, device IDs, model names, station IDs, alarm codes, part numbers, and equipment codes exactly.
5. Do not invent missing information, causes, actions, measurements, or context that is not in the source.
6. If the source is ambiguous, translate only the meaning that is present and keep the wording neutral.
7. Return each translated block using the same markers and do not add explanations, notes, or commentary:
[BLOCK 1]
{target_language} translation
[/BLOCK 1]
8. Preserve text already written in {target_language}; translate only the {source_language} content.
9. For PowerPoint content, each BLOCK contains the complete text from one text box. Read the entire BLOCK first and translate its complete meaning as one semantic unit; line or paragraph breaks inside it may be visual formatting rather than separate meanings.
10. Do not translate PowerPoint text line by line. Produce one natural, accurate translation for continuous prose. Preserve separate list items only when the source is clearly a list.
11. Keep every BLOCK marker separate because each BLOCK maps to one source text box.

Company terminology detected:
{unique_terms}

Protected codes detected:
{unique_codes}

Source blocks:
{item_text}
""".strip()


def parse_batch_translation(output_text: str, item_ids: list[int]) -> dict[int, str]:
    translations = {}
    for item_id in item_ids:
        pattern = re.compile(
            rf"\[BLOCK {item_id}\]\s*(.*?)\s*\[/BLOCK {item_id}\]",
            re.DOTALL,
        )
        match = pattern.search(output_text)
        if match:
            translations[item_id] = match.group(1).strip()
    return translations


def batch_powerpoint_quality_issues(
    chunk: list[TextBlock],
    items: list[tuple[int, str, list[TermHit], list[str], str]],
    parsed: dict[int, str],
    translation_mode: str,
    translation_direction: str,
) -> dict[int, list[str]]:
    if translation_mode != POWERPOINT_TRANSLATION_MODE or translation_direction != TRANSLATION_DIRECTION_JP_EN:
        return {}
    hits_by_id = {item_id: hits for item_id, _source, hits, _codes, _context in items}
    issues_by_id = {}
    for item_id, _source, _hits, _codes, _context in items:
        if item_id not in parsed:
            continue
        block = chunk[item_id - 1]
        issues = powerpoint_translation_quality_issues(
            block.text,
            parsed[item_id],
            hits_by_id.get(item_id, []),
        )
        if issues:
            issues_by_id[item_id] = issues
    return issues_by_id


def translation_quality_retry_instructions(issues_by_id: dict[int, list[str]]) -> str:
    details = "\n".join(
        f"- BLOCK {item_id}: {'; '.join(issues)}"
        for item_id, issues in issues_by_id.items()
    )
    return f"""

QUALITY CORRECTION REQUIRED:
The prior candidate failed semantic coverage checks:
{details}

Translate the affected BLOCKS again. Preserve every actor, condition, action, object, negation,
number, unit, protected code, and required glossary term. Concise wording is allowed only after
all source meaning is retained. Return the complete marked batch again.
""".rstrip()


def translate_batch_chunk(
    chunk: list[TextBlock],
    glossary: pd.DataFrame,
    translation_mode: str,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> tuple[dict[str, str], list[TermHit], TokenUsage]:
    items = []
    chunk_hits = []
    direct_translations = {}

    for offset, block in enumerate(chunk, start=1):
        exact_translation, exact_hits = exact_controlled_term_match(block.text, glossary, translation_direction)
        if exact_translation is not None:
            translation = post_process_translation(exact_translation, translation_mode, translation_direction)
            direct_translations[block.location] = restore_missing_enclosed_markers(block.text, translation)
            chunk_hits.extend(exact_hits)
            continue

        glossary_applied_text, hits = apply_glossary_to_source(
            block.text,
            glossary,
            replace_source=translation_mode == PLC_TRANSLATION_MODE,
            translation_direction=translation_direction,
        )
        protected_codes = find_protected_codes(block.text)
        items.append((offset, glossary_applied_text, hits, protected_codes, block.context))
        chunk_hits.extend(hits)

    parsed = {}
    token_usage = TokenUsage()
    last_error = None
    if items:
        machine_translations = None
        quality_retry = ""
        if translation_direction == TRANSLATION_DIRECTION_JP_EN and not any(item[4] for item in items):
            try:
                machine_translations = machine_translate_texts([item[1] for item in items], translation_direction)
            except Exception as exc:
                last_error = exc

        if machine_translations is not None:
            parsed = {
                item[0]: post_process_translation(translated, translation_mode, translation_direction)
                for item, translated in zip(items, machine_translations)
            }
            machine_quality_issues = batch_powerpoint_quality_issues(
                chunk,
                items,
                parsed,
                translation_mode,
                translation_direction,
            )
            if machine_quality_issues:
                quality_retry = translation_quality_retry_instructions(machine_quality_issues)
                parsed = {}
                machine_translations = None
        if machine_translations is None:
            client = openai_client()
            for attempt in range(1, MAX_TRANSLATION_RETRIES + 1):
                try:
                    prompt = build_batch_prompt(items, translation_mode, translation_direction) + quality_retry
                    response = client.responses.create(
                        model=openai_model(),
                        input=prompt,
                        temperature=0,
                        timeout=openai_timeout_seconds(),
                    )
                    token_usage.add(response_token_usage(response))
                    parsed = parse_batch_translation(response.output_text.strip(), [item[0] for item in items])
                    missing_ids = [item[0] for item in items if item[0] not in parsed]
                    if missing_ids:
                        raise ValueError(f"Translation response missed block marker(s): {missing_ids}")
                    quality_issues = batch_powerpoint_quality_issues(
                        chunk,
                        items,
                        parsed,
                        translation_mode,
                        translation_direction,
                    )
                    if quality_issues:
                        quality_retry = translation_quality_retry_instructions(quality_issues)
                        if attempt == MAX_TRANSLATION_RETRIES:
                            raise ValueError(
                                "PowerPoint semantic quality validation failed: "
                                + " | ".join(
                                    f"BLOCK {item_id}: {', '.join(issues)}"
                                    for item_id, issues in quality_issues.items()
                                )
                            )
                        continue
                    break
                except Exception as exc:
                    last_error = exc
                    if attempt == MAX_TRANSLATION_RETRIES:
                        raise
                    time.sleep(5 * attempt)

    if items and not parsed and last_error:
        raise last_error

    chunk_translations = dict(direct_translations)
    for offset, block in enumerate(chunk, start=1):
        if block.location in chunk_translations:
            continue
        translation = post_process_translation(parsed[offset], translation_mode, translation_direction)
        chunk_translations[block.location] = restore_missing_enclosed_markers(block.text, translation)

    return chunk_translations, chunk_hits, token_usage


def translate_batch_chunk_resilient(
    chunk: list[TextBlock],
    glossary: pd.DataFrame,
    translation_mode: str,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> tuple[dict[str, str], list[TermHit], TokenUsage]:
    try:
        return translate_batch_chunk(chunk, glossary, translation_mode, translation_direction)
    except Exception:
        if len(chunk) <= 1:
            raise

    midpoint = max(len(chunk) // 2, 1)
    combined_translations = {}
    combined_hits = []
    combined_usage = TokenUsage()
    for part in (chunk[:midpoint], chunk[midpoint:]):
        part_translations, part_hits, part_usage = translate_batch_chunk_resilient(
            part,
            glossary,
            translation_mode,
            translation_direction,
        )
        combined_translations.update(part_translations)
        combined_hits.extend(part_hits)
        combined_usage.add(part_usage)
    return combined_translations, combined_hits, combined_usage


def translation_memory_key(text: str) -> str:
    return clean_text(text)


def block_has_translation_context(block: TextBlock) -> bool:
    return bool(block.context and clean_text(block.context) != clean_text(block.text))


def block_translation_key(block: TextBlock) -> str:
    if block_has_translation_context(block):
        return f"context:{block.location}"
    return translation_memory_key(block.text)


def document_translation_chunks(blocks: list[TextBlock]) -> list[list[TextBlock]]:
    units = []
    current_unit = []
    current_group = ""
    for block in blocks:
        group = block.context_group if block_has_translation_context(block) else ""
        if current_unit and (not group or group != current_group):
            units.append(current_unit)
            current_unit = []
        current_unit.append(block)
        current_group = group
        if not group:
            units.append(current_unit)
            current_unit = []
            current_group = ""
    if current_unit:
        units.append(current_unit)

    chunks = []
    chunk = []
    for unit in units:
        if chunk and len(chunk) + len(unit) > DOCUMENT_BATCH_SIZE:
            chunks.append(chunk)
            chunk = []
        chunk.extend(unit)
        if len(chunk) >= DOCUMENT_BATCH_SIZE:
            chunks.append(chunk)
            chunk = []
    if chunk:
        chunks.append(chunk)
    return chunks


def translate_blocks_batch(
    blocks: list[TextBlock],
    glossary: pd.DataFrame,
    translation_mode: str,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
    checkpoint_path=None,
    progress_callback=None,
    should_continue=None,
) -> tuple[dict[str, str], list[TermHit], TokenUsage]:
    def ensure_continue() -> None:
        if should_continue is not None and not should_continue():
            raise RuntimeError("Translation job was stopped.")

    translations = load_checkpoint(checkpoint_path) if checkpoint_path else {}
    translatable_blocks = [block for block in blocks if should_translate(block.text, translation_direction)]
    memory_mode = translation_cache_mode(translation_mode, translation_direction)
    hydrate_translation_memory_from_checkpoint(blocks, translations, memory_mode)
    source_by_key = {}
    for block in translatable_blocks:
        if not block_has_translation_context(block):
            source_by_key.setdefault(translation_memory_key(block.text), block.text)
    memory_hits = translation_memory_lookup(list(source_by_key.values()), memory_mode)
    memory_applied = 0
    for block in translatable_blocks:
        if block.location in translations:
            continue
        if block_has_translation_context(block):
            continue
        memory_translation = memory_hits.get(translation_memory_key(block.text))
        if memory_translation:
            translations[block.location] = memory_translation
            memory_applied += 1
    if memory_applied and checkpoint_path:
        save_checkpoint(checkpoint_path, translations)
    source_memory = {}
    for block in translatable_blocks:
        if block.location in translations:
            source_memory.setdefault(block_translation_key(block), translations[block.location])

    for block in translatable_blocks:
        key = block_translation_key(block)
        if block.location not in translations and key in source_memory:
            translations[block.location] = source_memory[key]

    pending_by_key = {}
    duplicate_locations_by_key = {}
    for block in translatable_blocks:
        if block.location in translations:
            continue
        key = block_translation_key(block)
        if key not in pending_by_key:
            pending_by_key[key] = block
            duplicate_locations_by_key[key] = []
        duplicate_locations_by_key[key].append(block.location)

    pending_blocks = list(pending_by_key.values())
    chunks = document_translation_chunks(pending_blocks)
    all_hits = []
    token_usage = TokenUsage()
    started_at = time.time()
    pending_location_count = sum(len(locations) for locations in duplicate_locations_by_key.values())
    completed_at_start = len(translatable_blocks) - pending_location_count
    total_batches = len(chunks)
    parallel_batches = min(max_parallel_batches(), max(total_batches, 1))

    ensure_continue()
    if progress_callback:
        progress_callback(
            completed_at_start,
            len(translatable_blocks),
            0,
            total_batches,
            0,
            "Resuming translation." if completed_at_start else "Starting translation.",
        )

    completed_batches = 0
    completed_blocks = completed_at_start

    if chunks:
        ensure_continue()
        if progress_callback:
            progress_callback(
                completed_blocks,
                len(translatable_blocks),
                0,
                total_batches,
                time.time() - started_at,
                f"Translating with {parallel_batches} parallel workers.",
            )
        with ThreadPoolExecutor(max_workers=parallel_batches) as executor:
            chunk_index = 0
            future_to_chunk = {}

            def submit_until_full() -> None:
                nonlocal chunk_index
                while len(future_to_chunk) < parallel_batches and chunk_index < len(chunks):
                    ensure_continue()
                    chunk = chunks[chunk_index]
                    chunk_index += 1
                    future_to_chunk[
                        executor.submit(
                            translate_batch_chunk_resilient,
                            chunk,
                            glossary,
                            translation_mode,
                            translation_direction,
                        )
                    ] = chunk

            submit_until_full()
            while future_to_chunk:
                done_futures, _ = wait(future_to_chunk, return_when=FIRST_COMPLETED)
                for future in done_futures:
                    chunk = future_to_chunk.pop(future)
                    ensure_continue()
                    chunk_translations, chunk_hits, chunk_token_usage = future.result()
                    expanded_translations = {}
                    for block in chunk:
                        key = block_translation_key(block)
                        translated_text = chunk_translations[block.location]
                        source_memory[key] = translated_text
                        for location in duplicate_locations_by_key.get(key, [block.location]):
                            expanded_translations[location] = translated_text

                    translations.update(expanded_translations)
                    save_translation_memory_pairs(
                        [
                            (block.text, chunk_translations[block.location])
                            for block in chunk
                            if block.location in chunk_translations
                            and not block_has_translation_context(block)
                        ],
                        memory_mode,
                    )
                    all_hits.extend(chunk_hits)
                    token_usage.add(chunk_token_usage)
                    completed_batches += 1
                    completed_blocks += len(expanded_translations)

                    if checkpoint_path:
                        save_checkpoint(checkpoint_path, translations)

                    if progress_callback:
                        progress_callback(
                            min(completed_blocks, len(translatable_blocks)),
                            len(translatable_blocks),
                            completed_batches,
                            total_batches,
                            time.time() - started_at,
                            "Translating",
                        )
                submit_until_full()

    return translations, all_hits, token_usage


def read_text_file(raw: bytes) -> str:
    return decode_document_text(raw)


def output_translation_for(
    location: str,
    source_text: str,
    translations: dict[str, str],
    keep_source_with_translation: bool = False,
) -> str:
    translated = translations.get(location, source_text)
    translated = restore_missing_enclosed_markers(source_text, translated)
    if not keep_source_with_translation or clean_text(translated) == clean_text(source_text):
        return translated
    return source_with_translation_lines(source_text, translated)


def source_with_translation_lines(source_text: str, translated_text: str) -> str:
    source_lines = str(source_text).splitlines()
    translated_lines = str(translated_text).splitlines()
    if len(source_lines) <= 1 or len(translated_lines) <= 1:
        return f"{source_text}\n{translated_text}"

    paired_lines = []
    max_lines = max(len(source_lines), len(translated_lines))
    for index in range(max_lines):
        if index < len(source_lines) and source_lines[index].strip():
            paired_lines.append(source_lines[index])
        if index < len(translated_lines) and translated_lines[index].strip():
            paired_lines.append(translated_lines[index])
    return "\n".join(paired_lines)


def write_text_file(
    blocks: list[TextBlock],
    translations: dict[str, str],
    keep_source_with_translation: bool = False,
) -> bytes:
    lines = []
    for block in blocks:
        lines.append(output_translation_for(block.location, block.text, translations, keep_source_with_translation))
    return "\n\n".join(lines).encode("utf-8-sig")


def robot_comment_segments(text: str) -> list[tuple[int, int, str]]:
    segments = []
    line_start = 0
    paired_comment_pattern = re.compile(r";([^;\r\n]*[\u3040-\u30ff\u3400-\u9fff][^;\r\n]*);")
    for line in text.splitlines(keepends=True):
        content = line.rstrip("\r\n")
        paired_matches = list(paired_comment_pattern.finditer(content))
        if paired_matches:
            for match in paired_matches:
                value = match.group(1).strip()
                if looks_like_mojibake(value):
                    continue
                segments.append((
                    line_start + match.start(1),
                    line_start + match.end(1),
                    value,
                ))
        else:
            comment_start = content.find(";")
            if comment_start >= 0:
                comment_text = content[comment_start + 1:]
                if has_japanese_text(comment_text) and not looks_like_mojibake(comment_text):
                    segments.append((
                        line_start + comment_start + 1,
                        line_start + len(content),
                        comment_text.strip(),
                    ))
        line_start += len(line)
    return segments


def robot_quoted_string_segments(text: str) -> list[tuple[int, int, str]]:
    segments = []
    for match in re.finditer(r'"([^"\r\n]*[\u3040-\u30ff\u3400-\u9fff][^"\r\n]*)"', text):
        value = match.group(1).strip()
        if value and not looks_like_mojibake(value):
            segments.append((match.start(1), match.end(1), value))
    return segments


def extract_robot_program_blocks(raw: bytes) -> list[TextBlock]:
    text, _encoding = decode_robot_program_text_with_encoding(raw)
    blocks = []
    for start, end, value in robot_comment_segments(text):
        if value:
            blocks.append(TextBlock(location=f"robot_comment:{start}:{end}", text=value))
    for start, end, value in robot_quoted_string_segments(text):
        if value:
            blocks.append(TextBlock(location=f"robot_string:{start}:{end}", text=value))
    return blocks


def build_translated_robot_program(
    raw: bytes,
    translations: dict[str, str],
    source_by_location: dict[str, str],
    keep_source_with_translation: bool = False,
) -> bytes:
    text, encoding = decode_robot_program_text_with_encoding(raw)
    replacements = []
    for location, source_text in source_by_location.items():
        if not location.startswith(("robot_comment:", "robot_string:")) or location not in translations:
            continue
        _prefix, start_text, end_text = location.split(":", 2)
        replacement = output_translation_for(
            location,
            source_text,
            translations,
            keep_source_with_translation,
        )
        if keep_source_with_translation:
            replacement = replacement.replace("\r\n", " / ").replace("\n", " / ").replace("\r", " / ")
        replacements.append((int(start_text), int(end_text), replacement))

    for start, end, replacement in sorted(replacements, reverse=True):
        text = text[:start] + replacement + text[end:]

    try:
        return text.encode(encoding)
    except UnicodeEncodeError:
        return text.encode("utf-8-sig")


def sniff_csv_dialect(text: str):
    sample = "\n".join(text.splitlines()[:100])
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        return csv.excel


def split_leading_code_cell(value: str) -> list[str]:
    cleaned = clean_text(value)
    match = LEADING_CODE_PATTERN.match(cleaned)
    if not match:
        return [value]
    code, remainder = match.groups()
    remainder = clean_text(remainder)
    if not remainder:
        return [value]
    return [code, remainder]


def normalize_csv_structure(rows: list[list[str]]) -> list[list[str]]:
    normalized = []
    for row in rows:
        if len(row) == 1:
            normalized.append(split_leading_code_cell(row[0]))
        else:
            normalized.append(row)
    return normalized


def read_csv_rows(raw: bytes) -> list[list[str]]:
    text = decode_document_text(raw)
    return parse_csv_rows_lenient(text)


def parse_csv_rows_lenient(text: str) -> list[list[str]]:
    if "ï¿½" in text and not has_japanese_text(text):
        raise ValueError(
            "The CSV text could not be decoded into readable Japanese. "
            "Please save the CSV as UTF-8 CSV or Excel .xlsx, then upload again."
        )
    dialect = sniff_csv_dialect(text)
    try:
        return normalize_csv_structure([row for row in csv.reader(io.StringIO(text, newline=""), dialect)])
    except csv.Error:
        rows = []
        for line in text.splitlines():
            try:
                rows.append(next(csv.reader([line], dialect)))
            except csv.Error:
                rows.append([line])
        return normalize_csv_structure(rows)


def extract_csv_blocks(raw: bytes) -> list[TextBlock]:
    rows = read_csv_rows(raw)
    blocks = []
    for row_index, row in enumerate(rows):
        for column_index, cell in enumerate(row):
            value = clean_text(cell)
            if value:
                blocks.append(TextBlock(location=f"csv:{row_index}:{column_index}", text=value))
    return blocks


def build_translated_csv(
    raw: bytes,
    translations: dict[str, str],
    source_by_location: dict[str, str],
    keep_source_with_translation: bool = False,
) -> bytes:
    rows = read_csv_rows(raw)
    for row_index, row in enumerate(rows):
        for column_index, cell in enumerate(row):
            key = f"csv:{row_index}:{column_index}"
            if key in translations:
                row[column_index] = output_translation_for(
                    key,
                    source_by_location.get(key, clean_text(cell)),
                    translations,
                    keep_source_with_translation,
                )

    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")


def extract_text_blocks(raw: bytes, file_name: str) -> list[TextBlock]:
    lower_name = file_name.lower()
    if lower_name.endswith((".as", ".ad")):
        return extract_robot_program_blocks(raw)

    if lower_name.endswith(".txt"):
        text = read_text_file(raw)
        return [TextBlock(location=f"text:{index}", text=part.strip()) for index, part in enumerate(text.split("\n\n")) if part.strip()]

    if lower_name.endswith(".csv"):
        return extract_csv_blocks(raw)

    if lower_name.endswith(".docx"):
        return extract_docx_blocks(raw)

    if lower_name.endswith(".pptx"):
        return extract_pptx_blocks(raw)

    if lower_name.endswith((".xlsx", ".xlsm")):
        return extract_xlsx_blocks(raw)

    if lower_name.endswith(".pdf"):
        return extract_pdf_blocks(raw)

    raise ValueError("Supported document types: CSV, TXT, AS, AD, DOCX, PPTX, XLSX, XLSM, PDF.")


def no_blocks_error_message(file_name: str) -> str:
    if file_name.lower().endswith((".as", ".ad")):
        return (
            "No Japanese text was found inside semicolon-delimited robot comments. "
            "For AS/AD files, this app translates only Japanese text between semicolons, such as ;æ¬é€é–‹å§‹;."
        )
    return (
        "No translatable text was found in this document. "
        "Upload a CSV, TXT, AS, AD, DOCX, PPTX, XLSX, XLSM, or text-based PDF with selectable text."
    )


def require_pdf_engine() -> None:
    if fitz is None:
        raise RuntimeError("PDF support requires PyMuPDF. Install it with: python -m pip install PyMuPDF")


def extract_pdf_blocks(raw: bytes) -> list[TextBlock]:
    """Extract visual PDF lines while preserving page and drawing coordinates.

    CAD and engineering PDFs commonly split one visible label into many font
    spans. Treating every span as a translation request creates excessive API
    work and loses sentence context. One visual line is the smallest useful
    translation and placement unit.
    """
    require_pdf_engine()
    blocks: list[TextBlock] = []
    with fitz.open(stream=raw, filetype="pdf") as document:
        for page_index, page in enumerate(document):
            page_dict = page.get_text("dict", flags=fitz.TEXTFLAGS_TEXT)
            for block_index, pdf_block in enumerate(page_dict.get("blocks", [])):
                if pdf_block.get("type") != 0:
                    continue
                for line_index, line in enumerate(pdf_block.get("lines", [])):
                    spans = [span for span in line.get("spans", []) if clean_text(span.get("text", ""))]
                    if not spans:
                        continue
                    text = clean_text("".join(str(span.get("text", "")) for span in spans))
                    boxes = [span.get("bbox") for span in spans if len(span.get("bbox") or ()) == 4]
                    if not text or not boxes:
                        continue
                    x0 = round(min(float(box[0]) for box in boxes), 3)
                    y0 = round(min(float(box[1]) for box in boxes), 3)
                    x1 = round(max(float(box[2]) for box in boxes), 3)
                    y1 = round(max(float(box[3]) for box in boxes), 3)
                    size = round(max(float(span.get("size", 8.0)) for span in spans), 3)
                    location = (
                        f"pdf:p{page_index}:b{block_index}:l{line_index}:s0:"
                        f"{x0},{y0},{x1},{y1}:{size}"
                    )
                    blocks.append(TextBlock(location=location, text=text))
    return blocks


PDF_LOCATION_PATTERN = re.compile(
    r"^pdf:p(?P<page>\d+):b\d+:l\d+:s\d+:"
    r"(?P<x0>-?[\d.]+),(?P<y0>-?[\d.]+),(?P<x1>-?[\d.]+),(?P<y1>-?[\d.]+):"
    r"(?P<size>[\d.]+)$"
)


def pdf_span_details(location: str):
    match = PDF_LOCATION_PATTERN.match(location)
    if not match:
        return None
    values = match.groupdict()
    return (
        int(values["page"]),
        fitz.Rect(float(values["x0"]), float(values["y0"]), float(values["x1"]), float(values["y1"])),
        float(values["size"]),
    )


PDF_ENGINEERING_FIELD_TERMS = {
    "寿命時間 [年]": "Service Life [years]",
    "寿命時間[年]": "Service Life [years]",
    "最短使命時間 [年]": "Minimum Mission Time [years]",
    "最短使命時間[年]": "Minimum Mission Time [years]",
    "参照記号": "Reference Designation",
    "技術的分類": "Technical Classification",
    "部品の製造者": "Component Manufacturer",
    "部品の識別子": "Component Identifier",
    "部品のグループ": "Component Group",
    "部品番号": "Part Number",
    "評価者": "Evaluator",
}


def canonical_pdf_translation(source_text: str, translated_text: str) -> str:
    """Lock recurring engineering form labels to stable English terminology."""
    source = clean_text(source_text)
    suffix = ":" if source.endswith(":") else ""
    core = source[:-1].strip() if suffix else source
    canonical = PDF_ENGINEERING_FIELD_TERMS.get(core)
    if canonical:
        return f"{canonical}{suffix}"
    for japanese, english in PDF_ENGINEERING_FIELD_TERMS.items():
        if source.startswith(japanese):
            remainder = source[len(japanese):]
            return f"{english}{remainder}"
    return " ".join(str(translated_text).split())


def dedupe_pdf_replacements(replacements):
    """Remove duplicate shadow-text placements exported by engineering PDFs."""
    accepted = []
    for rect, font_size, translated in sorted(
        replacements, key=lambda item: (item[0].y0, item[0].x0, -item[0].get_area())
    ):
        duplicate = False
        for prior_rect, _, _ in accepted[-12:]:
            intersection = rect & prior_rect
            smaller = min(rect.get_area(), prior_rect.get_area())
            if smaller > 0 and intersection.get_area() / smaller >= 0.82:
                duplicate = True
                break
            if (
                abs(rect.x0 - prior_rect.x0) <= 1.2
                and abs(rect.y0 - prior_rect.y0) <= 1.2
                and abs(rect.x1 - prior_rect.x1) <= 1.2
                and abs(rect.y1 - prior_rect.y1) <= 1.2
            ):
                duplicate = True
                break
        if not duplicate:
            accepted.append((rect, font_size, translated))
    return accepted


def consolidate_pdf_row_fragments(replacements):
    """Merge adjacent fragments that visually form one engineering-form field."""
    items = dedupe_pdf_replacements(replacements)
    rows = []
    for item in items:
        rect = item[0]
        for row in rows:
            row_rect = row[-1][0]
            if abs(rect.y0 - row_rect.y0) <= 1.8 and abs(rect.y1 - row_rect.y1) <= 1.8:
                row.append(item)
                break
        else:
            rows.append([item])
    consolidated = []
    for row in rows:
        row.sort(key=lambda item: item[0].x0)
        current = None
        for rect, font_size, translated in row:
            if current is None:
                current = [fitz.Rect(rect), font_size, translated]
                continue
            gap = rect.x0 - current[0].x1
            if gap <= 12:
                current[0] |= rect
                current[1] = max(current[1], font_size)
                if clean_text(translated) != clean_text(current[2]):
                    current[2] = f"{current[2]} {translated}".strip()
            else:
                consolidated.append(tuple(current))
                current = [fitz.Rect(rect), font_size, translated]
        if current is not None:
            consolidated.append(tuple(current))
    return consolidated


def pdf_available_text_rect(page, source_rect, page_words=None):
    """Use empty space on the same form row without covering the next value."""
    page_width = page.rect.width
    right_limit = page_width - 8
    for item in (page_words if page_words is not None else page.get_text("words")):
        other = fitz.Rect(item[:4])
        if other.x0 <= source_rect.x1 + 0.5:
            continue
        overlap = min(source_rect.y1, other.y1) - max(source_rect.y0, other.y0)
        if overlap >= min(source_rect.height, other.height) * 0.35:
            right_limit = min(right_limit, other.x0 - 2)
    return fitz.Rect(
        max(0, source_rect.x0 - 0.4),
        max(0, source_rect.y0 - 0.4),
        max(source_rect.x1 + 1, right_limit),
        source_rect.y1 + max(1.0, source_rect.height * 0.25),
    )


def insert_fitted_pdf_text(page, rect, text: str, original_size: float, page_words=None) -> None:
    """Place readable English in available row space with bounded fitting."""
    cover = fitz.Rect(rect.x0 - 0.4, rect.y0 - 0.4, rect.x1 + 1.0, rect.y1 + 0.8)
    target = pdf_available_text_rect(page, rect, page_words)
    page.draw_rect(cover, color=None, fill=(1, 1, 1), overlay=True)
    clean = " ".join(str(text).split())
    estimated = target.width / max(len(clean) * 0.52, 1.0)
    font_size = max(4.5, min(original_size, target.height * 0.72, estimated))
    for trial_size in (font_size, max(4.5, font_size * 0.9), 4.5):
        remaining = page.insert_textbox(
            target, clean, fontname="helv", fontsize=trial_size,
            color=(0, 0, 0), align=fitz.TEXT_ALIGN_LEFT, overlay=True,
        )
        if remaining >= 0:
            return
    page.insert_text(
        (target.x0, target.y1 - 0.5),
        clean,
        fontname="helv",
        fontsize=4.5,
        color=(0, 0, 0),
        overlay=True,
    )


def build_translated_pdf(
    raw: bytes,
    translations: dict[str, str],
    blocks: list[TextBlock],
    keep_source_with_translation: bool = False,
) -> bytes:
    require_pdf_engine()
    with fitz.open(stream=raw, filetype="pdf") as document:
        replacements_by_page: dict[int, list[tuple[fitz.Rect, float, str]]] = {}
        for block in blocks:
            if block.location not in translations:
                continue
            details = pdf_span_details(block.location)
            if details is None:
                continue
            page_index, rect, font_size = details
            translated = output_translation_for(
                block.location,
                block.text,
                translations,
                keep_source_with_translation,
            )
            translated = canonical_pdf_translation(block.text, translated)
            replacements_by_page.setdefault(page_index, []).append((rect, font_size, translated))

        for page_index, replacements in replacements_by_page.items():
            page = document[page_index]
            page_words = page.get_text("words")
            for rect, font_size, translated in consolidate_pdf_row_fragments(replacements):
                insert_fitted_pdf_text(page, rect, translated, font_size, page_words)
        metadata = document.metadata or {}
        metadata["subject"] = "Term1 glossary-first translation review copy; verify before controlled engineering use."
        document.set_metadata(metadata)
        return document.tobytes(garbage=1, deflate=True)


def extract_docx_blocks(raw: bytes) -> list[TextBlock]:
    blocks = []
    with ZipFile(io.BytesIO(raw)) as archive:
        xml_names = [
            name for name in archive.namelist()
            if name == "word/document.xml" or name.startswith("word/header") or name.startswith("word/footer")
        ]
        for xml_name in xml_names:
            root = ET.fromstring(archive.read(xml_name))
            for index, paragraph in enumerate(root.findall(".//w:p", WORD_NS)):
                text = "".join(node.text or "" for node in paragraph.findall(".//w:t", WORD_NS)).strip()
                if text:
                    blocks.append(TextBlock(location=f"{xml_name}#{index}", text=text))
    return blocks


def ppt_paragraph_text(paragraph: ET.Element) -> str:
    parts = []
    for child in list(paragraph):
        if child.tag == f"{{{PPT_NS['a']}}}br":
            parts.append("\n")
            continue
        parts.extend(node.text or "" for node in child.findall(".//a:t", PPT_NS))
    return "".join(parts).strip()


def ppt_text_bodies(root: ET.Element) -> list[ET.Element]:
    return [element for element in root.iter() if element.tag.endswith("}txBody")]


def ppt_text_body_text(text_body: ET.Element) -> str:
    return "\n".join(
        text
        for text in (ppt_paragraph_text(paragraph) for paragraph in text_body.findall("a:p", PPT_NS))
        if text
    ).strip()


def extract_pptx_blocks(raw: bytes) -> list[TextBlock]:
    blocks = []
    with ZipFile(io.BytesIO(raw)) as archive:
        slide_names = sorted(
            name for name in archive.namelist()
            if name.startswith("ppt/slides/slide") and name.endswith(".xml")
        )
        for slide_name in slide_names:
            root = ET.fromstring(archive.read(slide_name))
            for index, text_body in enumerate(ppt_text_bodies(root)):
                text = ppt_text_body_text(text_body)
                if text:
                    blocks.append(
                        TextBlock(
                            location=f"{slide_name}#textbox:{index}",
                            text=text,
                        )
                    )
    return blocks


def read_shared_strings(archive: ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []

    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    strings = []
    for item in root.findall("a:si", EXCEL_NS):
        strings.append("".join(text.text or "" for text in item.findall(".//a:t", EXCEL_NS)))
    return strings


def cell_text(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(text.text or "" for text in cell.findall(".//a:t", EXCEL_NS)).strip()

    value_node = cell.find("a:v", EXCEL_NS)
    if value_node is None or value_node.text is None:
        return ""

    value = value_node.text
    if cell_type == "s" and value.isdigit():
        index = int(value)
        if 0 <= index < len(shared_strings):
            return shared_strings[index].strip()

    if cell_type in {"str", "e"}:
        return value.strip()

    return ""


def extract_xlsx_blocks(raw: bytes) -> list[TextBlock]:
    blocks = []
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=False)
        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            sheet_name = f"xl/worksheets/sheet{sheet_index}.xml"
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.strip():
                        blocks.append(TextBlock(location=f"{sheet_name}#{cell.coordinate}", text=clean_text(value)))
        workbook.close()
        if blocks:
            return blocks
    except Exception:
        blocks = []

    with ZipFile(io.BytesIO(raw)) as archive:
        shared_strings = read_shared_strings(archive)
        sheet_names = sorted(
            name for name in archive.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )

        for sheet_name in sheet_names:
            root = ET.fromstring(archive.read(sheet_name))
            for cell in root.findall(".//a:c", EXCEL_NS):
                ref = cell.attrib.get("r", "")
                text = cell_text(cell, shared_strings)
                if text:
                    blocks.append(TextBlock(location=f"{sheet_name}#{ref}", text=text))
    return blocks


def first_run_with_text(paragraph: ET.Element) -> ET.Element | None:
    for run in paragraph.findall(".//w:r", WORD_NS):
        if run.findall(".//w:t", WORD_NS):
            return run
    return None


def replace_text_in_paragraph(paragraph: ET.Element, text_nodes: list[ET.Element], translated_text: str) -> None:
    if not text_nodes:
        return
    lines = str(translated_text).splitlines() or [""]
    text_nodes[0].text = clean_office_xml_text(lines[0])
    for node in text_nodes[1:]:
        node.text = ""
    if len(lines) <= 1:
        return

    first_run = first_run_with_text(paragraph)
    if first_run is None:
        return
    for line in lines[1:]:
        ET.SubElement(first_run, f"{{{WORD_NS['w']}}}br")
        text_node = ET.SubElement(first_run, f"{{{WORD_NS['w']}}}t")
        text_node.text = clean_office_xml_text(line)


def build_translated_document(
    raw: bytes,
    file_name: str,
    translations: dict[str, str],
    blocks: list[TextBlock],
    keep_source_with_translation: bool = False,
) -> bytes:
    lower_name = file_name.lower()
    source_by_location = {block.location: block.text for block in blocks}
    if lower_name.endswith((".as", ".ad")):
        return build_translated_robot_program(raw, translations, source_by_location, keep_source_with_translation)

    if lower_name.endswith(".txt"):
        return write_text_file(blocks, translations, keep_source_with_translation)

    if lower_name.endswith(".csv"):
        return build_translated_csv(raw, translations, source_by_location, keep_source_with_translation)

    if lower_name.endswith(".docx"):
        return build_translated_docx(raw, translations, source_by_location, keep_source_with_translation)

    if lower_name.endswith(".pptx"):
        return build_translated_pptx(raw, translations, source_by_location, keep_source_with_translation)

    if lower_name.endswith((".xlsx", ".xlsm")):
        return build_translated_xlsx(raw, translations, source_by_location, keep_source_with_translation)

    if lower_name.endswith(".pdf"):
        return build_translated_pdf(raw, translations, blocks, keep_source_with_translation)

    raise ValueError("Supported document types: CSV, TXT, AS, AD, DOCX, PPTX, XLSX, XLSM, PDF.")


def build_translated_docx(
    raw: bytes,
    translations: dict[str, str],
    source_by_location: dict[str, str],
    keep_source_with_translation: bool = False,
) -> bytes:
    source = io.BytesIO(raw)
    target = io.BytesIO()

    with ZipFile(source) as input_zip, ZipFile(target, "w", ZIP_DEFLATED) as output_zip:
        for item in input_zip.infolist():
            data = input_zip.read(item.filename)
            if item.filename == "word/document.xml" or item.filename.startswith("word/header") or item.filename.startswith("word/footer"):
                root = ET.fromstring(data)
                for index, paragraph in enumerate(root.findall(".//w:p", WORD_NS)):
                    key = f"{item.filename}#{index}"
                    if key in translations:
                        translated = output_translation_for(
                            key,
                            source_by_location.get(key, ""),
                            translations,
                            keep_source_with_translation,
                        )
                        replace_text_in_paragraph(paragraph, paragraph.findall(".//w:t", WORD_NS), translated)
                data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            output_zip.writestr(item, data)

    return target.getvalue()


def replace_text_in_ppt_paragraph(paragraph: ET.Element, translated_text: str) -> None:
    runs = paragraph.findall("a:r", PPT_NS)
    first_run = next((run for run in runs if run.find("a:t", PPT_NS) is not None), None)
    if first_run is None:
        return

    lines = str(translated_text).splitlines() or [""]
    first_text = first_run.find("a:t", PPT_NS)
    if first_text is None:
        return
    first_text.text = clean_office_xml_text(lines[0])

    for child in list(paragraph):
        if child is first_run:
            continue
        if child.tag in {
            f"{{{PPT_NS['a']}}}r",
            f"{{{PPT_NS['a']}}}br",
            f"{{{PPT_NS['a']}}}fld",
        }:
            paragraph.remove(child)

    insertion_index = list(paragraph).index(first_run) + 1
    run_properties = first_run.find("a:rPr", PPT_NS)
    for line in lines[1:]:
        paragraph.insert(insertion_index, ET.Element(f"{{{PPT_NS['a']}}}br"))
        insertion_index += 1
        new_run = ET.Element(f"{{{PPT_NS['a']}}}r")
        if run_properties is not None:
            new_run.append(deepcopy(run_properties))
        new_text = ET.SubElement(new_run, f"{{{PPT_NS['a']}}}t")
        new_text.text = clean_office_xml_text(line)
        paragraph.insert(insertion_index, new_run)
        insertion_index += 1


def ppt_paragraph_is_list_item(paragraph: ET.Element) -> bool:
    text = ppt_paragraph_text(paragraph).lstrip()
    if text.startswith(("・", "•", "-", "–", "—", "*")):
        return True
    paragraph_properties = paragraph.find("a:pPr", PPT_NS)
    if paragraph_properties is None:
        return False
    return any(
        child.tag.endswith(("}buChar", "}buAutoNum", "}buBlip"))
        for child in paragraph_properties
    )


def replace_text_in_ppt_text_body(text_body: ET.Element, translated_text: str) -> ET.Element | None:
    paragraphs = text_body.findall("a:p", PPT_NS)
    populated = [paragraph for paragraph in paragraphs if ppt_paragraph_text(paragraph)]
    if not populated:
        return None

    translated_lines = [line.strip() for line in str(translated_text).splitlines() if line.strip()]
    preserve_list = (
        len(populated) > 1
        and all(ppt_paragraph_is_list_item(paragraph) for paragraph in populated)
        and len(translated_lines) == len(populated)
    )
    if preserve_list:
        for paragraph, line in zip(populated, translated_lines):
            replace_text_in_ppt_paragraph(paragraph, line)
        return populated[0]

    first_paragraph = populated[0]
    replace_text_in_ppt_paragraph(first_paragraph, translated_text)
    for paragraph in paragraphs:
        if paragraph is not first_paragraph:
            text_body.remove(paragraph)
    return first_paragraph


def enable_ppt_text_autofit(paragraph: ET.Element, parent_map: dict[ET.Element, ET.Element]) -> None:
    current = paragraph
    text_body = None
    while current in parent_map:
        current = parent_map[current]
        if current.tag.endswith("}txBody"):
            text_body = current
            break
    if text_body is None:
        return

    body_properties = text_body.find(f"{{{PPT_NS['a']}}}bodyPr")
    if body_properties is None:
        return
    autofit_tags = {
        f"{{{PPT_NS['a']}}}noAutofit",
        f"{{{PPT_NS['a']}}}normAutofit",
        f"{{{PPT_NS['a']}}}spAutoFit",
    }
    for child in list(body_properties):
        if child.tag in autofit_tags:
            body_properties.remove(child)
    ET.SubElement(
        body_properties,
        f"{{{PPT_NS['a']}}}normAutofit",
        {"fontScale": "70000", "lnSpcReduction": "20000"},
    )


def build_translated_pptx(
    raw: bytes,
    translations: dict[str, str],
    source_by_location: dict[str, str],
    keep_source_with_translation: bool = False,
) -> bytes:
    source = io.BytesIO(raw)
    target = io.BytesIO()

    with ZipFile(source) as input_zip, ZipFile(target, "w", ZIP_DEFLATED) as output_zip:
        for item in input_zip.infolist():
            data = input_zip.read(item.filename)
            if item.filename.startswith("ppt/slides/slide") and item.filename.endswith(".xml"):
                root = ET.fromstring(data)
                parent_map = {child: parent for parent in root.iter() for child in parent}
                for index, text_body in enumerate(ppt_text_bodies(root)):
                    key = f"{item.filename}#textbox:{index}"
                    if key in translations:
                        source_text = source_by_location.get(key, "")
                        translated = output_translation_for(
                            key,
                            source_text,
                            translations,
                            keep_source_with_translation,
                        )
                        first_paragraph = replace_text_in_ppt_text_body(text_body, translated)
                        if first_paragraph is not None and len(clean_text(translated)) > max(len(clean_text(source_text)) * 1.15, 12):
                            enable_ppt_text_autofit(first_paragraph, parent_map)
                data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            output_zip.writestr(item, data)

    return target.getvalue()


def replace_excel_cell_text(cell: ET.Element, translated_text: str) -> None:
    for child in list(cell):
        cell.remove(child)

    cell.attrib["t"] = "inlineStr"
    inline_string = ET.SubElement(cell, f"{{{EXCEL_NS['a']}}}is")
    text_node = ET.SubElement(inline_string, f"{{{EXCEL_NS['a']}}}t")
    text_node.text = clean_office_xml_text(translated_text)


def serialize_excel_xml(root: ET.Element) -> bytes:
    used_uris = set()
    for element in root.iter():
        if element.tag.startswith("{"):
            used_uris.add(element.tag[1:].split("}", 1)[0])
        for attr_name in element.attrib:
            if attr_name.startswith("{"):
                used_uris.add(attr_name[1:].split("}", 1)[0])

    ignorable_attr = "{http://schemas.openxmlformats.org/markup-compatibility/2006}Ignorable"
    if ignorable_attr in root.attrib:
        kept_prefixes = [
            prefix
            for prefix in root.attrib[ignorable_attr].split()
            if EXCEL_SERIALIZE_NAMESPACES.get(prefix) in used_uris
        ]
        if kept_prefixes:
            root.attrib[ignorable_attr] = " ".join(kept_prefixes)
        else:
            root.attrib.pop(ignorable_attr, None)

    for prefix, uri in EXCEL_SERIALIZE_NAMESPACES.items():
        ET.register_namespace(prefix, uri)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def build_translated_xlsx(
    raw: bytes,
    translations: dict[str, str],
    source_by_location: dict[str, str],
    keep_source_with_translation: bool = False,
) -> bytes:
    source = io.BytesIO(raw)
    target = io.BytesIO()

    with ZipFile(source) as input_zip, ZipFile(target, "w", ZIP_DEFLATED) as output_zip:
        for item in input_zip.infolist():
            data = input_zip.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet") and item.filename.endswith(".xml"):
                root = ET.fromstring(data)
                for cell in root.findall(".//a:c", EXCEL_NS):
                    key = f"{item.filename}#{cell.attrib.get('r', '')}"
                    if key in translations:
                        translated = output_translation_for(
                            key,
                            source_by_location.get(key, ""),
                            translations,
                            keep_source_with_translation,
                        )
                        replace_excel_cell_text(cell, translated)
                data = serialize_excel_xml(root)
            output_zip.writestr(item, data)

    return target.getvalue()


def output_file_name(file_name: str) -> str:
    path = Path(file_name)
    return f"Translated-{path.stem}{path.suffix or '.txt'}"


def mime_type(file_name: str) -> str:
    lower_name = file_name.lower()
    if lower_name.endswith(".csv"):
        return "text/csv"
    if lower_name.endswith((".as", ".ad")):
        return "text/plain"
    if lower_name.endswith(".docx"):
        return "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    if lower_name.endswith(".pptx"):
        return "application/vnd.openxmlformats-officedocument.presentationml.presentation"
    if lower_name.endswith(".xlsx"):
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if lower_name.endswith(".xlsm"):
        return "application/vnd.ms-excel.sheet.macroEnabled.12"
    if lower_name.endswith(".pdf"):
        return "application/pdf"
    return "text/plain"


def safe_storage_name(file_name: str) -> str:
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(file_name).name).strip("_")
    return safe_name or "document.txt"


def ensure_job_storage_dirs() -> None:
    JOB_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    JOB_RESULT_DIR.mkdir(parents=True, exist_ok=True)


def job_upload_path(job_id: str, file_name: str) -> Path:
    ensure_job_storage_dirs()
    return JOB_UPLOAD_DIR / f"{job_id}_{safe_storage_name(file_name)}"


def job_result_path(job_id: str, file_name: str) -> Path:
    ensure_job_storage_dirs()
    translated_name = output_file_name(file_name)
    return JOB_RESULT_DIR / f"{job_id}_{safe_storage_name(translated_name)}"


def should_use_large_pdf_renderer(
    file_name: str,
    raw_document: bytes,
    blocks: list[TextBlock],
) -> bool:
    if not file_name.lower().endswith(".pdf") or fitz is None:
        return False
    with fitz.open(stream=raw_document, filetype="pdf") as document:
        return (
            document.page_count >= LARGE_PDF_PAGE_THRESHOLD
            or len(blocks) >= LARGE_PDF_BLOCK_THRESHOLD
        )


def run_large_pdf_renderer_for_job(
    job_id: str,
    raw_document: bytes,
    file_name: str,
    checkpoint_path: Path,
    result_path: Path,
) -> None:
    """Generate a large translated PDF in resumable chunks outside Streamlit."""
    renderer = BASE_DIR / "generate_large_pdf.py"
    if not renderer.exists():
        raise RuntimeError(f"Large PDF renderer is missing: {renderer}")

    source_path = job_upload_path(job_id, file_name)
    if not source_path.exists():
        source_path.write_bytes(raw_document)
    work_dir = LARGE_PDF_WORK_DIR / job_id
    work_dir.mkdir(parents=True, exist_ok=True)
    progress_path = work_dir / "generation-progress.json"
    stdout_path = work_dir / "renderer-out.log"
    stderr_path = work_dir / "renderer-err.log"

    command = [
        sys.executable,
        str(renderer),
        "--source", str(source_path),
        "--checkpoint", str(checkpoint_path),
        "--output", str(result_path),
        "--work-dir", str(work_dir),
        "--chunk-pages", "100",
    ]
    creation_flags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    with stdout_path.open("a", encoding="utf-8") as stdout_log, stderr_path.open(
        "a", encoding="utf-8"
    ) as stderr_log:
        process = subprocess.Popen(
            command,
            cwd=BASE_DIR,
            stdout=stdout_log,
            stderr=stderr_log,
            creationflags=creation_flags,
        )
        while process.poll() is None:
            if not translation_job_is_active(job_id):
                process.terminate()
                raise RuntimeError("Translation job was stopped.")
            if progress_path.exists():
                try:
                    progress = json.loads(progress_path.read_text(encoding="utf-8"))
                    generated = int(progress.get("generated_pages", 0))
                    total = int(progress.get("total_pages", 0))
                    percent = 0.0 if total <= 0 else 100 * generated / total
                    update_translation_job(
                        job_id,
                        generated_pages=generated,
                        total_pages=total,
                        progress_message=(
                            f"Generating English PDF pages: {generated:,}/{total:,} "
                            f"({percent:.1f}%)."
                        ),
                    )
                except (OSError, ValueError, json.JSONDecodeError):
                    pass
            time.sleep(2)

    if process.returncode != 0:
        detail = ""
        try:
            detail = stderr_path.read_text(encoding="utf-8", errors="replace")[-2000:]
        except OSError:
            pass
        raise RuntimeError(
            f"Large PDF renderer exited with code {process.returncode}. {detail}".strip()
        )
    if not result_path.exists() or result_path.stat().st_size == 0:
        raise RuntimeError("Large PDF renderer finished without creating an output file.")


def is_valid_email(value: str) -> bool:
    return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", value.strip()))


def is_smtp_configured() -> bool:
    smtp_host = os.getenv("SMTP_HOST", "").strip()
    smtp_from = os.getenv("SMTP_FROM", os.getenv("SMTP_USERNAME", "")).strip()
    return bool(smtp_host and smtp_from)


def manual_email_status() -> str:
    return "Manual email draft ready. Download the translated file and attach it yourself."


def translation_mailto_link(to_email: str, file_name: str, result_file_name: str) -> str:
    subject = f"Term1 translation completed: {result_file_name}"
    body = "\n".join(
        [
            "Your Term1 translation job is complete.",
            "",
            f"Source file: {file_name}",
            f"Translated file: {result_file_name}",
            "",
            "Please attach the downloaded translated file before sending this email.",
        ]
    )
    return f"mailto:{quote(to_email)}?subject={quote(subject)}&body={quote(body)}"


def send_completed_translation_email(to_email: str, file_name: str, result_path: Path, result_file_name: str) -> str:
    if not to_email:
        return ""
    if not is_valid_email(to_email):
        return "Invalid email address."

    smtp_host = os.getenv("SMTP_HOST", "").strip()
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USERNAME", "").strip()
    smtp_password = os.getenv("SMTP_PASSWORD", "")
    smtp_from = os.getenv("SMTP_FROM", smtp_user).strip()
    smtp_tls = os.getenv("SMTP_TLS", "true").lower() in {"1", "true", "yes"}

    if not smtp_host or not smtp_from:
        return manual_email_status()
    if not result_path.exists():
        return "Email not sent: translated file was not found."
    if result_path.stat().st_size > MAX_EMAIL_ATTACHMENT_BYTES:
        return "Email not sent: translated file is larger than the email attachment limit. Download it from the app."

    message = EmailMessage()
    message["From"] = smtp_from
    message["To"] = to_email
    message["Subject"] = f"Term1 translation completed: {result_file_name}"
    message.set_content(
        "\n".join(
            [
                "Your Term1 translation job is complete.",
                "",
                f"Source file: {file_name}",
                f"Translated file: {result_file_name}",
                "",
                "The translated file is attached.",
            ]
        )
    )

    content_type, _ = mimetypes.guess_type(result_file_name)
    maintype, subtype = (content_type or "application/octet-stream").split("/", 1)
    message.add_attachment(
        result_path.read_bytes(),
        maintype=maintype,
        subtype=subtype,
        filename=result_file_name,
    )

    with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
        if smtp_tls:
            server.starttls()
        if smtp_user and smtp_password:
            server.login(smtp_user, smtp_password)
        server.send_message(message)

    return "Email sent."


@st.cache_resource
def background_job_executor() -> ThreadPoolExecutor:
    return ThreadPoolExecutor(max_workers=4)


def prepare_and_run_document_translation_job(
    job_id: str,
    raw_document: bytes,
    file_name: str,
    glossary: pd.DataFrame,
    translation_mode: str,
    translation_direction: str,
    keep_source_with_translation: bool,
    owner_session_id: str,
) -> None:
    try:
        update_translation_job(job_id, status="running", error_message="", progress_message="Preparing file.")
        blocks = extract_text_blocks(raw_document, file_name)
        translatable_blocks = [block for block in blocks if should_translate(block.text, translation_direction)]
        memory_mode = translation_cache_mode(translation_mode, translation_direction)
        checkpoint_path = checkpoint_path_for(
            file_name,
            raw_document,
            translation_mode,
            translation_direction,
            owner_session_id,
        )
        saved_translations = load_checkpoint(checkpoint_path)
        hydrate_translation_memory_from_checkpoint(blocks, saved_translations, memory_mode)
        source_by_key = {}
        for block in translatable_blocks:
            if not block_has_translation_context(block):
                source_by_key.setdefault(translation_memory_key(block.text), block.text)
        memory_hits = translation_memory_lookup(list(source_by_key.values()), memory_mode)
        memory_applied = 0
        for block in translatable_blocks:
            if block.location in saved_translations:
                continue
            if block_has_translation_context(block):
                continue
            memory_translation = memory_hits.get(translation_memory_key(block.text))
            if memory_translation:
                saved_translations[block.location] = memory_translation
                memory_applied += 1
        if memory_applied:
            save_checkpoint(checkpoint_path, saved_translations)
        saved_memory_keys = {
            block_translation_key(block)
            for block in translatable_blocks
            if block.location in saved_translations
        }
        saved_count = sum(1 for block in translatable_blocks if block.location in saved_translations)
        pending_by_key = {}
        for block in translatable_blocks:
            key = block_translation_key(block)
            if block.location not in saved_translations and key not in saved_memory_keys:
                pending_by_key.setdefault(key, block)
        pending_for_batches = list(pending_by_key.values())
        pending_unique_count = len(pending_for_batches)
        batch_count = len(document_translation_chunks(pending_for_batches))
        update_translation_job(
            job_id,
            status="running",
            total_blocks=len(blocks),
            translatable_blocks=len(translatable_blocks),
            completed_blocks=saved_count,
            total_batches=batch_count,
            completed_batches=0,
            progress_message=(
                f"Preflight complete. Unique source text: {len(source_by_key):,}. "
                f"TM/checkpoint hits: {saved_count:,}. Remaining unique: {pending_unique_count:,}."
            ),
        )

        if not blocks:
            update_translation_job(
                job_id,
                status="failed",
                error_message=no_blocks_error_message(file_name),
                finished_at=time.strftime("%Y-%m-%d %H:%M:%S"),
            )
            return

        if not translatable_blocks:
            translated_document = build_translated_document(
                raw_document,
                file_name,
                {},
                blocks,
                keep_source_with_translation,
            )
            translated_name = output_file_name(file_name)
            result_path = job_result_path(job_id, file_name)
            result_path.write_bytes(translated_document)
            update_translation_job(
                job_id,
                status="completed",
                result_file_name=translated_name,
                result_file_path=str(result_path),
                result_mime=mime_type(translated_name),
                progress_message=f"No {direction_language_names(translation_direction)[0]} text found. Original document is ready.",
                finished_at=time.strftime("%Y-%m-%d %H:%M:%S"),
            )
            return

        run_document_translation_job(
            job_id,
            raw_document,
            file_name,
            blocks,
            glossary,
            translation_mode,
            translation_direction,
            keep_source_with_translation,
            checkpoint_path,
            batch_count,
        )
    except Exception as exc:
        if str(exc) == "Translation job was stopped.":
            return
        update_translation_job(
            job_id,
            status="failed",
            error_message=format_translation_error(exc),
            finished_at=time.strftime("%Y-%m-%d %H:%M:%S"),
        )


def run_document_translation_job(
    job_id: str,
    raw_document: bytes,
    file_name: str,
    blocks: list[TextBlock],
    glossary: pd.DataFrame,
    translation_mode: str,
    translation_direction: str,
    keep_source_with_translation: bool,
    checkpoint_path: Path,
    total_batches: int,
) -> None:
    started_at = time.time()

    def update_progress(done, total, done_batches, total_batches, elapsed, message):
        update_translation_job(
            job_id,
            status="running",
            completed_blocks=done,
            completed_batches=done_batches,
            progress_message=message,
        )

    try:
        update_translation_job(job_id, status="running", error_message="", progress_message="Starting translation.")
        translations, _, token_usage = translate_blocks_batch(
            blocks,
            glossary,
            translation_mode,
            translation_direction,
            checkpoint_path=checkpoint_path,
            progress_callback=update_progress,
            should_continue=lambda: translation_job_is_active(job_id),
        )
        translated_name = output_file_name(file_name)
        result_path = job_result_path(job_id, file_name)
        if should_use_large_pdf_renderer(file_name, raw_document, blocks):
            update_translation_job(
                job_id,
                generated_pages=0,
                total_pages=0,
                progress_message="Translation complete. Preparing the English PDF generator.",
            )
            run_large_pdf_renderer_for_job(
                job_id,
                raw_document,
                file_name,
                checkpoint_path,
                result_path,
            )
        else:
            translated_document = build_translated_document(
                raw_document,
                file_name,
                translations,
                blocks,
                keep_source_with_translation,
            )
            result_path.write_bytes(translated_document)
        notify_email = ""
        with sqlite3.connect(JOB_DB_PATH) as conn:
            row = conn.execute(
                "SELECT notify_email FROM translation_jobs WHERE job_id = ?",
                (job_id,),
            ).fetchone()
            notify_email = row[0] if row and row[0] else ""
        notification_status = ""
        if notify_email:
            try:
                notification_status = send_completed_translation_email(
                    notify_email,
                    file_name,
                    result_path,
                    translated_name,
                )
            except Exception as exc:
                notification_status = f"Email failed: {exc}"
        update_translation_job(
            job_id,
            status="completed",
            completed_blocks=sum(1 for block in blocks if should_translate(block.text, translation_direction)),
            completed_batches=total_batches,
            input_tokens=token_usage.input_tokens,
            output_tokens=token_usage.output_tokens,
            total_tokens=token_usage.total_tokens,
            result_file_name=translated_name,
            result_file_path=str(result_path),
            result_mime=mime_type(translated_name),
            notification_status=notification_status,
            progress_message=f"Completed in {format_duration(time.time() - started_at)}.",
            finished_at=time.strftime("%Y-%m-%d %H:%M:%S"),
        )
    except Exception as exc:
        if str(exc) == "Translation job was stopped.":
            return
        update_translation_job(
            job_id,
            status="failed",
            error_message=format_translation_error(exc),
            finished_at=time.strftime("%Y-%m-%d %H:%M:%S"),
        )


def terminology_report(
    hits: list[TermHit],
    glossary: pd.DataFrame | None = None,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> pd.DataFrame:
    if glossary is not None and not glossary.empty:
        usage_counts: dict[tuple[str, str], int] = {}
        ordered_terms: list[tuple[str, str]] = []
        for hit in hits:
            key = (hit.jp, hit.en)
            if key not in usage_counts:
                ordered_terms.append(key)
                usage_counts[key] = 0
            usage_counts[key] += hit.count

        rows: list[dict[str, object]] = []
        for jp, en in ordered_terms:
            matched_records = glossary[
                (glossary["JP"].astype(str).map(clean_text) == jp)
                & (glossary["EN"].astype(str).map(clean_text) == en)
            ]
            glossary_record = matched_records.iloc[0].to_dict() if not matched_records.empty else {
                "JP": jp,
                "EN": en,
            }
            row = {
                str(column): glossary_record.get(column, "")
                for column in glossary.columns
            }
            row["Used Count"] = usage_counts[(jp, en)]
            rows.append(row)
        return pd.DataFrame(rows)

    if translation_direction == TRANSLATION_DIRECTION_EN_JP:
        return pd.DataFrame(
            [{"English": hit.en, "Required Japanese": hit.jp, "Count": hit.count} for hit in hits]
        )
    return pd.DataFrame(
        [{"Japanese": hit.jp, "Required English": hit.en, "Count": hit.count} for hit in hits]
    )


def ai_model_version_text() -> str:
    return "\n".join(
        [
            "OpenAI",
            openai_model(),
        ]
    )


def glossary_version_text() -> str:
    glossary_path = next((path for path in DEFAULT_GLOSSARY_PATHS if path.exists()), None)
    if glossary_path is None:
        return "Glossary file was not found."

    raw = glossary_path.read_bytes()
    digest = hashlib.sha256(raw).hexdigest()[:12]
    modified = glossary_path.stat().st_mtime
    modified_text = datetime.fromtimestamp(modified, ZoneInfo("America/New_York")).strftime(
        "%Y-%m-%d %I:%M %p ET"
    )
    try:
        workbook = load_workbook(glossary_path, read_only=True, data_only=True)
        try:
            worksheet = workbook[workbook.sheetnames[0]]
            non_empty_a_cells = sum(
                1
                for (value,) in worksheet.iter_rows(min_col=1, max_col=1, values_only=True)
                if str(value or "").strip()
            )
            term_count = max(non_empty_a_cells - 1, 0)
        finally:
            workbook.close()
    except Exception:
        term_count = "Unavailable"
    term_count_text = f"{term_count:,}" if isinstance(term_count, int) else str(term_count)

    return "\n".join(
        [
            f"Terms: {term_count_text}",
            f"Last Updated: {modified_text}",
            "Design & Governance Owner: Aoi Minamoto (Controls)",
        ]
    )


def plc_rules_version_text() -> str:
    rule_path = next((path for path in DEFAULT_PLC_RULE_PATHS if path.exists()), None)
    if rule_path is None:
        expected = ", ".join(path.name for path in DEFAULT_PLC_RULE_PATHS)
        return f"PLC rule file was not found. Expected one of: {expected}"

    raw = rule_path.read_bytes()
    digest = hashlib.sha256(raw).hexdigest()[:12]
    modified = rule_path.stat().st_mtime
    modified_text = datetime.fromtimestamp(modified, ZoneInfo("America/New_York")).strftime(
        "%Y-%m-%d %I:%M %p ET"
    )
    try:
        rule_count = len(normalize_plc_rules(read_plc_rules()))
    except Exception:
        rule_count = "Unavailable"

    return "\n".join(
        [
            "PLC Rules Information",
            "",
            "Version: v1.0",
            f"Rules: {rule_count}",
            f"Last Updated: {modified_text}",
            "Owner: Aoi Minamoto",
            "",
            f"Source: {rule_path.name}",
        ]
    )


def apply_compact_style() -> None:
    st.markdown(
        """
        <style>
        div[data-testid="stMarkdownContainer"] p,
        div[data-testid="stCaptionContainer"],
        div[data-testid="stWidgetLabel"],
        div[data-testid="stFileUploader"],
        div[data-testid="stRadio"] label,
        div[data-testid="stCodeBlock"] pre {
            font-size: 1.5rem !important;
            line-height: 1.35;
        }

        button[data-baseweb="tab"],
        div[data-testid="stExpander"] summary {
            font-size: 1.5rem !important;
            font-weight: 600;
        }

        div.stButton > button,
        div.stDownloadButton > button {
            width: auto;
            min-width: 190px;
            border-radius: 4px;
            padding: 0.56rem 1.1rem;
            font-size: 1.25rem;
            font-weight: 500;
            box-shadow: none;
        }

        div.stButton > button[kind="primary"],
        div.stDownloadButton > button[kind="primary"] {
            background-color: #ff4b4b;
            border: 1px solid #d63b3b;
        }

        div.stButton > button:hover,
        div.stDownloadButton > button:hover {
            box-shadow: 0 1px 2px rgba(15, 23, 42, 0.16);
        }

        div[data-testid="stHorizontalBlock"] {
            align-items: center;
        }

        div[data-testid="stProgress"] > div > div > div {
            height: 14px;
        }

        section[data-testid="stSidebar"] div[data-testid="stSidebarContent"] {
            padding-top: 0 !important;
            padding-bottom: 190px;
        }

        section[data-testid="stSidebar"] div[data-testid="stSidebarContent"] > div:first-child,
        section[data-testid="stSidebar"] .block-container {
            padding-top: 0 !important;
        }

        section[data-testid="stSidebar"] div[data-testid="stVerticalBlock"] {
            gap: 0.65rem;
        }

        section[data-testid="stSidebar"] div[data-testid="stExpander"] details,
        section[data-testid="stSidebar"] div[data-testid="stExpander"] details > summary {
            background: #ffffff;
        }

        section[data-testid="stSidebar"] div[data-testid="stExpander"] details {
            border: 1px solid #d0d7de;
            border-radius: 6px;
        }

        .sidebar-brand-top {
            position: sticky;
            top: 0;
            z-index: 70;
            display: flex;
            justify-content: center;
            align-items: flex-start;
            height: 80px;
            overflow: hidden;
            margin: 0 0 -0.65rem;
            padding: 0;
            border-bottom: 1px solid rgba(148, 163, 184, 0.24);
            background: #f0f4f8;
        }

        .sidebar-brand-top img {
            position: relative;
            top: -35px;
            width: 150px;
            max-width: none;
            height: auto;
            display: block;
            filter: drop-shadow(0 1px 1px rgba(15, 23, 42, 0.12));
        }

        .sidebar-brand-footer {
            position: fixed;
            left: 10px;
            bottom: 34px;
            width: 318px;
            max-width: calc(100vw - 20px);
            z-index: 50;
            pointer-events: none;
            padding: 10px 10px;
            border-top: 1px solid rgba(148, 163, 184, 0.30);
            background: rgba(240, 244, 248, 0.94);
        }

        .sidebar-brand-footer img {
            width: 100%;
            height: auto;
            display: block;
            object-fit: contain;
        }

        div[data-testid="stFileUploaderDropzoneInstructions"] small,
        div[data-testid="stFileUploaderDropzoneInstructions"] > div:last-child {
            display: none;
        }

        div[data-testid="stFileUploaderFile"] {
            max-width: 100%;
        }

        div[data-testid="stFileUploaderFileName"],
        div[data-testid="stFileUploaderFile"] span {
            max-width: 100%;
            overflow: visible;
            text-overflow: clip;
            white-space: normal;
            overflow-wrap: anywhere;
            word-break: break-word;
        }

        .usage-card {
            border: 1px solid #d0d7de;
            border-radius: 6px;
            background: #ffffff;
            padding: 10px 12px 12px;
            margin-bottom: 8px;
        }

        .usage-card-label {
            color: #475569;
            font-size: 1rem;
            font-weight: 650;
            letter-spacing: 0;
            line-height: 1.2;
            margin-bottom: 6px;
        }

        .usage-card-value {
            color: #111827;
            font-size: 2.45rem;
            font-weight: 600;
            line-height: 1;
        }

        .usage-card-divider {
            border-top: 1px solid #e2e8f0;
            margin: 12px 0 9px;
        }

        .usage-card-since {
            color: #64748b;
            font-size: 0.76rem;
            font-weight: 600;
            margin-bottom: 6px;
        }

        .usage-card-row {
            display: flex;
            justify-content: space-between;
            gap: 8px;
            color: #334155;
            font-size: 0.84rem;
            line-height: 1.45;
        }

        .usage-card-row strong {
            color: #111827;
            font-weight: 700;
        }

        .knowledge-base-bar {
            border: 1px solid #d0d7de;
            border-radius: 6px;
            background: #ffffff;
            padding: 10px 12px;
            margin: 8px 0;
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 8px;
        }

        .knowledge-base-title {
            color: #111827;
            font-size: 1.05rem;
            font-weight: 600;
            line-height: 1.2;
        }

        .knowledge-base-status {
            color: #64748b;
            font-size: 0.82rem;
            font-weight: 650;
            line-height: 1;
            white-space: nowrap;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def image_file_data_uri(path: Path) -> str:
    mime_type = mimetypes.guess_type(path.name)[0] or "image/png"
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:{mime_type};base64,{encoded}"


def render_sidebar_logo() -> None:
    if not SIDEBAR_LOGO_PATH.exists():
        return
    data_uri = image_file_data_uri(SIDEBAR_LOGO_PATH)
    st.markdown(
        f"""
        <div class="sidebar-brand-top">
          <img src="{html.escape(data_uri)}" alt="Controls logo">
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_sidebar_footer() -> None:
    if not SIDEBAR_FOOTER_PATH.exists():
        return
    data_uri = image_file_data_uri(SIDEBAR_FOOTER_PATH)
    st.markdown(
        f"""
        <div class="sidebar-brand-footer">
          <img src="{html.escape(data_uri)}" alt="Toyota Production Engineering Battery">
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_feedback_support() -> None:
    with st.expander("Feedback & Support"):
        st.caption("Request a term, report an issue, or share feedback.")
        support_email = os.getenv("TERM1_SUPPORT_EMAIL", "").strip()
        if is_valid_email(support_email):
            st.caption("Email:")
            st.code(support_email, language="text")
        else:
            st.caption("Feedback email is being configured.")
        st.caption(
            "Controls leads weekly glossary validation with GBX engineering support. "
            "Approved updates are version-controlled."
        )


def render_usage_card(usage_count: int) -> None:
    text_uses = read_translation_usage_count(TEXT_TRANSLATION_USAGE_KEY)
    document_uses = read_translation_usage_count(DOCUMENT_TRANSLATION_USAGE_KEY)
    image_uses = read_translation_usage_count(IMAGE_TRANSLATION_USAGE_KEY)
    st.markdown(
        f"""
        <div class="usage-card">
          <div class="usage-card-label">App use times</div>
          <div class="usage-card-value">{usage_count:,}</div>
          <div class="usage-card-divider"></div>
          <div class="usage-card-since">Section use since {TRANSLATION_USAGE_SINCE_LABEL}</div>
          <div class="usage-card-row"><span>Text</span><strong>{text_uses:,}</strong></div>
          <div class="usage-card-row"><span>Documents</span><strong>{document_uses:,}</strong></div>
          <div class="usage-card-row"><span>Image / HMI</span><strong>{image_uses:,}</strong></div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_knowledge_base_bar(title: str, status: str = "in progress") -> None:
    st.markdown(
        f"""
        <div class="knowledge-base-bar">
          <div class="knowledge-base-title">{html.escape(title)}</div>
          <div class="knowledge-base-status">({html.escape(status)})</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_translation_result(
    translated_text: str,
    key: str = "translation_result",
    label: str = "Translation",
) -> None:
    line_count = max(translated_text.count("\n") + 1, 4)
    height = min(max(210, line_count * 28 + 92), 520)
    st.text_area(
        label,
        value=translated_text,
        height=height,
        key=key,
    )


def render_text_translation(glossary: pd.DataFrame, plc_rules: pd.DataFrame) -> None:
    st.caption("Step 1 - Select translation direction (required)")
    translation_direction = st.radio(
        "Text translation direction",
        TRANSLATION_DIRECTIONS,
        index=None,
        horizontal=True,
        key="text_translation_direction",
        help="Choose a direction before entering or translating text.",
    )
    if translation_direction is None:
        st.info("Select JP → EN or EN → JP to continue.")
        return

    source_language, target_language = direction_language_names(translation_direction)
    st.caption("Step 2 - Select translation style")
    translation_mode = st.radio(
        "Translation Style",
        TEXT_TRANSLATION_MODES,
        format_func=lambda mode: TEXT_TRANSLATION_MODE_LABELS.get(mode, mode),
        horizontal=True,
        key="text_translation_mode",
    )
    source_input = st.text_area(
        f"Input or paste {source_language} text",
        height=220,
        placeholder=f"Example: paste {source_language} manufacturing text here.",
        key=f"text_input_{translation_direction_key(translation_direction)}",
    )
    source_text, user_guidance = split_text_translation_input(source_input)
    current_text_key = f"{translation_direction}::{translation_mode}::{source_text}::{user_guidance}"

    if st.button("Translate Text", type="primary"):
        increment_usage_action("translate_text")
        if not source_text:
            st.warning(f"Please paste {source_language} text first.")
            return
        if not should_translate(source_text, translation_direction):
            st.warning(f"No translatable {source_language} text was detected.")
            return
        increment_translation_usage_count(TEXT_TRANSLATION_USAGE_KEY)

        progress = st.progress(0)
        status = st.empty()
        active_glossary = glossary_for_mode(glossary, plc_rules, translation_mode)

        try:
            status.write("Preparing glossary terms and protected codes...")
            progress.progress(0.2)
            translated_text, hits, token_usage = translate_block(
                source_text,
                active_glossary,
                translation_mode,
                user_guidance,
                translation_direction,
            )
            progress.progress(1.0)
            status.success("Translation complete.")
            if token_usage.total_tokens == 0 and hits:
                st.success("Translated by controlled rule. OpenAI API was not called.")

            st.session_state["last_text_translation_key"] = current_text_key
            st.session_state["last_text_translation"] = translated_text
            st.session_state["last_text_translation_terms"] = (
                terminology_report(hits, active_glossary, translation_direction) if hits else None
            )
            result_key = f"translation_result_{abs(hash(current_text_key))}"
            render_translation_result(translated_text, key=result_key, label=f"{target_language} Translation")

            st.download_button(
                "Download translation",
                data=translated_text.encode("utf-8-sig"),
                file_name="Translated-text.txt",
                mime="text/plain",
            )

            if hits:
                st.subheader("Detected Terminology")
                st.dataframe(
                    terminology_report(hits, active_glossary, translation_direction),
                    use_container_width=True,
                    hide_index=True,
                )
            else:
                st.info("No glossary terms were detected in this text.")
        except Exception as exc:
            status.error("Translation failed.")
            st.error(f"Translation failed: {format_translation_error(exc)}")
    elif (
        st.session_state.get("last_text_translation_key") == current_text_key
        and st.session_state.get("last_text_translation")
    ):
        result_key = f"translation_result_{abs(hash(current_text_key))}"
        render_translation_result(
            st.session_state["last_text_translation"],
            key=result_key,
            label=f"{target_language} Translation",
        )
        terms = st.session_state.get("last_text_translation_terms")
        if terms is not None:
            st.subheader("Detected Terminology")
            st.dataframe(terms, use_container_width=True, hide_index=True)
@st.fragment(run_every="15s")
def render_active_document_job(
    active_job_id: str,
    owner_session_id: str,
    glossary: pd.DataFrame,
    plc_rules: pd.DataFrame,
    translation_mode: str,
    translation_direction: str,
) -> None:
    detail = translation_job_detail(active_job_id, owner_session_id)
    if detail.empty:
        return

    active_job = detail.iloc[0].to_dict()
    active_direction = active_job.get("translation_direction") or translation_direction
    result_path_text = str(active_job.get("result_file_path") or "")
    result_path = Path(result_path_text) if result_path_text else None
    if active_job["status"] == "completed" and result_path is not None and result_path.exists():
        st.caption(active_job["file_name"])
        result_file_name = active_job["result_file_name"] or output_file_name(active_job["file_name"])
        render_download_ready(
            data=result_path.read_bytes(),
            file_name=result_file_name,
            mime=active_job["result_mime"] or mime_type(active_job["file_name"]),
            key=f"active_download_{active_job_id}",
        )
        source_path_text = str(active_job.get("source_file_path") or "")
        source_path = Path(source_path_text) if source_path_text else None
        if source_path is not None and source_path.exists():
            active_glossary = glossary_for_mode(
                glossary,
                plc_rules,
                active_job["translation_mode"] or translation_mode,
            )
            render_translation_pairs_preview(
                source_path.read_bytes(),
                active_job["file_name"],
                active_job["translation_mode"] or translation_mode,
                active_direction,
                owner_session_id,
                glossary=active_glossary,
            )
        if active_job["notify_email"]:
            st.link_button(
                "Open Email Draft",
                translation_mailto_link(active_job["notify_email"], active_job["file_name"], result_file_name),
            )
        if st.button("Dismiss", key=f"dismiss_completed_{active_job_id}"):
            increment_usage_action("dismiss_completed_document_job")
            st.session_state.pop("active_document_job_id", None)
            rerun_app()
    elif active_job["status"] == "failed":
        st.caption(active_job["file_name"])
        error_message = active_job["error_message"] or "No error detail."
        if error_message.lower().startswith("stopped by user"):
            st.warning(f"Stopped | {error_message}")
        else:
            st.error(f"Failed | {error_message}")
        action_cols = st.columns([1, 1, 2])
        if action_cols[0].button("Dismiss", key=f"dismiss_failed_{active_job_id}"):
            increment_usage_action("dismiss_failed_document_job")
            st.session_state.pop("active_document_job_id", None)
            rerun_app()
        if action_cols[1].button("Retry", key=f"retry_{active_job_id}"):
            increment_usage_action("retry_document_translation")
            source_path_text = str(active_job.get("source_file_path") or "")
            source_path = Path(source_path_text) if source_path_text else None
            if source_path is None or not source_path.exists():
                st.warning("Source file not found.")
            else:
                retry_raw = source_path.read_bytes()
                retry_blocks = extract_text_blocks(retry_raw, active_job["file_name"])
                retry_mode = active_job["translation_mode"] or translation_mode
                retry_direction = active_job.get("translation_direction") or translation_direction
                retry_glossary = glossary_for_mode(glossary, plc_rules, retry_mode)
                retry_progress_path = checkpoint_path_for(
                    active_job["file_name"],
                    retry_raw,
                    retry_mode,
                    retry_direction,
                    owner_session_id,
                )
                retry_batch_count = int(active_job["total_batches"] or 0)
                retry_job_id = start_background_translation_job(
                    retry_raw,
                    active_job["file_name"],
                    retry_blocks,
                    retry_glossary,
                    retry_mode,
                    retry_direction,
                    False,
                    active_job["notify_email"] or "",
                    retry_batch_count,
                    retry_progress_path,
                    owner_session_id,
                )
                st.session_state["active_document_job_id"] = retry_job_id
                rerun_app()
    else:
        st.caption(active_job["file_name"])
        st.info("Document is still being translated. You do not need to upload the file again.")
        stop_cols = st.columns([1, 1, 2])
        if stop_cols[0].button("Stop Translation", key=f"stop_translation_{active_job_id}"):
            increment_usage_action("stop_document_translation")
            stop_translation_job(active_job_id, owner_session_id)
            st.session_state.pop("active_document_job_id", None)
            rerun_app()
        if stop_cols[1].button("Stop All", key=f"stop_all_translation_{active_job_id}"):
            increment_usage_action("stop_all_document_translations")
            stop_all_active_translation_jobs(owner_session_id)
            st.session_state.pop("active_document_job_id", None)
            rerun_app()
        active_done = int(active_job["completed_blocks"] or 0)
        active_total = int(active_job["translatable_blocks"] or 0)
        active_done_batches = int(active_job["completed_batches"] or 0)
        active_total_batches = int(active_job["total_batches"] or 0)
        generated_pages = int(active_job.get("generated_pages") or 0)
        total_pages = int(active_job.get("total_pages") or 0)
        active_elapsed = elapsed_since_timestamp(active_job.get("created_at", ""))
        updated_elapsed = elapsed_since_timestamp(active_job.get("updated_at", ""))
        updated_at = parse_timestamp(active_job.get("updated_at", ""))
        job_is_orphaned = updated_at is not None and updated_at < APP_STARTED_AT
        active_ratio = 0.0 if active_total == 0 else min(active_done / active_total, 1.0)
        visual_ratio = active_ratio
        if active_total > 0 and active_done == 0:
            visual_ratio = 0.04
        elif active_total > 0 and active_done > 0:
            visual_ratio = max(active_ratio, 0.04)
        if total_pages > 0:
            page_ratio = min(generated_pages / total_pages, 1.0)
            st.progress(max(page_ratio, 0.04 if generated_pages == 0 else page_ratio))
            st.write(f"Generating English PDF | {100 * page_ratio:.2f}%")
        else:
            st.progress(visual_ratio)
            st.write(f"{progress_text(active_done, active_total, active_elapsed)} | {progress_percent(active_done, active_total)}")
        progress_message = str(active_job.get("progress_message") or "")
        if progress_message:
            st.caption(progress_message)
        remaining_batches = max(active_total_batches - active_done_batches, 0)
        updated_label = "unknown"
        if updated_elapsed is not None:
            updated_label = f"{format_duration(updated_elapsed)} ago"
        metric_cols = st.columns(4)
        if total_pages > 0:
            metric_cols[0].metric("Generated pages", f"{generated_pages:,}")
            metric_cols[1].metric("Total pages", f"{total_pages:,}")
            metric_cols[2].metric("Remaining pages", f"{max(total_pages - generated_pages, 0):,}")
        else:
            metric_cols[0].metric("Completed batches", f"{active_done_batches:,}")
            metric_cols[1].metric("Total batches", f"{active_total_batches:,}")
            metric_cols[2].metric("Remaining batches", f"{remaining_batches:,}")
        metric_cols[3].metric("Last update", updated_label)
        st.caption(f"Translated blocks: {active_done:,}/{active_total:,}")
        if active_total <= 0:
            st.caption(f"Reading file and calculating batches | Updated: {active_job.get('updated_at', '')}")
        else:
            st.caption(
                f"Blocks: {active_done:,}/{active_total:,} | "
                f"Batches: {active_done_batches:,}/{active_total_batches:,} | "
                f"Updated: {active_job.get('updated_at', '')}"
            )
        job_is_stalled = (updated_elapsed is not None and updated_elapsed > 300) or job_is_orphaned
        if job_is_orphaned:
            st.warning("This job was started by an older app process. Continue can attach a new background worker.")
        elif job_is_stalled:
            st.warning("No progress update for more than 5 minutes. Continue can restart from saved progress.")
        source_path_text = str(active_job.get("source_file_path") or "")
        source_path = Path(source_path_text) if source_path_text else None
        if job_is_stalled and source_path is not None and source_path.exists():
            if st.button("Continue Translation", type="primary", key=f"continue_translation_{active_job_id}"):
                increment_usage_action("continue_document_translation")
                source_raw = source_path.read_bytes()
                restart_mode = active_job["translation_mode"] or translation_mode
                restart_direction = active_job.get("translation_direction") or translation_direction
                restart_glossary = glossary_for_mode(glossary, plc_rules, restart_mode)
                update_translation_job(
                    active_job_id,
                    status="failed",
                    error_message="Stopped by user to continue with a new background worker.",
                    finished_at=time.strftime("%Y-%m-%d %H:%M:%S"),
                )
                restart_job_id = start_queued_document_translation_job(
                    source_raw,
                    active_job["file_name"],
                    restart_glossary,
                    restart_mode,
                    restart_direction,
                    False,
                    active_job["notify_email"] or "",
                    owner_session_id,
                )
                st.session_state["active_document_job_id"] = restart_job_id
                rerun_app()


def render_current_document_job(
    glossary: pd.DataFrame,
    plc_rules: pd.DataFrame,
    translation_mode: str,
    translation_direction: str,
    owner_session_id: str,
) -> bool:
    session_job_id = str(st.session_state.get("active_document_job_id") or "")
    active_job_id = session_job_id or latest_running_translation_job_id(owner_session_id)
    if not active_job_id:
        return False
    if translation_job_detail(active_job_id, owner_session_id).empty:
        st.session_state.pop("active_document_job_id", None)
        return False
    st.session_state["active_document_job_id"] = active_job_id
    st.caption("Current translation")
    render_active_document_job(
        active_job_id,
        owner_session_id,
        glossary,
        plc_rules,
        translation_mode,
        translation_direction,
    )
    return True


def image_mime_type(file_name: str) -> str:
    lower_name = file_name.lower()
    if lower_name.endswith(".png"):
        return "image/png"
    if lower_name.endswith((".jpg", ".jpeg")):
        return "image/jpeg"
    return "application/octet-stream"


def hmi_glossary_for_mode(glossary: pd.DataFrame, plc_rules: pd.DataFrame) -> pd.DataFrame:
    if plc_rules.empty:
        return glossary
    combined = pd.concat([plc_rules, glossary], ignore_index=True, sort=False).fillna("")
    combined = combined.drop_duplicates(subset=["JP"], keep="first")
    combined["term_length"] = combined["JP"].str.len()
    return combined.sort_values("term_length", ascending=False).drop(columns=["term_length"]).reset_index(drop=True)


def image_mode_translation_mode(image_mode: str) -> str:
    if image_mode == IMAGE_MODE_HMI:
        return HMI_SCREEN_TRANSLATION_MODE
    return GENERAL_TRANSLATION_MODE


def image_mode_guidance(
    image_mode: str,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> str:
    source_language, _target_language = direction_language_names(translation_direction)
    if image_mode == IMAGE_MODE_HMI:
        return (
            "Use this for machine HMI screens: buttons, menus, alarms, parameter tables, "
            "screen navigation, and short operator labels. The tool numbers visible boxes/cells "
            f"left-to-right and top-to-bottom, then translates the {source_language} inside each box."
        )
    if image_mode == IMAGE_MODE_ENGINEERING:
        return (
            "Use this for engineering drawings, process standard sheets, inspection sheets, "
            "tables, diagrams, title blocks, callouts, and equipment/process notes."
        )
    return (
        f"Use this for general photos or screenshots with {source_language} text. For formal drawings "
        "or HMI screens, choose the specific mode above for better structure."
    )


def validate_image_translation_input(
    raw: bytes,
    file_name: str,
    image_width: int,
    image_height: int,
    image_mode: str,
) -> tuple[bool, str, TokenUsage]:
    encoded = base64.b64encode(raw).decode("ascii")
    if image_mode == IMAGE_MODE_HMI:
        expected_input = (
            "an original factory HMI, PLC, machine interface, operator panel, alarm screen, "
            "parameter screen, or machine setting screen"
        )
    elif image_mode == IMAGE_MODE_ENGINEERING:
        expected_input = (
            "an original engineering drawing, CAD screenshot, inspection sheet, process sheet, "
            "standard sheet, table, diagram, or equipment/process image"
        )
    else:
        expected_input = (
            "an original photo, conceptual diagram, training image, document image, or screenshot "
            "that contains readable Japanese or English text"
        )

    prompt = f"""
You are checking whether an uploaded image contains readable Japanese or English text for translation.

Selected image mode: {image_mode}
Expected input: {expected_input}
Image size: {image_width} x {image_height}

Return only valid JSON with this exact shape:
{{
  "is_suitable": true,
  "reason": "short user-facing reason",
  "detected_type": "short image type"
}}

Rules:
1. Accept any original photo, screenshot, HMI/PLC screen, drawing, document, table, diagram, or conceptual image that contains readable Japanese or English text.
2. The selected mode is a processing preference, not a reason to reject an otherwise translatable image.
3. A conceptual diagram or general photo is suitable when it contains readable Japanese or English text.
4. Reject generated review outputs from this tool, especially images with a title like "HMI Translation Review Map", numbered yellow markers, or a side-by-side translation table.
5. Set is_suitable to false only when there is clearly no readable Japanese or English text, or when the image is a generated translation/review output.
6. Do not reject an image merely because its detected type differs from the selected mode.
""".strip()
    response = openai_client().responses.create(
        model=openai_model(),
        input=[{
            "role": "user",
            "content": [
                {"type": "input_text", "text": prompt},
                {
                    "type": "input_image",
                    "image_url": f"data:{image_mime_type(file_name)};base64,{encoded}",
                },
            ],
        }],
        temperature=0,
        timeout=openai_timeout_seconds(),
    )
    payload = extract_json_payload(response.output_text)
    is_suitable = bool(payload.get("is_suitable"))
    reason = clean_text(payload.get("reason") or "")
    detected_type = clean_text(payload.get("detected_type") or "")
    if not reason:
        reason = "This image does not match the selected image translation mode."
    if detected_type:
        reason = f"{reason} Detected type: {detected_type}."
    return is_suitable, reason, response_token_usage(response)


def extract_json_payload(text: str) -> dict:
    cleaned = str(text).strip()
    cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
    cleaned = re.sub(r"\s*```$", "", cleaned)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start >= 0 and end > start:
            return json.loads(cleaned[start:end + 1])
        raise


def clamp_region(value, low: int, high: int) -> int:
    try:
        number = int(round(float(value)))
    except (TypeError, ValueError):
        number = low
    return max(low, min(number, high))


def prepare_hmi_vision_image(raw: bytes, scale: int = 2) -> tuple[bytes, int, int, float]:
    if Image is None:
        return raw, 0, 0, 1.0
    original = Image.open(io.BytesIO(raw)).convert("RGB")
    image = original
    if scale > 1:
        image = image.resize((image.width * scale, image.height * scale), Image.Resampling.LANCZOS)
    if ImageEnhance is not None:
        image = ImageEnhance.Contrast(image).enhance(1.25)
        image = ImageEnhance.Sharpness(image).enhance(1.6)
    if ImageFilter is not None:
        image = image.filter(ImageFilter.SHARPEN)

    grid = Image.new("RGBA", image.size, (0, 0, 0, 0))
    grid_draw = ImageDraw.Draw(grid)
    grid_step = 50
    fine_step = 25
    for x in range(0, original.width + 1, fine_step):
        color = (255, 0, 0, 55) if x % grid_step else (255, 0, 0, 105)
        width = 1 if x % grid_step else 2
        grid_draw.line((x * scale, 0, x * scale, image.height), fill=color, width=width)
    for y in range(0, original.height + 1, fine_step):
        color = (255, 0, 0, 55) if y % grid_step else (255, 0, 0, 105)
        width = 1 if y % grid_step else 2
        grid_draw.line((0, y * scale, image.width, y * scale), fill=color, width=width)

    label_font = load_hmi_font(12)
    for x in range(0, original.width + 1, grid_step):
        grid_draw.text((x * scale + 2, 2), str(x), fill=(255, 255, 0, 230), font=label_font)
    for y in range(0, original.height + 1, grid_step):
        grid_draw.text((2, y * scale + 2), str(y), fill=(255, 255, 0, 230), font=label_font)
    image = Image.alpha_composite(image.convert("RGBA"), grid).convert("RGB")
    output = io.BytesIO()
    image.save(output, format="PNG")
    return output.getvalue(), image.width, image.height, float(scale)


def extract_hmi_text_regions_from_crop_with_vision(
    raw: bytes,
    file_name: str,
    image_width: int,
    image_height: int,
    crop_name: str,
    image_mode: str = IMAGE_MODE_HMI,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> tuple[list[HmiTextRegion], TokenUsage]:
    vision_raw, vision_width, vision_height, _scale = prepare_hmi_vision_image(raw)
    if not vision_width or not vision_height:
        vision_width, vision_height = image_width, image_height
    encoded = base64.b64encode(vision_raw).decode("ascii")
    source_language, _target_language = direction_language_names(translation_direction)
    if image_mode == IMAGE_MODE_HMI:
        image_description = "one cropped area of a factory HMI screenshot"
        extraction_target = f"{source_language} text inside visible HMI squares, rectangular cells, buttons, table cells, screen-name boxes, alarm/status boxes, and parameter-name boxes"
        object_description = "physical HMI square/rectangle, table cell, button, title box, navigation label box, or parameter label box"
        mode_note = (
            f"For HMI screens, first identify the visible physical boxes/cells/buttons, then read the {source_language} inside each one. "
            "Use the full box/cell/button bounds as x/y/width/height, not only the glyph bounds. "
            f"If one box contains two rows of {source_language} text, keep both rows together as one region and one translation item."
        )
    elif image_mode == IMAGE_MODE_ENGINEERING:
        image_description = "an engineering image, CAD screenshot, drawing, standard sheet, table, or process diagram"
        extraction_target = f"{source_language} headings, table cells, drawing notes, CAD callouts, diagram labels, title-block labels, and process/equipment text"
        object_description = "logical text block, table cell, callout, drawing note, title-block field, or diagram label"
        mode_note = "For drawings and CAD images, keep full related text blocks together and preserve nearby numbers, units, tolerances, model names, process numbers, and product numbers in note."
    else:
        image_description = "a general photo, conceptual diagram, training image, document image, or screenshot"
        extraction_target = f"all readable {source_language} titles, labels, captions, callouts, table cells, signs, buttons, and logical text blocks"
        object_description = "logical text block, title, label, caption, callout, table cell, sign, or button"
        mode_note = (
            "For general images, scan the full image systematically from top-left to bottom-right. "
            "Keep related multi-line text together, but do not merge separate labels or callouts."
        )

    prompt = f"""
You are extracting {source_language} text from {image_description} for engineering translation.

Return only valid JSON with this exact shape:
{{
  "regions": [
    {{
      "jp": "{source_language} text exactly as shown",
      "x": 0,
      "y": 0,
      "width": 100,
      "height": 30,
      "confidence": 0.95,
      "kind": "parameter_label",
      "note": ""
    }}
  ]
}}

Rules:
1. The uploaded reference image is a cropped HMI area enhanced for readability. Return coordinates in THIS CROP's pixel coordinate system: width {image_width}, height {image_height}, origin at top-left.
2. The reference image includes a red coordinate grid and yellow coordinate labels. Use that grid to estimate the original screenshot x/y coordinates as accurately as possible.
3. Image type: {image_mode}. Crop name: {crop_name}. Extract {extraction_target} visible inside this crop.
4. Return only regions containing the selected source language ({source_language}); do not return pure numbers, dates, times, units, or text that is only in the other language. Preserve associated numbers/units in note as engineering context.
5. Keep {source_language} text exactly as visible. Do not translate in this step.
6. Return one region per {object_description}.
7. For HMI screens, identify every visible square/rectangular label box that contains {source_language}. Do not skip a box, do not merge neighboring boxes, and do not reverse the visual order.
8. Numbering will be handled by the app after extraction. To support correct numbering, return boxes in visual reading order: left to right within the same row, then top to bottom by rows.
9. Never merge repeated labels from different cells. For example, if 予備 appears in seven cells, return seven separate regions with seven separate boxes.
10. For table headers or buttons with {source_language} split across multiple lines, group only the lines inside the same cell/button.
11. For HMI boxed labels, use the full visible square/rectangle/cell/button as the bounding box. If there is no visible box, use a tight bounding box around the logical text block.
12. Do not include the blue numeric value rows as separate regions unless the value text itself contains {source_language}. If a numeric value row belongs directly under a {source_language} parameter label, record the visible value/unit in note.
13. Pay special attention to small colored top-row buttons, right-side buttons, and bottom navigation buttons. Do not miss short labels in the selected source language.
14. Do not infer repeated labels from table structure. Return only {source_language} text you can see, at the location where it is actually visible.
15. Include confidence from 0 to 1.
16. If text is uncertain, include it with lower confidence and explain briefly in note.
17. The note field should preserve nearby engineering context when visible, such as units (mm, mm/s, msec), axis names, parameter values, or button identifiers. Keep this concise.
18. Set kind to one of: screen_title, top_button, navigation_button, action_button, parameter_label, alarm_label, status_label, value_field, unit_label, table_header, other.
19. For HMI review quality, classify bottom menu items and right-side buttons as navigation_button or action_button; classify the large centered title as screen_title; classify left/right table label cells as parameter_label; classify pure value/unit cells as value_field or unit_label.
20. {mode_note}
""".strip()
    response = openai_client().responses.create(
        model=openai_model(),
        input=[{
            "role": "user",
            "content": [
                {"type": "input_text", "text": prompt},
                {
                    "type": "input_image",
                    "image_url": f"data:{image_mime_type(file_name)};base64,{encoded}",
                    "detail": "high",
                },
            ],
        }],
        temperature=0,
        timeout=openai_timeout_seconds(),
    )
    payload = extract_json_payload(response.output_text)
    regions = []
    seen = set()
    for index, item in enumerate(payload.get("regions", []), start=1):
        jp = clean_text(item.get("jp", ""))
        if not jp or not should_translate(jp, translation_direction):
            continue
        x = clamp_region(item.get("x", 0), 0, image_width - 1)
        y = clamp_region(item.get("y", 0), 0, image_height - 1)
        width = clamp_region(item.get("width", 1), 1, image_width - x)
        height = clamp_region(item.get("height", 1), 1, image_height - y)
        try:
            confidence = max(0.0, min(float(item.get("confidence", 0.0)), 1.0))
        except (TypeError, ValueError):
            confidence = 0.0
        key = (jp, x, y, width, height)
        if key in seen:
            continue
        seen.add(key)
        regions.append(HmiTextRegion(
            location=f"{crop_name}:{index}",
            jp=jp,
            x=x,
            y=y,
            width=width,
            height=height,
            confidence=confidence,
            note=clean_text(item.get("note", "")),
            kind=clean_text(item.get("kind", "")),
        ))
    return regions, response_token_usage(response)


def hmi_detection_crops(raw: bytes, image_mode: str = IMAGE_MODE_HMI) -> list[tuple[str, int, int, bytes]]:
    image = Image.open(io.BytesIO(raw)).convert("RGB")
    width, height = image.size
    if image_mode != IMAGE_MODE_HMI:
        specs = [("full_image", 0, 0, width, height)]
        # Large diagrams and screenshots are also scanned in overlapping tiles.
        # This prevents small text from disappearing when the full image is resized
        # by the vision model while retaining the full-image pass for context.
        if width >= 1200 or height >= 900 or width * height >= 1_200_000:
            overlap_x = max(24, int(width * 0.08))
            overlap_y = max(24, int(height * 0.08))
            mid_x = width // 2
            mid_y = height // 2
            if width >= height * 1.35:
                specs.extend([
                    ("left_detail", 0, 0, min(width, mid_x + overlap_x), height),
                    ("right_detail", max(0, mid_x - overlap_x), 0, width, height),
                ])
            elif height >= width * 1.35:
                specs.extend([
                    ("top_detail", 0, 0, width, min(height, mid_y + overlap_y)),
                    ("bottom_detail", 0, max(0, mid_y - overlap_y), width, height),
                ])
            else:
                specs.extend([
                    ("top_left_detail", 0, 0, min(width, mid_x + overlap_x), min(height, mid_y + overlap_y)),
                    ("top_right_detail", max(0, mid_x - overlap_x), 0, width, min(height, mid_y + overlap_y)),
                    ("bottom_left_detail", 0, max(0, mid_y - overlap_y), min(width, mid_x + overlap_x), height),
                    ("bottom_right_detail", max(0, mid_x - overlap_x), max(0, mid_y - overlap_y), width, height),
                ])
        crops = []
        for name, x1, y1, x2, y2 in specs:
            output = io.BytesIO()
            image.crop((x1, y1, x2, y2)).save(output, format="PNG")
            crops.append((name, x1, y1, output.getvalue()))
        return crops
    specs = [
        ("title_top_bar", 0, 0, width, max(70, int(height * 0.16))),
        ("top_buttons", int(width * 0.55), 0, width, max(70, int(height * 0.16))),
        ("right_side_buttons", int(width * 0.80), 0, width, height),
        ("bottom_navigation", 0, int(height * 0.80), width, height),
        ("main_table", 0, int(height * 0.14), int(width * 0.92), int(height * 0.78)),
    ]
    crops = []
    for name, x1, y1, x2, y2 in specs:
        x1 = max(0, min(x1, width - 1))
        y1 = max(0, min(y1, height - 1))
        x2 = max(x1 + 1, min(x2, width))
        y2 = max(y1 + 1, min(y2, height))
        output = io.BytesIO()
        image.crop((x1, y1, x2, y2)).save(output, format="PNG")
        crops.append((name, x1, y1, output.getvalue()))
    return crops


def hmi_regions_overlap(a: HmiTextRegion, b: HmiTextRegion) -> bool:
    ax2, ay2 = a.x + a.width, a.y + a.height
    bx2, by2 = b.x + b.width, b.y + b.height
    overlap_w = max(0, min(ax2, bx2) - max(a.x, b.x))
    overlap_h = max(0, min(ay2, by2) - max(a.y, b.y))
    if overlap_w <= 0 or overlap_h <= 0:
        return False
    overlap_area = overlap_w * overlap_h
    smaller_area = max(min(a.width * a.height, b.width * b.height), 1)
    return overlap_area / smaller_area > 0.45


def hmi_regions_are_duplicate(a: HmiTextRegion, b: HmiTextRegion) -> bool:
    a_jp = re.sub(r"[\s\u3000]+", "", unicodedata.normalize("NFKC", clean_text(a.jp))).casefold()
    b_jp = re.sub(r"[\s\u3000]+", "", unicodedata.normalize("NFKC", clean_text(b.jp))).casefold()
    if a_jp != b_jp:
        return False
    if hmi_regions_overlap(a, b):
        return True
    a_center_x = a.x + a.width / 2
    a_center_y = a.y + a.height / 2
    b_center_x = b.x + b.width / 2
    b_center_y = b.y + b.height / 2
    x_tolerance = max(a.width, b.width, 24) * 1.35
    y_tolerance = max(a.height, b.height, 16) * 1.8
    if abs(a_center_x - b_center_x) <= x_tolerance and abs(a_center_y - b_center_y) <= y_tolerance:
        return True
    return False


def grouped_line_centers(indices: list[int], max_gap: int = 2) -> list[int]:
    if not indices:
        return []
    groups: list[list[int]] = [[indices[0]]]
    for value in indices[1:]:
        if value - groups[-1][-1] <= max_gap:
            groups[-1].append(value)
        else:
            groups.append([value])
    return [int(round(sum(group) / len(group))) for group in groups]


def longest_true_runs(mask, axis: int) -> "np.ndarray":
    if np is None:
        return []
    lines = mask.T if axis == 0 else mask
    scores = []
    for line in lines:
        best = 0
        current = 0
        for value in line:
            if value:
                current += 1
                if current > best:
                    best = current
            else:
                current = 0
        scores.append(best)
    return np.asarray(scores)


def box_iou(a: HmiDetectedBox, b: HmiDetectedBox) -> float:
    ax2, ay2 = a.x + a.width, a.y + a.height
    bx2, by2 = b.x + b.width, b.y + b.height
    overlap_w = max(0, min(ax2, bx2) - max(a.x, b.x))
    overlap_h = max(0, min(ay2, by2) - max(a.y, b.y))
    if overlap_w <= 0 or overlap_h <= 0:
        return 0.0
    intersection = overlap_w * overlap_h
    union = max(a.width * a.height + b.width * b.height - intersection, 1)
    return intersection / union


def order_hmi_boxes_row_major(boxes: list[HmiDetectedBox]) -> list[HmiDetectedBox]:
    if not boxes:
        return []
    row_groups: list[dict[str, object]] = []
    for box in sorted(boxes, key=lambda item: (item.y + item.height / 2, item.x)):
        center_y = box.y + box.height / 2
        matched_group = None
        tolerance = max(10, min(32, box.height * 0.75))
        for group in row_groups:
            if abs(center_y - float(group["center_y"])) <= tolerance:
                matched_group = group
                break
        if matched_group is None:
            row_groups.append({"center_y": center_y, "boxes": [box]})
            continue
        group_boxes = matched_group["boxes"]
        group_boxes.append(box)
        matched_group["center_y"] = (
            float(matched_group["center_y"]) * (len(group_boxes) - 1) + center_y
        ) / len(group_boxes)

    ordered: list[HmiDetectedBox] = []
    for group in sorted(row_groups, key=lambda item: float(item["center_y"])):
        ordered.extend(sorted(group["boxes"], key=lambda item: item.x))
    return [
        HmiDetectedBox(index, box.x, box.y, box.width, box.height)
        for index, box in enumerate(ordered, start=1)
    ]


def detect_hmi_candidate_boxes(raw: bytes) -> list[HmiDetectedBox]:
    if Image is None or np is None:
        return []
    try:
        image = Image.open(io.BytesIO(raw)).convert("RGB")
    except Exception:
        return []

    gray = np.asarray(image.convert("L")).astype("int16")
    image_height, image_width = gray.shape
    if image_width < 40 or image_height < 40:
        return []

    vertical_edge = np.zeros((image_height, image_width), dtype=bool)
    horizontal_edge = np.zeros((image_height, image_width), dtype=bool)
    vertical_edge[:, 1:] = np.abs(np.diff(gray, axis=1)) > 70
    horizontal_edge[1:, :] = np.abs(np.diff(gray, axis=0)) > 70
    horizontal_scores = longest_true_runs(horizontal_edge, axis=1)
    vertical_scores = longest_true_runs(vertical_edge, axis=0)
    h_threshold = max(30, int(image_width * 0.06))
    v_threshold = max(30, int(image_height * 0.09))
    h_lines = grouped_line_centers(np.where(horizontal_scores >= h_threshold)[0].tolist(), max_gap=3)
    v_lines = grouped_line_centers(np.where(vertical_scores >= v_threshold)[0].tolist(), max_gap=3)
    h_lines = [y for y in h_lines if 0 <= y < image_height]
    v_lines = [x for x in v_lines if 0 <= x < image_width]

    boxes: list[HmiDetectedBox] = []
    for top, bottom in zip(h_lines, h_lines[1:]):
        height = bottom - top
        if height < 14 or height > min(90, image_height * 0.35):
            continue
        for left, right in zip(v_lines, v_lines[1:]):
            width = right - left
            if width < 18 or width > min(260, image_width * 0.55):
                continue

            inner_pad = 1
            y_slice = slice(max(0, top - 1), min(image_height, bottom + 2))
            x_slice = slice(max(0, left - 1), min(image_width, right + 2))
            top_band = horizontal_edge[max(0, top - 1):min(image_height, top + 2), x_slice]
            bottom_band = horizontal_edge[max(0, bottom - 1):min(image_height, bottom + 2), x_slice]
            left_band = vertical_edge[y_slice, max(0, left - inner_pad):min(image_width, left + inner_pad + 1)]
            right_band = vertical_edge[y_slice, max(0, right - inner_pad):min(image_width, right + inner_pad + 1)]
            top_score = top_band.mean() if top_band.size else 0
            bottom_score = bottom_band.mean() if bottom_band.size else 0
            left_score = left_band.mean() if left_band.size else 0
            right_score = right_band.mean() if right_band.size else 0
            if min(top_score, bottom_score, left_score, right_score) < 0.08:
                continue

            candidate = HmiDetectedBox(0, left, top, width, height)
            if any(box_iou(candidate, existing) > 0.72 for existing in boxes):
                continue
            boxes.append(candidate)

    # Prefer the smallest valid cells/buttons; remove boxes that mainly contain smaller boxes.
    boxes = sorted(boxes, key=lambda item: item.width * item.height)
    kept: list[HmiDetectedBox] = []
    for box in boxes:
        contains_smaller = False
        for existing in kept:
            if (
                existing.x >= box.x
                and existing.y >= box.y
                and existing.x + existing.width <= box.x + box.width
                and existing.y + existing.height <= box.y + box.height
                and existing.width * existing.height < box.width * box.height * 0.72
            ):
                contains_smaller = True
                break
        if not contains_smaller:
            kept.append(box)

    return order_hmi_boxes_row_major(kept[:220])


def render_hmi_box_number_overlay(raw: bytes, boxes: list[HmiDetectedBox]) -> bytes:
    image = Image.open(io.BytesIO(raw)).convert("RGB")
    draw = ImageDraw.Draw(image, "RGBA")
    marker_font = load_hmi_font(13)
    for box in boxes:
        draw.rectangle(
            (box.x, box.y, box.x + box.width, box.y + box.height),
            outline=(255, 230, 0, 220),
            width=2,
        )
        label = str(box.no)
        bbox = draw.textbbox((0, 0), label, font=marker_font)
        label_w = bbox[2] - bbox[0]
        label_h = bbox[3] - bbox[1]
        pad_x, pad_y = 4, 2
        label_x = max(0, min(box.x + 2, image.width - label_w - pad_x * 2))
        label_y = max(0, min(box.y + 2, image.height - label_h - pad_y * 2))
        draw.rectangle(
            (label_x, label_y, label_x + label_w + pad_x * 2, label_y + label_h + pad_y * 2),
            fill=(255, 230, 0, 245),
            outline=(175, 40, 35, 245),
            width=1,
        )
        draw.text((label_x + pad_x, label_y + pad_y - 1), label, fill=(0, 0, 0, 255), font=marker_font)

    output = io.BytesIO()
    image.save(output, format="PNG")
    return output.getvalue()


def hmi_box_crop_image(raw_image, box: HmiDetectedBox):
    pad = 3
    left = max(0, box.x - pad)
    top = max(0, box.y - pad)
    right = min(raw_image.width, box.x + box.width + pad)
    bottom = min(raw_image.height, box.y + box.height + pad)
    crop = raw_image.crop((left, top, right, bottom)).convert("RGB")
    min_width = 180
    if crop.width < min_width:
        scale = min(4.0, max(1.0, min_width / max(crop.width, 1)))
        crop = crop.resize((int(crop.width * scale), int(crop.height * scale)), Image.Resampling.LANCZOS)
    if ImageEnhance is not None:
        crop = ImageEnhance.Contrast(crop).enhance(1.18)
        crop = ImageEnhance.Sharpness(crop).enhance(1.35)
    return crop


def image_to_png_data_url(image) -> str:
    output = io.BytesIO()
    image.save(output, format="PNG")
    encoded = base64.b64encode(output.getvalue()).decode("ascii")
    return f"data:image/png;base64,{encoded}"


def extract_hmi_box_regions_with_vision(
    raw: bytes,
    file_name: str,
    boxes: list[HmiDetectedBox],
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
    progress_callback=None,
) -> tuple[list[HmiTextRegion], TokenUsage]:
    if not boxes:
        return [], TokenUsage()

    original = Image.open(io.BytesIO(raw)).convert("RGB")
    box_by_no = {box.no: box for box in boxes}
    regions: list[HmiTextRegion] = []
    seen: set[int] = set()
    total_usage = TokenUsage()
    boxes = boxes[:HMI_MAX_VISION_BOXES]
    batch_size = HMI_BOX_VISION_BATCH_SIZE
    source_language, _target_language = direction_language_names(translation_direction)

    for start in range(0, len(boxes), batch_size):
        batch = boxes[start:start + batch_size]
        if progress_callback is not None:
            batch_no = start // batch_size + 1
            total_batches = max(1, (len(boxes) + batch_size - 1) // batch_size)
            progress_callback("boxes", batch_no, total_batches)
        prompt = f"""
You are reading clean cropped cells from a factory HMI/PLC screen for engineering translation.

Each image below is one detected physical HMI box/cell/button cropped from the original screenshot. The text label before each image gives the stable box number. Use these box numbers exactly.

Return only valid JSON with this exact shape:
{{
  "items": [
    {{
      "box_no": 1,
      "jp": "{source_language} text exactly as shown inside this box",
      "confidence": 0.95,
      "kind": "parameter_label",
      "note": ""
    }}
  ]
}}

Rules:
1. Read every supplied crop carefully. This batch contains box numbers: {", ".join(str(box.no) for box in batch)}.
2. Return every box that contains {source_language} text, including small buttons, top-row labels, bottom menu labels, and two-line labels.
3. Do not return pure numbers, dates, times, units, or labels written only in the other language.
4. If a box contains {source_language} plus numbers/units, keep the source text in jp and preserve nearby numbers/units in note.
5. Keep {source_language} text exactly as visible. Do not translate in this step.
6. Do not merge neighboring boxes. One crop equals one item.
7. If text is uncertain, still return it with lower confidence and explain briefly in note.
8. Set kind to one of: screen_title, top_button, navigation_button, action_button, parameter_label, alarm_label, status_label, value_field, unit_label, table_header, other.
9. Use screen_title for the main page title, top_button for small top-row mode buttons, navigation_button for page/menu movement, action_button for register/cancel/call/start buttons, parameter_label for setting names, unit_label for unit-only cells, and value_field for numeric/value-only fields.
""".strip()
        content = [{"type": "input_text", "text": prompt}]
        for box in batch:
            crop = hmi_box_crop_image(original, box)
            content.append({"type": "input_text", "text": f"Box {box.no}"})
            content.append({
                "type": "input_image",
                "image_url": image_to_png_data_url(crop),
                "detail": "high",
            })

        response = openai_client().responses.create(
            model=openai_model(),
            input=[{"role": "user", "content": content}],
            temperature=0,
            timeout=openai_timeout_seconds(),
        )
        total_usage.add(response_token_usage(response))
        payload = extract_json_payload(response.output_text)
        for item in payload.get("items", []):
            try:
                box_no = int(item.get("box_no", 0))
            except (TypeError, ValueError):
                continue
            if box_no in seen or box_no not in box_by_no:
                continue
            jp = clean_text(item.get("jp", ""))
            if not jp or not should_translate(jp, translation_direction):
                continue
            try:
                confidence = max(0.0, min(float(item.get("confidence", 0.0)), 1.0))
            except (TypeError, ValueError):
                confidence = 0.0
            box = box_by_no[box_no]
            regions.append(HmiTextRegion(
                location=f"hmi_box:{box_no}",
                jp=jp,
                x=box.x,
                y=box.y,
                width=box.width,
                height=box.height,
                confidence=confidence,
                note=clean_text(item.get("note", "")),
                kind=clean_text(item.get("kind", "")),
            ))
            seen.add(box_no)
    return order_hmi_regions_row_major(regions), total_usage


def nearest_hmi_box_for_region(image, region: HmiTextRegion) -> tuple[int, int, int, int] | None:
    if np is None or Image is None:
        return None
    try:
        gray = np.asarray(image.convert("L"))
    except Exception:
        return None
    image_height, image_width = gray.shape
    if image_width <= 1 or image_height <= 1:
        return None

    center_x = region.x + region.width / 2
    center_y = region.y + region.height / 2
    margin_y = max(10, int(region.height * 0.9))
    y1 = max(0, region.y - margin_y)
    y2 = min(image_height, region.y + region.height + margin_y)
    if y2 <= y1:
        return None

    bright = gray > 105
    vertical_scores = bright[y1:y2, :].sum(axis=0)
    vertical_threshold = max(5, int((y2 - y1) * 0.42))
    vertical_indices = np.where(vertical_scores >= vertical_threshold)[0].tolist()
    vertical_centers = grouped_line_centers(vertical_indices)
    left_candidates = [x for x in vertical_centers if x < center_x - 3]
    right_candidates = [x for x in vertical_centers if x > center_x + 3]
    if not left_candidates or not right_candidates:
        return None

    left = max(left_candidates)
    right = min(right_candidates)
    if right - left < max(region.width + 4, 12):
        return None

    x1_band = max(0, left)
    x2_band = min(image_width, right + 1)
    horizontal_scores = bright[:, x1_band:x2_band].sum(axis=1)
    horizontal_threshold = max(5, int((x2_band - x1_band) * 0.35))
    horizontal_indices = np.where(horizontal_scores >= horizontal_threshold)[0].tolist()
    horizontal_centers = grouped_line_centers(horizontal_indices)
    top_candidates = [y for y in horizontal_centers if y < center_y - 3]
    bottom_candidates = [y for y in horizontal_centers if y > center_y + 3]
    if not top_candidates or not bottom_candidates:
        return None

    top = max(top_candidates)
    bottom = min(bottom_candidates)
    if bottom - top < max(region.height + 4, 12):
        return None

    return (
        clamp_region(left, 0, image_width - 1),
        clamp_region(top, 0, image_height - 1),
        clamp_region(right - left, 1, image_width - left),
        clamp_region(bottom - top, 1, image_height - top),
    )


def snap_hmi_regions_to_detected_boxes(raw: bytes, regions: list[HmiTextRegion]) -> list[HmiTextRegion]:
    if not regions or Image is None:
        return regions
    try:
        image = Image.open(io.BytesIO(raw)).convert("RGB")
    except Exception:
        return regions

    snapped: list[HmiTextRegion] = []
    for region in regions:
        box = nearest_hmi_box_for_region(image, region)
        if box is None:
            snapped.append(region)
            continue
        x, y, width, height = box
        note = region.note
        if "box-snapped" not in note:
            note = f"{note}; box-snapped".strip("; ")
        snapped.append(HmiTextRegion(
            location=region.location,
            jp=region.jp,
            x=x,
            y=y,
            width=width,
            height=height,
            confidence=region.confidence,
            note=note,
            kind=region.kind,
        ))
    return snapped


def merge_hmi_regions(regions: list[HmiTextRegion]) -> list[HmiTextRegion]:
    merged: list[HmiTextRegion] = []
    for region in sorted(regions, key=lambda item: (item.y, item.x, -item.confidence)):
        duplicate_index = None
        for index, existing in enumerate(merged):
            if hmi_regions_are_duplicate(existing, region):
                duplicate_index = index
                break
        if duplicate_index is None:
            merged.append(region)
        elif region.confidence > merged[duplicate_index].confidence:
            merged[duplicate_index] = region
    return [
        HmiTextRegion(
            location=f"hmi:{index}",
            jp=region.jp,
            x=region.x,
            y=region.y,
            width=region.width,
            height=region.height,
            confidence=region.confidence,
            note=region.note,
            kind=region.kind,
        )
        for index, region in enumerate(merged, start=1)
    ]


def extract_hmi_text_regions_with_vision(
    raw: bytes,
    file_name: str,
    image_width: int,
    image_height: int,
    image_mode: str = IMAGE_MODE_HMI,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
    progress_callback=None,
) -> tuple[list[HmiTextRegion], TokenUsage]:
    box_regions: list[HmiTextRegion] = []
    box_usage = TokenUsage()
    if image_mode == IMAGE_MODE_HMI:
        boxes = detect_hmi_candidate_boxes(raw)
        if len(boxes) >= 6:
            try:
                box_regions, box_usage = extract_hmi_box_regions_with_vision(
                    raw,
                    file_name,
                    boxes,
                    translation_direction,
                    progress_callback,
                )
            except Exception:
                box_regions = []
                box_usage = TokenUsage()

    all_regions: list[HmiTextRegion] = []
    total_usage = TokenUsage()
    total_usage.add(box_usage)
    all_regions.extend(box_regions)
    crop_errors: list[Exception] = []
    detection_crops = hmi_detection_crops(raw, image_mode)
    if image_mode == IMAGE_MODE_HMI and box_regions:
        # Box extraction already covers the main table. Scan only the areas most
        # likely to contain labels that do not have a complete rectangular border.
        detection_crops = [
            crop for crop in detection_crops
            if crop[0] in {"title_top_bar", "bottom_navigation"}
        ]
    for crop_index, (crop_name, offset_x, offset_y, crop_raw) in enumerate(detection_crops, start=1):
        if progress_callback is not None:
            progress_callback("crops", crop_index, max(1, len(detection_crops)))
        crop_image = Image.open(io.BytesIO(crop_raw)).convert("RGB")
        crop_regions: list[HmiTextRegion] = []
        crop_usage = TokenUsage()
        for attempt in range(1, MAX_TRANSLATION_RETRIES + 1):
            try:
                crop_regions, crop_usage = extract_hmi_text_regions_from_crop_with_vision(
                    crop_raw,
                    file_name,
                    crop_image.width,
                    crop_image.height,
                    crop_name,
                    image_mode,
                    translation_direction,
                )
                break
            except Exception as exc:
                if attempt == MAX_TRANSLATION_RETRIES:
                    crop_errors.append(exc)
                    break
                time.sleep(2 * attempt)
        total_usage.add(crop_usage)
        for region in crop_regions:
            all_regions.append(HmiTextRegion(
                location=region.location,
                jp=region.jp,
                x=clamp_region(region.x + offset_x, 0, image_width - 1),
                y=clamp_region(region.y + offset_y, 0, image_height - 1),
                width=clamp_region(region.width, 1, image_width - (region.x + offset_x)),
                height=clamp_region(region.height, 1, image_height - (region.y + offset_y)),
                confidence=region.confidence,
                note=f"{crop_name}: {region.note}".strip(": "),
                kind=region.kind,
            ))
    if not all_regions and crop_errors:
        raise crop_errors[-1]
    if image_mode == IMAGE_MODE_HMI:
        supplemental_regions = snap_hmi_regions_to_detected_boxes(raw, [
            region for region in all_regions if not region.location.startswith("hmi_box:")
        ])
        merged_regions = list(box_regions)
        for region in supplemental_regions:
            if any(hmi_regions_are_duplicate(existing, region) for existing in merged_regions):
                continue
            merged_regions.append(region)
        merged_regions = order_hmi_regions_row_major(merged_regions)
    else:
        merged_regions = merge_hmi_regions(all_regions)
    return merged_regions, total_usage


def build_hmi_translation_prompt(
    items: list[tuple[int, HmiTextRegion, str, list[TermHit], list[str]]],
    screen_context: str,
    image_mode: str,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> str:
    source_language, target_language = direction_language_names(translation_direction)

    def required_terms_for_block(hits: list[TermHit]) -> str:
        if not hits:
            return "None"
        if translation_direction == TRANSLATION_DIRECTION_EN_JP:
            return "; ".join(f"{hit.en} -> {hit.jp}" for hit in hits)
        return "; ".join(f"{hit.jp} -> {hit.en}" for hit in hits)

    item_text = "\n\n".join(
        "\n".join([
            f"[BLOCK {item_id}]",
            f"No: {item_id}",
            f"{source_language}: {source_text}",
            f"Required controlled terminology: {required_terms_for_block(hits)}",
            f"Position: x={region.x}, y={region.y}, width={region.width}, height={region.height}",
            f"Nearby context: {region.note or 'None'}",
            f"[/BLOCK {item_id}]",
        ])
        for item_id, region, source_text, hits, _ in items
    )
    terms = []
    codes = []
    for _, _, _, hits, protected_codes in items:
        terms.extend(f"{hit.jp} = {hit.en}" for hit in hits)
        codes.extend(protected_codes)
    unique_terms = "\n".join(sorted(set(terms))) or "None"
    unique_codes = ", ".join(sorted(set(codes))) if codes else "None detected"
    mode_rules = mode_rules_for(image_mode_translation_mode(image_mode), translation_direction)

    return f"""
You are translating {source_language} text detected from one factory HMI/PLC/engineering image into {target_language}.

Image mode: {image_mode}
Translation mode: {image_mode_translation_mode(image_mode)}
Translation direction: {translation_direction}

{mode_rules}

Screen context from the same image:
{screen_context}

Company terminology detected:
{unique_terms}

Protected codes detected:
{unique_codes}

Critical engineering-image translation rules:
1. Translate each block using the screen context, position, and nearby context. Do not translate each item as an isolated dictionary word.
2. Preserve the exact meaning first. Keep output concise only after the meaning is correct.
3. Use concise {target_language} HMI/operator wording for buttons, screen names, alarm labels, settings, or parameters.
4. Do not invent status, cause, action, or equipment details not visible in the source.
5. Do not copy {source_language} prose into the {target_language} output unless it is a proper name, code, model, or unavoidable label.
6. Preserve numbers, units, PLC/HMI codes, arrows, punctuation, and separators when they are part of the source text.
7. Use consistent {target_language} for repeated source labels across the same image.
8. If a block is a navigation or action button, use short standard UI wording in {target_language} when accurate.
9. Required controlled terminology is mandatory. When a block lists a required term, use its target wording exactly; do not replace it with a synonym or paraphrase.
10. Return each translated block using the same markers and do not add explanations:
[BLOCK 1]
{target_language} translation
[/BLOCK 1]
11. Preserve text already written in {target_language}; translate only the {source_language} content.

Source blocks:
{item_text}
""".strip()


def hmi_screen_context(regions: list[HmiTextRegion], limit: int = 80) -> str:
    ordered = order_hmi_regions_row_major(regions)
    lines = []
    for index, region in enumerate(ordered[:limit], start=1):
        note = f" | note: {region.note}" if region.note else ""
        lines.append(f"{index}. {region.jp}{note}")
    if len(ordered) > limit:
        lines.append(f"... {len(ordered) - limit} more item(s)")
    return "\n".join(lines) or "None"


def normalize_hmi_english(text: str) -> str:
    normalized = clean_text(text)
    normalized = re.sub(r"\b(Data|Unit)(\d+)\b", r"\1 \2", normalized)
    normalized = re.sub(r"\b(Data|Unit)\s+(\d+)\s+(Data|Unit)\s+(\d+)\b", r"\1 \2 / \3 \4", normalized)
    return normalized


def normalize_hmi_translation(text: str, translation_direction: str) -> str:
    if translation_direction == TRANSLATION_DIRECTION_EN_JP:
        return clean_text(text)
    return normalize_hmi_english(text)


def normalize_hmi_region_kind(kind: str) -> str:
    compact = re.sub(r"[^a-z_]+", "", clean_text(kind).lower().replace("-", "_").replace(" ", "_"))
    allowed = {
        "screen_title",
        "top_button",
        "navigation_button",
        "action_button",
        "parameter_label",
        "alarm_label",
        "status_label",
        "value_field",
        "unit_label",
        "table_header",
        "other",
    }
    return compact if compact in allowed else ""


def infer_hmi_region_kind(region: HmiTextRegion, translation: str, image_height: int, image_width: int = 0) -> str:
    existing = normalize_hmi_region_kind(region.kind)
    if existing:
        return existing

    jp = re.sub(r"\s+", "", clean_text(region.jp))
    english = clean_text(translation).lower()
    compact_en = re.sub(r"[^a-z0-9]+", "", english)
    bottom_navigation = image_height > 0 and region.y >= image_height * 0.80
    top_area = image_height > 0 and region.y <= image_height * 0.14
    right_edge = image_width > 0 and region.x >= image_width * 0.72

    if bottom_navigation:
        return "navigation_button"
    if top_area and any(term in jp for term in ["応援", "職制", "英文", "ハードコピー"]):
        return "top_button"
    if any(term in jp for term in ["画面", "前頁", "選択"]):
        return "navigation_button"
    if any(term in jp for term in ["登録", "キャンセル", "呼出", "開始"]):
        return "action_button"
    if any(term in jp for term in ["異常", "警報", "アラーム"]):
        return "alarm_label"
    if any(term in jp for term in ["運転", "停止", "完了", "中"]):
        return "status_label"
    if jp.startswith("単位") or compact_en.startswith("unit"):
        return "unit_label"
    if jp.startswith(("データ", "ﾃﾞｰﾀ", "デ－タ")) or compact_en.startswith("data"):
        return "value_field"
    if top_area or right_edge:
        return "action_button"
    return "parameter_label"


def translate_hmi_regions(
    regions: list[HmiTextRegion],
    glossary: pd.DataFrame,
    image_mode: str = IMAGE_MODE_HMI,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> tuple[dict[str, str], list[TermHit], TokenUsage]:
    if not regions:
        return {}, [], TokenUsage()

    translation_mode = image_mode_translation_mode(image_mode)
    translations: dict[str, str] = {}
    all_hits: list[TermHit] = []
    token_usage = TokenUsage()
    context = hmi_screen_context(regions)
    client = openai_client()
    chunk_size = 24

    for start in range(0, len(regions), chunk_size):
        chunk = regions[start:start + chunk_size]
        items = []
        item_regions: dict[int, HmiTextRegion] = {}
        direct_translations: dict[str, str] = {}

        for item_id, region in enumerate(chunk, start=1):
            exact_translation, exact_hits = exact_controlled_term_match(
                region.jp,
                glossary,
                translation_direction,
            )
            if exact_translation is not None:
                translation = post_process_translation(
                    exact_translation,
                    translation_mode,
                    translation_direction,
                )
                direct_translations[region.location] = normalize_hmi_translation(
                    restore_missing_enclosed_markers(region.jp, translation),
                    translation_direction,
                )
                all_hits.extend(exact_hits)
                continue

            glossary_applied_text, hits = apply_glossary_to_source(
                region.jp,
                glossary,
                replace_source=False,
                translation_direction=translation_direction,
            )
            protected_codes = find_protected_codes(region.jp)
            items.append((item_id, region, glossary_applied_text, hits, protected_codes))
            item_regions[item_id] = region
            all_hits.extend(hits)

        translations.update(direct_translations)
        if not items:
            continue

        parsed: dict[int, str] = {}
        for attempt in range(1, MAX_TRANSLATION_RETRIES + 1):
            try:
                response = client.responses.create(
                    model=openai_model(),
                    input=build_hmi_translation_prompt(
                        items,
                        context,
                        image_mode,
                        translation_direction,
                    ),
                    temperature=0,
                    timeout=openai_timeout_seconds(),
                )
                token_usage.add(response_token_usage(response))
                parsed = parse_batch_translation(response.output_text.strip(), [item[0] for item in items])
                missing_ids = [item[0] for item in items if item[0] not in parsed]
                if missing_ids:
                    raise ValueError(f"HMI translation response missed block marker(s): {missing_ids}")
                break
            except Exception:
                if attempt == MAX_TRANSLATION_RETRIES:
                    raise
                time.sleep(4 * attempt)

        for item_id, translated in parsed.items():
            region = item_regions[item_id]
            translation = post_process_translation(
                translated,
                translation_mode,
                translation_direction,
            )
            translations[region.location] = normalize_hmi_translation(
                restore_missing_enclosed_markers(region.jp, translation)
                , translation_direction
            )

    return translations, all_hits, token_usage


def hmi_regions_are_translated_duplicates(
    a: HmiTextRegion,
    b: HmiTextRegion,
    translations: dict[str, str],
) -> bool:
    a_english = clean_text(translations.get(a.location, ""))
    b_english = clean_text(translations.get(b.location, ""))
    if not a_english or a_english != b_english:
        return False
    if a_english.lower() in {"spare", "reserved"}:
        return False
    if hmi_regions_are_duplicate(a, b):
        return True
    a_center_x = a.x + a.width / 2
    a_center_y = a.y + a.height / 2
    b_center_x = b.x + b.width / 2
    b_center_y = b.y + b.height / 2
    same_hmi_row = abs(a_center_y - b_center_y) <= max(a.height, b.height, 24) * 1.8
    nearby_button = abs(a_center_x - b_center_x) <= max(a.width, b.width, 32) * 1.8
    return same_hmi_row and nearby_button


def order_hmi_regions_row_major(regions: list[HmiTextRegion]) -> list[HmiTextRegion]:
    row_groups: list[dict[str, object]] = []
    for region in sorted(regions, key=lambda item: (item.y + item.height / 2, item.x)):
        center_y = region.y + region.height / 2
        matched_group = None
        for group in row_groups:
            if abs(center_y - float(group["center_y"])) <= 28:
                matched_group = group
                break
        if matched_group is None:
            row_groups.append({"center_y": center_y, "regions": [region]})
            continue
        group_regions = matched_group["regions"]
        group_regions.append(region)
        matched_group["center_y"] = (
            float(matched_group["center_y"]) * (len(group_regions) - 1) + center_y
        ) / len(group_regions)

    ordered: list[HmiTextRegion] = []
    for group in sorted(row_groups, key=lambda item: float(item["center_y"])):
        ordered.extend(sorted(group["regions"], key=lambda item: item.x))
    return ordered


def deduplicate_hmi_regions_for_output(
    regions: list[HmiTextRegion],
    translations: dict[str, str],
) -> list[HmiTextRegion]:
    deduped: list[HmiTextRegion] = []
    for region in sorted(regions, key=lambda item: (item.y, item.x, -item.confidence)):
        duplicate_index = None
        for index, existing in enumerate(deduped):
            if hmi_regions_are_duplicate(existing, region) or hmi_regions_are_translated_duplicates(
                existing,
                region,
                translations,
            ):
                duplicate_index = index
                break
        if duplicate_index is None:
            deduped.append(region)
        elif region.confidence > deduped[duplicate_index].confidence:
            deduped[duplicate_index] = region
    return order_hmi_regions_row_major(deduped)


def is_low_value_hmi_placeholder(region: HmiTextRegion, translation: str) -> bool:
    jp = clean_text(region.jp)
    english = clean_text(translation)
    compact_jp = re.sub(r"\s+", "", jp)
    compact_en = re.sub(r"[\s/]+", "", english).lower()

    if re.fullmatch(r"(?:(?:データ|ﾃﾞｰﾀ|デ－タ)\d+)+", compact_jp):
        return True
    if re.fullmatch(r"(?:単位\d+)+", compact_jp):
        return True
    if re.fullmatch(r"(?:data\d+)+", compact_en):
        return True
    if re.fullmatch(r"(?:unit\d+)+", compact_en):
        return True
    if compact_jp in {"装置", "号機", "番号", "予備", "スペア"}:
        return True
    if compact_en in {"equipment", "machinenumber", "equipmentnumber", "unitnumber", "spare", "reserved"}:
        return True
    return False


def hmi_review_label_key(region: HmiTextRegion, translation: str) -> str:
    jp_key = re.sub(r"\s+", "", clean_text(region.jp))
    en_key = re.sub(r"[^a-z0-9]+", "", clean_text(translation).lower())
    return f"{jp_key}|{en_key}"


def hmi_review_semantic_key(region: HmiTextRegion, translation: str) -> str:
    english = clean_text(translation).lower()
    compact_en = re.sub(r"[^a-z0-9]+", "", english)
    compact_jp = re.sub(r"\s+", "", clean_text(region.jp))
    if compact_en and not re.fullmatch(r"(?:data|unit)\d+", compact_en):
        return compact_en
    return compact_jp


def is_important_hmi_region(region: HmiTextRegion, translation: str, image_height: int) -> bool:
    jp = clean_text(region.jp)
    english = clean_text(translation)
    compact_jp = re.sub(r"\s+", "", jp)
    compact_en = re.sub(r"\s+", "", english).lower()
    kind = infer_hmi_region_kind(region, english, image_height)

    if not jp or is_low_value_hmi_placeholder(region, english):
        return False

    if kind in {"value_field", "unit_label", "other"}:
        return False
    if kind in {
        "screen_title",
        "top_button",
        "navigation_button",
        "action_button",
        "alarm_label",
        "status_label",
        "table_header",
    }:
        return True

    important_jp_patterns = [
        "画面", "呼出", "登録", "キャンセル", "前頁", "英文", "開始",
        "番号", "設定", "操作", "異常", "情報", "管理", "モニタ",
        "メイン", "タクト", "サーボ", "制限", "位置", "負荷", "生産",
        "ライン", "装置", "主軸", "実効", "ワーク",
    ]
    if any(pattern in compact_jp for pattern in important_jp_patterns):
        return True

    important_en_patterns = [
        "screen", "call", "register", "cancel", "previous", "start",
        "number", "setting", "operation", "alarm", "fault", "info",
        "management", "monitor", "takt", "servo", "limit", "position",
        "load", "production", "line", "equipment", "workpiece",
    ]
    if any(pattern in compact_en for pattern in important_en_patterns):
        return True

    if kind == "parameter_label":
        parameter_terms = [
            "番号", "設定", "制限", "位置", "負荷", "生産", "ライン",
            "主軸", "実効", "サーボ", "ワーク", "地点",
        ]
        parameter_en_terms = [
            "number", "setting", "limit", "position", "load", "production",
            "line", "servo", "spindle", "workpiece", "site",
        ]
        if any(pattern in compact_jp for pattern in parameter_terms):
            return True
        if any(pattern in compact_en for pattern in parameter_en_terms):
            return True
        return False

    bottom_navigation = image_height > 0 and region.y >= image_height * 0.80
    if bottom_navigation and not is_low_value_hmi_placeholder(region, english):
        return True

    return False


def filter_hmi_regions_for_review(
    regions: list[HmiTextRegion],
    translations: dict[str, str],
    image_height: int,
    review_detail: str,
) -> list[HmiTextRegion]:
    ordered = order_hmi_regions_row_major(regions)
    if review_detail == HMI_REVIEW_ALL:
        return ordered
    filtered = [
        region for region in ordered
        if is_important_hmi_region(region, translations.get(region.location, ""), image_height)
    ]
    unique: list[HmiTextRegion] = []
    seen_labels: set[str] = set()
    seen_semantic_labels: set[str] = set()
    for region in filtered:
        english = translations.get(region.location, "")
        key = hmi_review_label_key(region, english)
        semantic_key = hmi_review_semantic_key(region, english)
        if key in seen_labels:
            continue
        if semantic_key and semantic_key in seen_semantic_labels:
            continue
        seen_labels.add(key)
        if semantic_key:
            seen_semantic_labels.add(semantic_key)
        unique.append(region)
    return unique or filtered or ordered


def enrich_hmi_region_kinds(
    regions: list[HmiTextRegion],
    translations: dict[str, str],
    image_height: int,
    image_width: int,
) -> list[HmiTextRegion]:
    enriched: list[HmiTextRegion] = []
    for region in regions:
        kind = infer_hmi_region_kind(
            region,
            translations.get(region.location, ""),
            image_height,
            image_width,
        )
        if normalize_hmi_region_kind(region.kind) == kind:
            enriched.append(region)
            continue
        enriched.append(HmiTextRegion(
            location=region.location,
            jp=region.jp,
            x=region.x,
            y=region.y,
            width=region.width,
            height=region.height,
            confidence=region.confidence,
            note=region.note,
            kind=kind,
        ))
    return enriched


def hmi_region_display_no(region: HmiTextRegion, fallback_index: int) -> int:
    return fallback_index


def hmi_review_rows(regions: list[HmiTextRegion], translations: dict[str, str], glossary: pd.DataFrame) -> list[dict[str, str | int | float]]:
    rows = []
    for index, region in enumerate(regions, start=1):
        _glossary_text, hits = apply_glossary_to_source(region.jp, glossary, replace_source=False)
        rows.append({
            "No.": hmi_region_display_no(region, index),
            "Location": region.location,
            "Japanese": region.jp,
            "English": translations.get(region.location, ""),
            "Region Type": infer_hmi_region_kind(region, translations.get(region.location, ""), 0),
            "Engineering Context / Numbers / Units": region.note,
            "OCR Confidence": round(region.confidence, 3),
            "X": region.x,
            "Y": region.y,
            "Width": region.width,
            "Height": region.height,
            "Glossary Matches": "; ".join(f"{hit.jp}={hit.en}" for hit in hits),
            "Review Status": "Needs review" if region.confidence < 0.80 else "",
            "Note": region.note,
        })
    return rows


def hmi_glossary_traceability_rows(
    regions: list[HmiTextRegion],
    glossary: pd.DataFrame,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> list[dict[str, object]]:
    usage_counts: dict[tuple[str, str], int] = {}
    ordered_terms: list[tuple[str, str]] = []
    for region in regions:
        _controlled_source, hits = apply_glossary_to_source(
            region.jp,
            glossary,
            replace_source=False,
            translation_direction=translation_direction,
        )
        for hit in hits:
            key = (hit.jp, hit.en)
            if key not in usage_counts:
                ordered_terms.append(key)
                usage_counts[key] = 0
            usage_counts[key] += hit.count

    rows: list[dict[str, object]] = []
    for jp, en in ordered_terms:
        matched_records = glossary[
            (glossary["JP"].astype(str).map(clean_text) == jp)
            & (glossary["EN"].astype(str).map(clean_text) == en)
        ]
        glossary_record = matched_records.iloc[0].to_dict() if not matched_records.empty else {
            "JP": jp,
            "EN": en,
        }
        term_row: dict[str, object] = {}
        for column in glossary.columns:
            if column in glossary_record:
                term_row[str(column)] = glossary_record[column]
        term_row["Used Count"] = usage_counts[(jp, en)]
        rows.append(term_row)
    return rows


def hmi_result_quality_report(
    regions: list[HmiTextRegion],
    translations: dict[str, str],
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
    glossary: pd.DataFrame | None = None,
) -> tuple[str, list[str]]:
    issues: list[str] = []
    total = len(regions)
    empty_translation_count = sum(
        1 for region in regions
        if not clean_text(translations.get(region.location, ""))
    )
    low_confidence_count = sum(1 for region in regions if region.confidence < 0.80)
    controlled_term_violations: list[str] = []
    if glossary is not None and not glossary.empty:
        for region in regions:
            output = clean_text(translations.get(region.location, ""))
            _source, hits = apply_glossary_to_source(
                region.jp,
                glossary,
                replace_source=False,
                translation_direction=translation_direction,
            )
            for hit in hits:
                source_term = hit.en if translation_direction == TRANSLATION_DIRECTION_EN_JP else hit.jp
                target_term = hit.jp if translation_direction == TRANSLATION_DIRECTION_EN_JP else hit.en
                target_pattern = controlled_term_pattern(
                    target_term,
                    ignore_case=translation_direction == TRANSLATION_DIRECTION_JP_EN,
                )
                if len(target_pattern.findall(output)) < hit.count:
                    controlled_term_violations.append(f"{source_term} -> {target_term}")
    overlap_count = 0
    for index, region in enumerate(regions):
        for other in regions[index + 1:]:
            if hmi_regions_are_duplicate(region, other):
                overlap_count += 1
                break

    expected_numbers = list(range(1, total + 1))
    actual_numbers = [hmi_region_display_no(region, index) for index, region in enumerate(regions, start=1)]
    if actual_numbers != expected_numbers:
        issues.append("Numbering is not continuous.")
    if empty_translation_count:
        _source_language, target_language = direction_language_names(translation_direction)
        issues.append(f"{empty_translation_count} row(s) have empty {target_language} output.")
    if overlap_count:
        issues.append(f"{overlap_count} row(s) may overlap with another detected region.")
    if low_confidence_count:
        issues.append(f"{low_confidence_count} row(s) have low OCR confidence.")
    if controlled_term_violations:
        examples = "; ".join(dict.fromkeys(controlled_term_violations[:4]))
        issues.append(
            f"{len(controlled_term_violations)} controlled-term application(s) require review"
            + (f": {examples}." if examples else ".")
        )

    summary = (
        f"Quality check: {total:,} row(s), continuous numbering, "
        f"{max(total - empty_translation_count, 0):,}/{total:,} translated."
    )
    return summary, issues


def hmi_regions_from_review_df(df: pd.DataFrame, image_width: int, image_height: int) -> tuple[list[HmiTextRegion], dict[str, str]]:
    regions = []
    translations = {}
    for index, row in df.fillna("").iterrows():
        location = clean_text(row.get("Location", "")) or f"hmi:{index + 1}"
        jp = clean_text(row.get("Japanese", ""))
        english = clean_text(row.get("English", ""))
        if not jp and not english:
            continue
        x = clamp_region(row.get("X", 0), 0, image_width - 1)
        y = clamp_region(row.get("Y", 0), 0, image_height - 1)
        width = clamp_region(row.get("Width", 1), 1, image_width - x)
        height = clamp_region(row.get("Height", 1), 1, image_height - y)
        try:
            confidence = max(0.0, min(float(row.get("OCR Confidence", 0.0)), 1.0))
        except (TypeError, ValueError):
            confidence = 0.0
        regions.append(HmiTextRegion(
            location=location,
            jp=jp or english,
            x=x,
            y=y,
            width=width,
            height=height,
            confidence=confidence,
            note=clean_text(row.get("Note", "")),
            kind=clean_text(row.get("Region Type", "")),
        ))
        translations[location] = english
    return regions, translations


def missing_known_hmi_terms(raw: bytes, regions: list[HmiTextRegion]) -> list[str]:
    return []
    detected_text = "\n".join(region.jp for region in regions)
    known_terms = [
        "å¿œæ´å‘¼å‡º",
        "è·åˆ¶å‘¼å‡º",
        "è‹±æ–‡",
        "ãƒãƒ¼ãƒ‰ã‚³ãƒ”ãƒ¼é–‹å§‹",
        "é¸æŠžç”»é¢",
        "ç™»éŒ²",
        "ã‚­ãƒ£ãƒ³ã‚»ãƒ«",
    ]
    return [term for term in known_terms if term not in detected_text]


def dataframe_to_excel_bytes(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="HMI Translation Review")
    return output.getvalue()


def load_hmi_font(size: int):
    if ImageFont is None:
        return None
    font_candidates = [
        r"C:\Windows\Fonts\YuGothR.ttc",
        r"C:\Windows\Fonts\YuGothM.ttc",
        r"C:\Windows\Fonts\BIZ-UDGothicR.ttc",
        r"C:\Windows\Fonts\meiryo.ttc",
        r"C:\Windows\Fonts\msgothic.ttc",
        r"C:\Windows\Fonts\ARIALUNI.ttf",
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\segoeui.ttf",
        r"C:\Windows\Fonts\calibri.ttf",
        "arial.ttf",
        "segoeui.ttf",
        "calibri.ttf",
    ]
    for font_name in font_candidates:
        try:
            return ImageFont.truetype(font_name, size=size)
        except OSError:
            continue
    return ImageFont.load_default()


def text_bbox_size(draw, text: str, font) -> tuple[int, int]:
    bbox = draw.multiline_textbbox((0, 0), text, font=font, spacing=1)
    return bbox[2] - bbox[0], bbox[3] - bbox[1]


def wrap_hmi_text(draw, text: str, font, max_width: int) -> str:
    words = str(text).split()
    if not words:
        return str(text)
    lines = []
    current = words[0]
    for word in words[1:]:
        candidate = f"{current} {word}"
        if text_bbox_size(draw, candidate, font)[0] <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return "\n".join(lines)


def compact_repeated_hmi_words(text: str) -> str:
    words = str(text).split()
    if len(words) > 1 and len(set(words)) == 1:
        return words[0]
    return str(text)


def wrap_long_word(draw, word: str, font, max_width: int) -> list[str]:
    if text_bbox_size(draw, word, font)[0] <= max_width:
        return [word]
    parts = []
    current = ""
    for char in word:
        candidate = current + char
        if current and text_bbox_size(draw, candidate, font)[0] > max_width:
            parts.append(current)
            current = char
        else:
            current = candidate
    if current:
        parts.append(current)
    return parts


def wrap_hmi_text_strict(draw, text: str, font, max_width: int) -> str:
    wrapped_lines = []
    for line in wrap_hmi_text(draw, text, font, max_width).splitlines():
        if text_bbox_size(draw, line, font)[0] <= max_width:
            wrapped_lines.append(line)
            continue
        for word in line.split():
            wrapped_lines.extend(wrap_long_word(draw, word, font, max_width))
    return "\n".join(wrapped_lines)


def fit_hmi_text(draw, text: str, max_width: int, max_height: int, font_scale: float = 1.0):
    text = compact_repeated_hmi_words(text)
    start_size = max(6, min(18, int(max_height * 0.40 * font_scale)))
    for size in range(start_size, 4, -1):
        font = load_hmi_font(size)
        wrapped = wrap_hmi_text_strict(draw, text, font, max_width)
        width, height = text_bbox_size(draw, wrapped, font)
        if width <= max_width and height <= max_height:
            return wrapped, font
    font = load_hmi_font(5)
    return wrap_hmi_text_strict(draw, text, font, max_width), font


def readable_text_color(rgb: tuple[int, int, int]) -> tuple[int, int, int]:
    r, g, b = rgb
    luminance = (0.299 * r) + (0.587 * g) + (0.114 * b)
    return (0, 0, 0) if luminance > 150 else (255, 255, 255)


def dominant_region_color(image, box: tuple[int, int, int, int]) -> tuple[int, int, int]:
    crop = image.crop(box)
    counts: dict[tuple[int, int, int], int] = {}
    for r, g, b in crop.getdata():
        key = (round(r / 16) * 16, round(g / 16) * 16, round(b / 16) * 16)
        counts[key] = counts.get(key, 0) + 1
    if not counts:
        return (0, 0, 0)
    return max(counts.items(), key=lambda item: item[1])[0]


def wrap_review_table_text(draw, text: str, font, max_width: int) -> list[str]:
    words = str(text).split()
    if not words:
        return [""]
    lines = []
    current = words[0]
    for word in words[1:]:
        candidate = f"{current} {word}"
        if text_bbox_size(draw, candidate, font)[0] <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


def hmi_marker_rect(x: int, y: int, radius: int) -> tuple[int, int, int, int]:
    return (x, y, x + radius * 2, y + radius * 2)


def hmi_rects_overlap(a: tuple[int, int, int, int], b: tuple[int, int, int, int], padding: int = 2) -> bool:
    return not (
        a[2] + padding < b[0]
        or b[2] + padding < a[0]
        or a[3] + padding < b[1]
        or b[3] + padding < a[1]
    )


def hmi_marker_positions(
    regions: list[HmiTextRegion],
    image_width: int,
    image_height: int,
    marker_radii: list[int] | None = None,
) -> list[tuple[int, int]]:
    positions: list[tuple[int, int]] = []
    occupied: list[tuple[int, int, int, int]] = []
    for index, region in enumerate(regions):
        radius = marker_radii[index] if marker_radii and index < len(marker_radii) else 10
        marker_size = radius * 2
        grid = 4
        base_x = int(round((region.x + 2) / grid) * grid)
        base_y = int(round((region.y + 2) / grid) * grid)
        step = max(marker_size + 3, 22)
        candidate_points = []
        for row_offset in (0, step, step * 2):
            for col_offset in (0, step, -step, step * 2, -step * 2):
                candidate_points.append((base_x + col_offset, base_y + row_offset))
        candidate_points.extend([
            (base_x, base_y - step),
            (base_x + step, base_y - step),
            (base_x - step, base_y - step),
        ])
        chosen = None
        for candidate_x, candidate_y in candidate_points:
            marker_x = max(2, min(int(candidate_x), image_width - marker_size - 2))
            marker_y = max(2, min(int(candidate_y), image_height - marker_size - 2))
            marker_x = int(round(marker_x / grid) * grid)
            marker_y = int(round(marker_y / grid) * grid)
            rect = hmi_marker_rect(marker_x, marker_y, radius)
            if not any(hmi_rects_overlap(rect, existing) for existing in occupied):
                chosen = (marker_x, marker_y, rect)
                break
        if chosen is None:
            marker_x = max(2, min(region.x + 2, image_width - marker_size - 2))
            marker_y = max(2, min(region.y + 2, image_height - marker_size - 2))
            chosen = (marker_x, marker_y, hmi_marker_rect(marker_x, marker_y, radius))
        positions.append((chosen[0], chosen[1]))
        occupied.append(chosen[2])
    return positions


def render_hmi_review_map_image(
    raw: bytes,
    regions: list[HmiTextRegion],
    translations: dict[str, str],
    numbered_raw: bytes | None = None,
    image_mode: str = IMAGE_MODE_HMI,
    translation_direction: str = TRANSLATION_DIRECTION_JP_EN,
) -> bytes:
    if Image is None or ImageDraw is None:
        raise RuntimeError("Pillow is not installed. Install pillow to create HMI review maps.")

    image_bytes = numbered_raw or raw
    original = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    source_language, target_language = direction_language_names(translation_direction)
    table_columns = 2 if len(regions) > 24 else 1
    panel_width = 860 if table_columns == 2 else 520
    row_height = 48 if table_columns == 2 else 54
    top_margin = 42
    rows_per_column = max((len(regions) + table_columns - 1) // table_columns, 1)
    table_height = top_margin + rows_per_column * row_height + 16
    canvas_width = original.width + panel_width
    canvas_height = max(original.height, table_height)
    canvas = Image.new("RGB", (canvas_width, canvas_height), (242, 244, 247))
    canvas.paste(original, (0, 0))

    draw = ImageDraw.Draw(canvas)
    marker_font = load_hmi_font(12)
    header_font = load_hmi_font(16)
    row_font = load_hmi_font(11)
    small_font = load_hmi_font(9)

    marker_radii = []
    for index, region in enumerate(regions, start=1):
        label = str(hmi_region_display_no(region, index))
        bbox = draw.textbbox((0, 0), label, font=marker_font)
        label_w = bbox[2] - bbox[0]
        marker_radii.append(max(10, label_w // 2 + 6))
    marker_positions = hmi_marker_positions(regions, original.width, original.height, marker_radii)
    for index, (region, (marker_x, marker_y)) in enumerate(zip(regions, marker_positions), start=1):
        if numbered_raw and region.location.startswith("hmi_box:"):
            continue
        label = str(hmi_region_display_no(region, index))
        bbox = draw.textbbox((0, 0), label, font=marker_font)
        label_w = bbox[2] - bbox[0]
        label_h = bbox[3] - bbox[1]
        radius = max(10, label_w // 2 + 6)
        draw.ellipse(
            (marker_x, marker_y, marker_x + radius * 2, marker_y + radius * 2),
            fill=(255, 230, 0),
            outline=(175, 40, 35),
            width=3,
        )
        draw.text(
            (marker_x + radius - label_w / 2, marker_y + radius - label_h / 2 - 1),
            label,
            fill=(0, 0, 0),
            font=marker_font,
        )

    panel_x = original.width
    draw.rectangle((panel_x, 0, canvas_width, canvas_height), fill=(248, 249, 251))
    if image_mode == IMAGE_MODE_HMI:
        map_type = "HMI"
    elif image_mode == IMAGE_MODE_ENGINEERING:
        map_type = "Engineering Image"
    else:
        map_type = "General Image"
    draw.text(
        (panel_x + 18, 12),
        f"{map_type} Translation Review Map ({translation_direction})",
        fill=(20, 27, 39),
        font=header_font,
    )
    draw.line((panel_x, 38, canvas_width, 38), fill=(200, 205, 215), width=1)

    table_left = panel_x + 10
    column_width = (panel_width - 20) // table_columns
    for column_index in range(table_columns):
        column_x = table_left + column_index * column_width
        col_no = column_x + 8
        col_jp = column_x + 44
        col_en = column_x + (176 if table_columns == 2 else 228)
        draw.text((col_no, top_margin - 18), "No.", fill=(55, 65, 81), font=small_font)
        draw.text((col_jp, top_margin - 18), source_language, fill=(55, 65, 81), font=small_font)
        draw.text((col_en, top_margin - 18), target_language, fill=(55, 65, 81), font=small_font)

    for index, region in enumerate(regions, start=1):
        column_index = (index - 1) // rows_per_column
        row_index = (index - 1) % rows_per_column
        column_x = table_left + column_index * column_width
        y = top_margin + row_index * row_height
        col_no = column_x + 8
        col_jp = column_x + 44
        col_en = column_x + (176 if table_columns == 2 else 228)
        row_right = min(column_x + column_width - 8, canvas_width - 8)
        jp_width = 122 if table_columns == 2 else 160
        en_width = max(145, row_right - col_en - 8)
        row_fill = (255, 255, 255) if index % 2 else (239, 242, 247)
        draw.rectangle((column_x, y - 4, row_right, y + row_height - 8), fill=row_fill, outline=(218, 223, 231))
        draw.text((col_no, y + 10), str(hmi_region_display_no(region, index)), fill=(17, 24, 39), font=row_font)
        jp_lines = wrap_review_table_text(draw, region.jp, row_font, jp_width)[:3]
        en_lines = wrap_review_table_text(draw, translations.get(region.location, ""), row_font, en_width)[:3]
        draw.multiline_text((col_jp, y + 4), "\n".join(jp_lines), fill=(17, 24, 39), font=row_font, spacing=1)
        draw.multiline_text((col_en, y + 4), "\n".join(en_lines), fill=(17, 24, 39), font=row_font, spacing=1)

    output = io.BytesIO()
    canvas.save(output, format="PNG")
    return output.getvalue()


def render_hmi_source_marker_map_image(raw: bytes, regions: list[HmiTextRegion]) -> bytes:
    if Image is None or ImageDraw is None:
        raise RuntimeError("Pillow is not installed. Install pillow to create HMI marker maps.")

    image = Image.open(io.BytesIO(raw)).convert("RGB")
    draw = ImageDraw.Draw(image)
    marker_font = load_hmi_font(12)

    marker_positions = hmi_marker_positions(regions, image.width, image.height)
    for index, (region, (marker_x, marker_y)) in enumerate(zip(regions, marker_positions), start=1):
        label = str(hmi_region_display_no(region, index))
        bbox = draw.textbbox((0, 0), label, font=marker_font)
        label_w = bbox[2] - bbox[0]
        label_h = bbox[3] - bbox[1]
        radius = max(9, label_w // 2 + 5)
        draw.ellipse(
            (marker_x, marker_y, marker_x + radius * 2, marker_y + radius * 2),
            fill=(255, 230, 0),
            outline=(175, 40, 35),
            width=2,
        )
        draw.text(
            (marker_x + radius - label_w / 2, marker_y + radius - label_h / 2 - 1),
            label,
            fill=(0, 0, 0),
            font=marker_font,
        )

    output = io.BytesIO()
    image.save(output, format="PNG")
    return output.getvalue()


def render_hmi_numbered_english_screen(raw: bytes, regions: list[HmiTextRegion], translations: dict[str, str]) -> bytes:
    if Image is None or ImageDraw is None:
        raise RuntimeError("Pillow is not installed. Install pillow to create HMI numbered English screens.")

    image = Image.open(io.BytesIO(raw)).convert("RGB")
    draw = ImageDraw.Draw(image, "RGBA")
    marker_font = load_hmi_font(10)
    text_font = load_hmi_font(10)

    for index, region in enumerate(regions, start=1):
        english = clean_text(translations.get(region.location, ""))
        if not english:
            continue

        display_no = hmi_region_display_no(region, index)
        marker_text = str(display_no)
        label_text = f"{display_no} {english}"
        max_label_width = 190
        label_lines = wrap_review_table_text(draw, label_text, text_font, max_label_width)
        label_text_wrapped = "\n".join(label_lines[:3])
        text_w, text_h = text_bbox_size(draw, label_text_wrapped, text_font)

        label_w = min(max(text_w + 12, 42), max_label_width + 12)
        label_h = text_h + 8
        x = max(0, min(region.x, image.width - label_w - 1))
        y = max(0, min(region.y, image.height - label_h - 1))

        # Use a compact callout anchored at the original Japanese text position.
        draw.rectangle(
            (x, y, x + label_w, y + label_h),
            fill=(20, 24, 33, 215),
            outline=(255, 230, 0, 230),
            width=1,
        )
        badge_size = 18
        draw.ellipse(
            (x + 2, y + 2, x + 2 + badge_size, y + 2 + badge_size),
            fill=(255, 230, 0, 245),
            outline=(160, 35, 35, 245),
            width=1,
        )
        badge_bbox = draw.textbbox((0, 0), marker_text, font=marker_font)
        badge_w = badge_bbox[2] - badge_bbox[0]
        badge_h = badge_bbox[3] - badge_bbox[1]
        draw.text(
            (x + 2 + badge_size / 2 - badge_w / 2, y + 2 + badge_size / 2 - badge_h / 2 - 1),
            marker_text,
            fill=(0, 0, 0, 255),
            font=marker_font,
        )
        english_only = "\n".join(wrap_review_table_text(draw, english, text_font, max(label_w - badge_size - 12, 20))[:3])
        draw.multiline_text(
            (x + badge_size + 8, y + 4),
            english_only,
            fill=(255, 255, 255, 255),
            font=text_font,
            spacing=1,
        )

    output = io.BytesIO()
    image.save(output, format="PNG")
    return output.getvalue()




def render_hmi_editable_position_excel_bytes(
    raw: bytes,
    regions: list[HmiTextRegion],
    translations: dict[str, str],
    review_rows: list[dict[str, object]] | None = None,
    detected_box_map: bytes | None = None,
    review_map_image: bytes | None = None,
    source_marker_map: bytes | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Review Table"
    sheet.sheet_view.showGridLines = False

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    thin_gray = Side(style="thin", color="A6A6A6")
    cell_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    rows = review_rows or hmi_review_rows(regions, translations, empty_terms_dataframe())
    headers = [
        "No.",
        "Japanese",
        "English",
        "Engineering Context / Numbers / Units",
        "OCR Confidence",
        "Review Status",
        "Glossary Matches",
        "X",
        "Y",
        "Width",
        "Height",
        "Location",
        "Note",
    ]
    widths = [8, 30, 34, 34, 14, 18, 32, 9, 9, 9, 9, 18, 34]

    sheet["A1"] = "Image Translation Review Table"
    sheet["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    sheet["A1"].fill = title_fill
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    sheet["A2"] = "Each row matches the numbered marker on the output image. Edit English and Review Status during engineering review."
    sheet["A2"].font = Font(size=10, color="667085")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    for col_index, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = sheet.cell(row=4, column=col_index, value=header)
        cell.font = Font(bold=True, color="111827")
        cell.fill = header_fill
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(col_index)].width = width

    for row_index, row in enumerate(rows, start=5):
        for col_index, header in enumerate(headers, start=1):
            value = row.get(header, "")
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if header in {"English", "Review Status", "Note"}:
                cell.fill = editable_fill
            if header == "No.":
                cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[row_index].height = 42

    sheet.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(len(rows) + 4, 4)}"
    sheet.freeze_panes = "A5"

    def add_image_sheet(title: str, image_bytes: bytes | None, caption: str) -> None:
        if not image_bytes:
            return
        image_sheet = workbook.create_sheet(title)
        image_sheet.sheet_view.showGridLines = False
        image_sheet["A1"] = caption
        image_sheet["A1"].font = Font(bold=True, size=14, color="1F2937")
        image_sheet["A2"] = "Use this image together with the Review Table No. column."
        image_sheet["A2"].font = Font(size=10, color="667085")
        image_file = io.BytesIO(image_bytes)
        excel_image = XLImage(image_file)
        max_width = 1100
        if excel_image.width > max_width:
            scale = max_width / excel_image.width
            excel_image.width = int(excel_image.width * scale)
            excel_image.height = int(excel_image.height * scale)
        image_sheet.add_image(excel_image, "A4")
        image_sheet.column_dimensions["A"].width = 140

    add_image_sheet("Review Map", review_map_image, "Image Review Map")
    add_image_sheet("Original Number Map", source_marker_map, "Original Image Number Map")
    add_image_sheet("All Detected Boxes", detected_box_map, "All Detected HMI Boxes")

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def render_hmi_translated_image(
    raw: bytes,
    regions: list[HmiTextRegion],
    translations: dict[str, str],
    font_scale: float = 1.0,
    padding_pixels: int = 1,
) -> bytes:
    if Image is None or ImageDraw is None:
        raise RuntimeError("Pillow is not installed. Install pillow to create translated HMI images.")
    image = Image.open(io.BytesIO(raw)).convert("RGB")
    draw = ImageDraw.Draw(image)
    for region in regions:
        english = clean_text(translations.get(region.location, ""))
        if not english:
            continue
        pad_x = max(0, int(padding_pixels))
        pad_y = max(0, int(padding_pixels))
        x1 = max(region.x - pad_x, 0)
        y1 = max(region.y - pad_y, 0)
        x2 = min(region.x + region.width + pad_x, image.width)
        y2 = min(region.y + region.height + pad_y, image.height)
        fill = dominant_region_color(image, (x1, y1, x2, y2))
        text_color = readable_text_color(fill)
        draw.rectangle((x1, y1, x2, y2), fill=fill)
        wrapped, font = fit_hmi_text(
            draw,
            english,
            max(x2 - x1 - 4, 1),
            max(y2 - y1 - 4, 1),
            font_scale=font_scale,
        )
        text_width, text_height = text_bbox_size(draw, wrapped, font)
        text_x = x1 + max((x2 - x1 - text_width) // 2, 2)
        text_y = y1 + max((y2 - y1 - text_height) // 2, 2)
        draw.multiline_text((text_x, text_y), wrapped, fill=text_color, font=font, align="center", spacing=1)
    output = io.BytesIO()
    image.save(output, format="PNG")
    return output.getvalue()


HMI_OUTPUT_STATE_KEYS = (
    "hmi_detected_box_map",
    "hmi_source_marker_map",
    "hmi_review_map_image",
    "hmi_numbered_english_screen",
    "hmi_translated_image",
    "hmi_position_excel",
    "hmi_last_raw_image",
    "hmi_last_image_size",
    "hmi_result_image_key",
    "hmi_glossary_traceability",
)
HMI_INPUT_CHECK_STATE_KEYS = (
    "hmi_input_check_key",
    "hmi_input_check_ok",
    "hmi_input_check_reason",
    "hmi_input_check_usage",
)


def clear_hmi_output_state() -> None:
    for key in HMI_OUTPUT_STATE_KEYS:
        st.session_state.pop(key, None)


def clear_hmi_input_check_state() -> None:
    for key in HMI_INPUT_CHECK_STATE_KEYS:
        st.session_state.pop(key, None)


def render_hmi_translation(glossary: pd.DataFrame, plc_rules: pd.DataFrame) -> None:
    st.caption("Step 1 - Select translation direction (required)")
    translation_direction = st.radio(
        "Image translation direction",
        TRANSLATION_DIRECTIONS,
        index=None,
        horizontal=True,
        key="image_translation_direction",
        help="Choose a direction before selecting an image type or uploading an image.",
    )
    if translation_direction is None:
        st.info("Select JP → EN or EN → JP to continue.")
        return

    source_language, target_language = direction_language_names(translation_direction)
    st.caption("Step 2 - Select image type")
    image_mode = st.radio(
        "Image type",
        IMAGE_TRANSLATION_MODES,
        horizontal=True,
        help=(
            "Choose HMI Screen for machine interface screenshots. Choose Other Image / CAD / Drawing "
            "for engineering drawings, standard sheets, CAD screenshots, tables, and process diagrams. "
            "Choose General Image for photos, conceptual diagrams, training material, or ordinary screenshots "
            "that contain Japanese or English text."
        ),
        key="image_translation_mode",
    )
    st.caption(image_mode_guidance(image_mode, translation_direction))
    review_detail = st.radio(
        "Review Detail",
        HMI_REVIEW_DETAIL_MODES,
        horizontal=True,
        help=(
            "Key review labels keeps the review map focused on buttons, navigation, screen titles, "
            "alarms, and meaningful parameters. All detected labels includes low-value placeholders such as Data 6 or Unit 18."
        ),
        key="hmi_review_detail",
    )
    uploaded_image = st.file_uploader(
        "Upload image",
        type=["png", "jpg", "jpeg"],
        label_visibility="collapsed",
        key="hmi_image_upload",
    )
    if uploaded_image is None:
        clear_hmi_output_state()
        clear_hmi_input_check_state()
        st.session_state.pop("hmi_image_key", None)
        return
    if Image is None:
        st.error("Pillow is required for HMI image translation. Please install pillow in the app environment.")
        return

    raw_image = uploaded_image.getvalue()
    image = Image.open(io.BytesIO(raw_image)).convert("RGB")
    image_key = document_fingerprint(
        f"{translation_direction}:{image_mode}:{review_detail}:{uploaded_image.name}",
        raw_image,
    )
    if st.session_state.get("hmi_image_key") != image_key:
        st.session_state["hmi_image_key"] = image_key
        clear_hmi_output_state()
        clear_hmi_input_check_state()
    st.image(raw_image, caption=f"Original image - {image_mode}", use_container_width=True)
    st.info(f"{uploaded_image.name} is ready. Image type: {image_mode}. Image size: {image.width} x {image.height} px.")
    if st.session_state.get("hmi_input_check_key") != image_key:
        with st.spinner("Checking image type..."):
            try:
                is_suitable, suitability_reason, suitability_usage = validate_image_translation_input(
                    raw_image,
                    uploaded_image.name,
                    image.width,
                    image.height,
                    image_mode,
                )
                st.session_state["hmi_input_check_key"] = image_key
                st.session_state["hmi_input_check_ok"] = is_suitable
                st.session_state["hmi_input_check_reason"] = suitability_reason
                st.session_state["hmi_input_check_usage"] = suitability_usage
            except Exception as exc:
                st.session_state["hmi_input_check_key"] = image_key
                st.session_state["hmi_input_check_ok"] = False
                st.session_state["hmi_input_check_reason"] = format_translation_error(exc)
                st.session_state["hmi_input_check_usage"] = TokenUsage()

    input_check_ok = bool(st.session_state.get("hmi_input_check_ok"))
    input_check_reason = clean_text(st.session_state.get("hmi_input_check_reason") or "")
    input_check_usage = st.session_state.get("hmi_input_check_usage") or TokenUsage()
    if input_check_ok:
        st.success("The image is ready for translation.")
    else:
        st.warning(
            "The preliminary check could not confirm readable Japanese or English text. "
            "You can still try translation; the detailed text scan will stop if no source text is found."
        )
        if input_check_reason:
            st.caption(input_check_reason)
    font_scale = 1.15
    overlay_padding = 0

    if st.button("Translate Image", type="primary", key="start_hmi_translation"):
        increment_usage_action("translate_image")
        increment_translation_usage_count(IMAGE_TRANSLATION_USAGE_KEY)
        clear_hmi_output_state()
        status = st.empty()
        progress = st.progress(0)
        try:
            active_glossary = hmi_glossary_for_mode(glossary, plc_rules)
            if image_mode == IMAGE_MODE_HMI:
                status.write(f"Finding HMI boxes/cells and reading {source_language} text with OpenAI Vision...")
            else:
                status.write(f"Reading {source_language} text and positions from {image_mode.lower()} with OpenAI Vision...")
            progress.progress(0.10)

            def update_image_detection_progress(stage: str, current: int, total: int) -> None:
                ratio = min(max(current / max(total, 1), 0.0), 1.0)
                if stage == "boxes":
                    status.write(
                        f"Reading HMI boxes with OpenAI Vision: batch {current}/{total}..."
                    )
                    progress.progress(0.10 + ratio * 0.28)
                else:
                    status.write(
                        f"Checking remaining HMI areas: region {current}/{total}..."
                    )
                    progress.progress(0.38 + ratio * 0.17)

            regions, vision_usage = extract_hmi_text_regions_with_vision(
                raw_image,
                uploaded_image.name,
                image.width,
                image.height,
                image_mode,
                translation_direction,
                update_image_detection_progress,
            )
            if not regions:
                st.warning(f"No {source_language} text was detected. Try a sharper image or crop.")
                return
            missing_terms = missing_known_hmi_terms(raw_image, regions) if translation_direction == TRANSLATION_DIRECTION_JP_EN else []
            if missing_terms:
                st.warning(
                    "Some common HMI labels were not detected by Vision and should be manually reviewed: "
                    + ", ".join(missing_terms)
                )

            status.write(f"Translating detected image text to {target_language} with controlled terminology...")
            progress.progress(0.60)
            translations, hits, translation_usage = translate_hmi_regions(
                regions,
                active_glossary,
                image_mode,
                translation_direction,
            )
            regions = enrich_hmi_region_kinds(regions, translations, image.height, image.width)
            clean_regions = deduplicate_hmi_regions_for_output(regions, translations)
            review_regions = filter_hmi_regions_for_review(
                clean_regions,
                translations,
                image.height,
                review_detail,
            )

            status.write("Rendering one clean image translation review result...")
            progress.progress(0.85)
            detected_box_map = b""
            if image_mode == IMAGE_MODE_HMI:
                detected_boxes = detect_hmi_candidate_boxes(raw_image)
                if detected_boxes:
                    detected_box_map = render_hmi_box_number_overlay(raw_image, detected_boxes)
            review_map_image = render_hmi_review_map_image(
                raw_image,
                review_regions,
                translations,
                numbered_raw=None,
                image_mode=image_mode,
                translation_direction=translation_direction,
            )
            source_marker_map = b""
            progress.progress(1.0)
            status.success("Image translation complete.")
            st.session_state["hmi_detected_box_map"] = detected_box_map
            st.session_state["hmi_review_map_image"] = review_map_image
            st.session_state["hmi_source_marker_map"] = source_marker_map
            st.session_state["hmi_position_excel"] = b""
            st.session_state["hmi_last_raw_image"] = raw_image
            st.session_state["hmi_last_image_size"] = (image.width, image.height)
            st.session_state["hmi_result_image_key"] = image_key
            st.session_state["hmi_glossary_traceability"] = pd.DataFrame(
                hmi_glossary_traceability_rows(
                    review_regions,
                    active_glossary,
                    translation_direction,
                )
            )
            total_usage = TokenUsage()
            total_usage.add(input_check_usage)
            total_usage.add(vision_usage)
            total_usage.add(translation_usage)
            st.caption(f"Detected text regions: {len(regions):,} | API tokens: {total_usage.display()}")
            st.caption(f"Review detail: {review_detail} | Review rows: {len(review_regions):,}/{len(clean_regions):,}")
            quality_summary, quality_issues = hmi_result_quality_report(
                review_regions,
                translations,
                translation_direction,
                active_glossary,
            )
            if quality_issues:
                st.warning(quality_summary + " " + " ".join(quality_issues))
            else:
                st.success(quality_summary)
            if hits:
                st.caption(f"Glossary matches: {len(hits):,}")
        except Exception as exc:
            clear_hmi_output_state()
            status.error("Image translation failed.")
            st.error(f"Image translation failed: {format_translation_error(exc)}")

    has_current_hmi_result = (
        st.session_state.get("hmi_result_image_key") == image_key
        and st.session_state.get("hmi_review_map_image")
    )
    if has_current_hmi_result:
        st.subheader("Image Review Map")
        st.image(st.session_state["hmi_review_map_image"], use_container_width=True)
        stem = Path(uploaded_image.name).stem
        download_key_suffix = image_key[:12]
        st.download_button(
            f"Download {translation_direction} review result",
            data=st.session_state["hmi_review_map_image"],
            file_name=f"Translation-Review-{translation_direction_key(translation_direction)}-{stem}.png",
            mime="image/png",
            key=f"download_hmi_review_result_{download_key_suffix}",
        )
        st.subheader("Controlled Glossary Terms Used")
        st.caption(
            "Complete glossary information for the controlled terms used in this image translation."
        )
        glossary_traceability = st.session_state.get("hmi_glossary_traceability")
        if isinstance(glossary_traceability, pd.DataFrame) and not glossary_traceability.empty:
            total_applications = int(glossary_traceability["Used Count"].sum())
            st.caption(
                f"Controlled terms used: {len(glossary_traceability):,} | "
                f"Total applications: {total_applications:,}"
            )
            st.dataframe(glossary_traceability, use_container_width=True, hide_index=True)
        else:
            st.info("No controlled glossary terms were matched in this image.")


def parse_batch_log_status() -> dict[str, str | int | float] | None:
    log_path = BASE_DIR / "batch_outputs" / "COMMENT_batch_translate.log"
    if not log_path.exists():
        return None
    lines = [line.strip() for line in log_path.read_text(encoding="utf-8", errors="replace").splitlines() if line.strip()]
    if not lines:
        return None
    latest = lines[-1]
    updated_text = latest.split("|", 1)[0].strip()
    updated_at = parse_timestamp(updated_text)
    status = {
        "updated_text": updated_text,
        "updated_age": elapsed_since_timestamp(updated_text),
        "latest": latest,
    }
    if updated_at is not None:
        status["updated_epoch"] = updated_at.timestamp()
    batch_match = re.search(r"batch=([\d,]+)/([\d,]+)", latest)
    saved_match = re.search(r"saved=([\d,]+)/([\d,]+)", latest)
    if batch_match:
        status["completed_batches"] = int(batch_match.group(1).replace(",", ""))
        status["total_batches"] = int(batch_match.group(2).replace(",", ""))
    if saved_match:
        status["saved_blocks"] = int(saved_match.group(1).replace(",", ""))
        status["total_blocks"] = int(saved_match.group(2).replace(",", ""))
    recent_points = []
    for line in lines[-60:]:
        line_time = parse_timestamp(line.split("|", 1)[0].strip())
        line_saved_match = re.search(r"saved=([\d,]+)/([\d,]+)", line)
        if line_time is not None and line_saved_match:
            recent_points.append((line_time, int(line_saved_match.group(1).replace(",", ""))))
    if len(recent_points) >= 2 and "saved_blocks" in status and "total_blocks" in status:
        first_time, first_saved = recent_points[0]
        last_time, last_saved = recent_points[-1]
        seconds = max((last_time - first_time).total_seconds(), 0)
        saved_delta = max(last_saved - first_saved, 0)
        if seconds > 0 and saved_delta > 0:
            remaining_blocks = max(int(status["total_blocks"]) - int(status["saved_blocks"]), 0)
            status["eta_seconds"] = remaining_blocks / (saved_delta / seconds)
    status["is_recent"] = updated_at is not None and (datetime.now() - updated_at).total_seconds() < 600
    return status


def render_batch_log_status(owner_session_id: str) -> bool:
    # The legacy COMMENT.csv batch log and output files are global to the server.
    # Never render them in the multi-user UI because they cannot be attributed
    # safely to one browser session.
    return False

    if not latest_running_translation_job_id(owner_session_id):
        return False
    status = parse_batch_log_status()
    if status is None:
        return False
    saved_blocks = int(status.get("saved_blocks", 0))
    total_blocks = int(status.get("total_blocks", 0))
    completed_batches = int(status.get("completed_batches", 0))
    total_batches = int(status.get("total_batches", 0))
    if total_blocks <= 0 and total_batches <= 0:
        return False

    ratio = min(saved_blocks / total_blocks, 1.0) if total_blocks else 0.0
    progress_label = "Complete" if ratio >= 1.0 else "Translating"
    updated_age = float(status.get("updated_age") or 0)
    partial_output_path = BASE_DIR / "batch_outputs" / "Translated-COMMENT.partial.csv"
    final_output_path = BASE_DIR / "batch_outputs" / "Translated-COMMENT.csv"
    latest_log_epoch = float(status.get("updated_epoch") or 0)

    st.subheader("Current translation")
    st.write("COMMENT.csv")
    if status.get("is_recent"):
        st.success("Translation is running. You can refresh this page or leave it open.")
    else:
        st.warning("Translation has not updated recently. The saved progress is still available.")
    st.progress(max(ratio, 0.02))
    st.write(f"{progress_label} | {ratio * 100:.2f}%")

    eta_seconds = status.get("eta_seconds")
    status_cols = st.columns(4)
    status_cols[0].metric("Live progress", f"{saved_blocks:,}/{total_blocks:,}")
    status_cols[1].metric("Last update", f"{format_duration(updated_age)} ago")
    status_cols[2].metric("Estimated finish", format_duration(float(eta_seconds)) if eta_seconds else "Calculating")

    partial_mtime = partial_output_path.stat().st_mtime if partial_output_path.exists() else 0
    partial_label = datetime.fromtimestamp(partial_mtime).strftime("%Y-%m-%d %H:%M:%S") if partial_mtime else "Not generated yet"
    status_cols[3].metric("Download file", partial_label)

    if latest_log_epoch and partial_mtime and partial_mtime + 300 < latest_log_epoch:
        st.info("The downloadable file is behind the live progress. Refresh it before downloading.")

    download_cols = st.columns([1, 1])
    if download_cols[0].button("Refresh Download File", key="update_comment_partial_output"):
        try:
            source_path = max(JOB_UPLOAD_DIR.glob("*COMMENT.csv"), key=lambda path: path.stat().st_mtime)
            raw = source_path.read_bytes()
            checkpoint_path = checkpoint_path_for("COMMENT.csv", raw, PLC_TRANSLATION_MODE)
            translations = load_checkpoint(checkpoint_path)
            blocks = extract_text_blocks(raw, "COMMENT.csv")
            partial_output_path.parent.mkdir(exist_ok=True)
            partial_output_path.write_bytes(build_translated_document(raw, "COMMENT.csv", translations, blocks))
            st.success("Download file refreshed from saved progress.")
        except Exception as exc:
            st.error(f"Could not refresh download file: {exc}")
    if partial_output_path.exists():
        download_cols[1].download_button(
            "Download Current File",
            data=partial_output_path.read_bytes(),
            file_name="Translated-COMMENT.partial.csv",
            mime="text/csv",
            key="download_batch_partial_output",
        )
    if ratio >= 1.0 and final_output_path.exists():
        final_updated = datetime.fromtimestamp(final_output_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        st.caption(f"Final output file updated: {final_updated}")
        st.download_button(
            "Download Final Output",
            data=final_output_path.read_bytes(),
            file_name="Translated-COMMENT.csv",
            mime="text/csv",
            key="download_batch_final_output",
        )
    with st.expander("Advanced details"):
        st.write(f"Completed batches: {completed_batches:,}/{total_batches:,}")
        st.write(f"Translated blocks: {saved_blocks:,}/{total_blocks:,}")
        st.write(f"Latest log: {status.get('latest', '')}")
    return True


def render_document_translation(
    glossary: pd.DataFrame,
    plc_rules: pd.DataFrame,
    owner_session_id: str,
) -> None:
    translation_direction = st.radio(
        "Translation direction",
        TRANSLATION_DIRECTIONS,
        index=None,
        horizontal=True,
        key="document_translation_direction",
    )
    if translation_direction is None:
        st.info("Select JP → EN or EN → JP to continue.")
        return

    source_language, _target_language = direction_language_names(translation_direction)
    st.markdown(
        'Supported files: <span style="color:#c62828; font-weight:700;">PDF, DOCX, PPTX, XLSX, XLSM, CSV, TXT</span> | Max. file size: 100 MB',
        unsafe_allow_html=True,
    )
    st.caption("Engineering PDFs: use Catalog / Specs and verify the output before use.")
    translation_mode = st.radio(
        "Content type",
        DOCUMENT_TRANSLATION_MODES,
        format_func=lambda mode: DOCUMENT_TRANSLATION_MODE_LABELS.get(mode, mode),
        horizontal=False,
        key="document_translation_mode",
    )
    keep_source_with_translation = False

    active_jobs = active_translation_job_count(owner_session_id)
    if active_jobs:
        control_cols = st.columns([1, 1, 3])
        control_cols[0].metric("Active jobs", f"{active_jobs:,}")
        if control_cols[1].button("Stop All Active Jobs", key="stop_all_active_document_jobs"):
            increment_usage_action("stop_all_active_document_jobs")
            stopped_count = stop_all_active_translation_jobs(owner_session_id)
            st.session_state.pop("active_document_job_id", None)
            st.success(f"Stopped {stopped_count:,} active job(s). Saved progress is preserved.")
            rerun_app()

    has_current_job = render_current_document_job(
        glossary,
        plc_rules,
        translation_mode,
        translation_direction,
        owner_session_id,
    )
    if has_current_job:
        return

    uploaded_document = st.file_uploader(
        "Upload file",
        type=["csv", "txt", "docx", "pptx", "xlsx", "xlsm", "pdf"],
        help="For mechanical or electrical drawing PDFs, select Catalog / Specs before starting translation.",
    )

    if uploaded_document is None:
        return

    raw_document = uploaded_document.getvalue()
    if len(raw_document) > MAX_UPLOAD_BYTES:
        st.warning("This document is larger than the 100 MB safety limit. Please split the file or test with a smaller copy.")
        return
    if uploaded_document.name.lower().endswith((".as", ".ad")) and translation_mode != ROBOT_PROGRAM_TRANSLATION_MODE:
        st.info("For AS/AD robot files, select Kawasaki Robot .as file. The app will translate readable Japanese robot comments/labels and write English back into those fields.")
    encoding_warning = robot_encoding_warning(raw_document, uploaded_document.name)
    if encoding_warning:
        warning_html = html.escape(encoding_warning).replace("\n", "<br>")
        st.markdown(
            f"""
            <div style="border: 3px solid #b91c1c; background: #fee2e2; color: #7f1d1d; padding: 18px 20px; margin: 14px 0; border-radius: 6px;">
              <div style="font-size: 2.1rem; font-weight: 800; line-height: 1.2; margin-bottom: 8px;">AS FILE WARNING</div>
              <div style="font-size: 1.45rem; font-weight: 600; line-height: 1.35;">{warning_html}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    document_key = (
        f"{document_fingerprint(uploaded_document.name, raw_document)}::"
        f"{translation_direction}::{translation_mode}::keep_source={keep_source_with_translation}"
    )
    progress_path = checkpoint_path_for(
        uploaded_document.name,
        raw_document,
        translation_mode,
        translation_direction,
        owner_session_id,
    )
    if st.session_state.get("translated_document_key") != document_key:
        st.session_state.pop("translated_document_bytes", None)
        st.session_state.pop("translated_document_name", None)
        st.session_state.pop("translated_document_mime", None)
        st.session_state.pop("translated_document_preview", None)
        st.session_state.pop("translated_document_terms", None)
        st.session_state["translated_document_key"] = document_key

    notify_email = ""
    if latest_running_translation_job_id(owner_session_id):
        st.warning("You already have another translation job active. Starting a second job may slow both jobs down.")

    st.caption("Ready")
    st.info(f"{uploaded_document.name} is ready to queue. Size: {len(raw_document) / (1024 * 1024):.1f} MB.")
    translate_clicked = st.button("Start Translation", type="primary")

    if translate_clicked:
        increment_usage_action("start_document_translation")
        increment_translation_usage_count(DOCUMENT_TRANSLATION_USAGE_KEY)
        active_glossary = glossary_for_mode(glossary, plc_rules, translation_mode)
        st.caption("Progress")
        status = st.empty()
        job_id = start_queued_document_translation_job(
            raw_document,
            uploaded_document.name,
            active_glossary,
            translation_mode,
            translation_direction,
            keep_source_with_translation,
            notify_email,
            owner_session_id,
        )
        st.session_state["active_document_job_id"] = job_id
        status.success("Queued. Preparing file in the background.")
        rerun_app()

    if st.session_state.get("translated_document_bytes"):
        render_download_ready(
            data=st.session_state["translated_document_bytes"],
            file_name=st.session_state["translated_document_name"],
            mime=st.session_state["translated_document_mime"],
        )
        active_glossary = glossary_for_mode(glossary, plc_rules, translation_mode)
        render_translation_pairs_preview(
            raw_document,
            uploaded_document.name,
            translation_mode,
            translation_direction,
            owner_session_id,
            glossary=active_glossary,
        )

def main() -> None:
    load_env()

    st.set_page_config(page_title="Manufacturing AI Translation Platform (JP ↔ EN)", layout="wide")
    owner_session_id = current_client_session_id()
    apply_compact_style()
    usage_count = increment_usage_session_once()
    st.title("Manufacturing AI Translation Platform (JP ↔ EN)")

    with st.sidebar:
        render_sidebar_logo()
        usage_card_slot = st.empty()
        with usage_card_slot:
            render_usage_card(usage_count)
        st.caption(f"AI Model: {openai_model()}")
        with st.expander("Controlled Glossary"):
            st.caption(glossary_version_text().replace("\n", "  \n"))
        render_feedback_support()
        render_sidebar_footer()

    try:
        glossary = normalize_glossary(read_glossary(None))
    except Exception as exc:
        st.error(f"Glossary error: {exc}")
        st.stop()

    plc_rules_error = ""
    try:
        plc_rules = normalize_plc_rules(read_plc_rules())
    except Exception as exc:
        plc_rules_error = str(exc)
        plc_rules = empty_terms_dataframe()

    if plc_rules_error:
        st.warning(f"PLC rule file could not be loaded, so the app will continue without PLC abbreviation rules: {plc_rules_error}")

    text_tab, document_tab, hmi_tab = st.tabs(["Text", "Document", "Image"])

    with text_tab:
        render_text_translation(glossary, plc_rules)

    with document_tab:
        render_document_translation(glossary, plc_rules, owner_session_id)

    with hmi_tab:
        render_hmi_translation(glossary, plc_rules)

    with usage_card_slot:
        render_usage_card(read_usage_count())


if __name__ == "__main__":
    main()
