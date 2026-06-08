import csv
import mimetypes
import html
import io
import hashlib
import json
import os
import re
import smtplib
import sqlite3
import time
import unicodedata
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from urllib.parse import quote
from zipfile import BadZipFile, ZipFile, ZIP_DEFLATED
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from dotenv import load_dotenv
from openpyxl import load_workbook
from openai import APIConnectionError, APIStatusError, AuthenticationError, OpenAI, RateLimitError

try:
    from langsmith.wrappers import wrap_openai
except ImportError:
    wrap_openai = None


BASE_DIR = Path(__file__).parent
ENV_PATH = BASE_DIR / "app.env"
DEFAULT_GLOSSARY_PATHS = [
    BASE_DIR / "glossary.xlsx",
    BASE_DIR / "glossary.csv",
]
DEFAULT_PLC_RULE_PATHS = [
    BASE_DIR / "plc_abbreviation_rules.xlsx",
    BASE_DIR / "plc_abbreviation_rules.csv",
]
DEFAULT_MODEL = "gpt-4.1-mini"
DOCUMENT_BATCH_SIZE = 120
MAX_PARALLEL_BATCHES = 3
MAX_TRANSLATION_RETRIES = 3
OPENAI_TIMEOUT_SECONDS = 120
MAX_UPLOAD_BYTES = 100 * 1024 * 1024
MAX_EMAIL_ATTACHMENT_BYTES = 20 * 1024 * 1024
EMAIL_DRAFT_RECOMMENDED_BYTES = 5 * 1024 * 1024
EMAIL_DRAFT_RECOMMENDED_BLOCKS = 1000
PROGRESS_DIR = BASE_DIR / ".term1_progress"
USAGE_COUNT_PATH = BASE_DIR / ".term1_usage_count.json"
JOB_DB_PATH = BASE_DIR / ".term1_jobs.db"
JOB_STORAGE_DIR = BASE_DIR / ".term1_job_storage"
JOB_UPLOAD_DIR = JOB_STORAGE_DIR / "uploads"
JOB_RESULT_DIR = JOB_STORAGE_DIR / "results"
GENERAL_TRANSLATION_MODE = "General Plant Translation"
PLC_TRANSLATION_MODE = "PLC/SPLC Comment Standardization"
TRANSLATION_MODES = [PLC_TRANSLATION_MODE, GENERAL_TRANSLATION_MODE]
PLC_DUPLICATE_STATUS_WORDS = [
    "ON",
    "OFF",
    "OK",
    "NG",
    "Mode",
    "Error",
    "Complete",
    "Confirm",
    "Request",
    "Command",
    "Present",
    "Absent",
]
PLC_SYNONYM_CLEANUPS = [
    (re.compile(r"\bpoor,\s*defective,\s*NG,\s*inoperative\b", re.IGNORECASE), "NG"),
    (re.compile(r"\bdefective,\s*NG\b", re.IGNORECASE), "NG"),
]
PROTECTED_PATTERN = re.compile(
    r"\b(?:[A-Z]{1,6}[-_]?\d{1,6}[A-Z]?|\d+[A-Z]{1,4}|[XYMDSZR][0-9]{1,5}|[A-Z]{2,}-[A-Z0-9-]+)\b"
)
LEADING_CODE_PATTERN = re.compile(r"^([A-Z]{1,6}[-_]?\d{1,6}[A-Z]?)(.*)$")

WORD_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
PPT_NS = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
EXCEL_NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
EXCEL_SERIALIZE_NAMESPACES = {
    "": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "mc": "http://schemas.openxmlformats.org/markup-compatibility/2006",
    "x14ac": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac",
    "xr": "http://schemas.microsoft.com/office/spreadsheetml/2014/revision",
    "xr2": "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2",
    "xr3": "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3",
}


@dataclass(frozen=True)
class TermHit:
    jp: str
    en: str
    count: int


@dataclass(frozen=True)
class TextBlock:
    location: str
    text: str


@dataclass
class TokenUsage:
    input_tokens: int = 0
    output_tokens: int = 0
    total_tokens: int = 0

    def add(self, other: "TokenUsage") -> None:
        self.input_tokens += other.input_tokens
        self.output_tokens += other.output_tokens
        self.total_tokens += other.total_tokens

    def display(self) -> str:
        if self.total_tokens <= 0:
            return "Token usage unavailable."
        return (
            f"Tokens: input {self.input_tokens:,}, "
            f"output {self.output_tokens:,}, total {self.total_tokens:,}."
        )


def load_env() -> None:
    if ENV_PATH.exists():
        load_dotenv(ENV_PATH)
    else:
        load_dotenv()


def clean_text(value: str) -> str:
    return unicodedata.normalize("NFKC", str(value)).strip()


def has_japanese_text(value: str) -> bool:
    return bool(re.search(r"[\u3040-\u30ff\u3400-\u9fff]", value))


def should_translate(text: str) -> bool:
    return has_japanese_text(clean_text(text))


def decode_document_text(raw: bytes) -> str:
    candidates = []
    has_utf16_bom = raw.startswith((b"\xff\xfe", b"\xfe\xff"))
    null_ratio = raw.count(b"\x00") / max(len(raw), 1)
    encodings = ["utf-8-sig", "utf-8", "cp932", "shift_jis", "euc_jp", "iso2022_jp"]
    if has_utf16_bom or null_ratio > 0.1:
        encodings = ["utf-16", "utf-16-le", "utf-16-be"] + encodings
    encodings.append("cp1252")

    for encoding in encodings:
        try:
            text = raw.decode(encoding)
        except UnicodeDecodeError:
            continue

        japanese_count = len(re.findall(r"[\u3040-\u30ff\u3400-\u9fff]", text))
        replacement_count = text.count("�")
        mojibake_count = sum(text.count(marker) for marker in ("ã", "Â", "Ã", "\x00"))
        control_count = sum(1 for char in text if ord(char) < 32 and char not in "\r\n\t")
        c1_count = sum(1 for char in text if 0x80 <= ord(char) <= 0x9F)
        ascii_count = sum(1 for char in text if ord(char) < 128)
        score = (
            japanese_count * 50
            + min(ascii_count, 200) * 0.01
            - replacement_count * 200
            - mojibake_count * 80
            - control_count * 60
            - c1_count * 40
        )
        candidates.append((score, encoding, text))

    if candidates:
        return max(candidates, key=lambda candidate: candidate[0])[2]

    return raw.decode("utf-8", errors="replace")


def document_fingerprint(file_name: str, raw: bytes) -> str:
    digest = hashlib.sha256(raw).hexdigest()[:16]
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(file_name).stem).strip("_") or "document"
    return f"{safe_name}_{len(raw)}_{digest}"


def checkpoint_path_for(file_name: str, raw: bytes, translation_mode: str = GENERAL_TRANSLATION_MODE) -> Path:
    mode_key = re.sub(r"[^A-Za-z0-9_.-]+", "_", translation_mode).strip("_").lower()
    return PROGRESS_DIR / f"{document_fingerprint(file_name, raw)}_{mode_key}.json"


def load_checkpoint(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    translations = data.get("translations", {})
    if not isinstance(translations, dict):
        return {}
    return {str(key): str(value) for key, value in translations.items()}


def save_checkpoint(path: Path, translations: dict[str, str]) -> None:
    PROGRESS_DIR.mkdir(exist_ok=True)
    payload = {
        "updated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "translations": translations,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def format_duration(seconds: float) -> str:
    seconds = max(int(seconds), 0)
    minutes, second = divmod(seconds, 60)
    hours, minute = divmod(minutes, 60)
    if hours:
        return f"{hours}h {minute}m"
    if minute:
        return f"{minute}m {second}s"
    return f"{second}s"


def format_file_size(size_bytes: int) -> str:
    if size_bytes >= 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.1f} MB"
    if size_bytes >= 1024:
        return f"{size_bytes / 1024:.0f} KB"
    return f"{size_bytes} B"


def estimate_remaining_time(done: int, total: int, elapsed_seconds: float) -> str:
    if done <= 0 or total <= 0 or done >= total or elapsed_seconds <= 0:
        return ""
    remaining_seconds = (elapsed_seconds / done) * max(total - done, 0)
    return format_duration(remaining_seconds)


def progress_text(done: int, total: int, elapsed_seconds: float | None = None) -> str:
    percent = 100 if total <= 0 else min(int((done / total) * 100), 100)
    text = f"{percent}%"
    if elapsed_seconds is not None:
        eta = estimate_remaining_time(done, total, elapsed_seconds)
        if eta:
            text += f" · ETA {eta}"
    return text


def elapsed_since_timestamp(timestamp_text: str) -> float | None:
    try:
        started_at = datetime.strptime(timestamp_text, "%Y-%m-%d %H:%M:%S")
    except (TypeError, ValueError):
        return None
    return max((datetime.now() - started_at).total_seconds(), 0)


def clean_office_xml_text(value: str) -> str:
    # Office XML files cannot contain most control characters.
    return "".join(
        char
        for char in str(value)
        if char in "\t\n\r" or ord(char) >= 32
    )


def rerun_app() -> None:
    rerun = getattr(st, "rerun", None) or getattr(st, "experimental_rerun", None)
    if rerun:
        rerun()


def read_usage_count() -> int:
    if not USAGE_COUNT_PATH.exists():
        return 0
    try:
        data = json.loads(USAGE_COUNT_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return 0
    return int(data.get("count", 0))


def increment_usage_count_once() -> int:
    if st.session_state.get("usage_counted"):
        return read_usage_count()

    count = read_usage_count() + 1
    payload = {
        "count": count,
        "updated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
    }
    USAGE_COUNT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    st.session_state["usage_counted"] = True
    return count


def init_job_store() -> None:
    with sqlite3.connect(JOB_DB_PATH) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS translation_jobs (
                job_id TEXT PRIMARY KEY,
                file_name TEXT NOT NULL,
                file_size_bytes INTEGER NOT NULL,
                status TEXT NOT NULL,
                total_blocks INTEGER DEFAULT 0,
                translatable_blocks INTEGER DEFAULT 0,
                completed_blocks INTEGER DEFAULT 0,
                total_batches INTEGER DEFAULT 0,
                completed_batches INTEGER DEFAULT 0,
                input_tokens INTEGER DEFAULT 0,
                output_tokens INTEGER DEFAULT 0,
                total_tokens INTEGER DEFAULT 0,
                result_file_name TEXT DEFAULT '',
                error_message TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                finished_at TEXT DEFAULT ''
            )
            """
        )
        columns = {
            row[1]
            for row in conn.execute("PRAGMA table_info(translation_jobs)").fetchall()
        }
        if "translation_mode" not in columns:
            conn.execute(
                "ALTER TABLE translation_jobs ADD COLUMN translation_mode TEXT DEFAULT ''"
            )
        for column_name in (
            "source_file_path",
            "result_file_path",
            "result_mime",
            "progress_message",
            "notify_email",
            "notification_status",
        ):
            if column_name not in columns:
                conn.execute(
                    f"ALTER TABLE translation_jobs ADD COLUMN {column_name} TEXT DEFAULT ''"
                )


def create_translation_job(
    file_name: str,
    file_size_bytes: int,
    total_blocks: int,
    translatable_blocks: int,
    total_batches: int,
    translation_mode: str = GENERAL_TRANSLATION_MODE,
    source_file_path: str = "",
    result_file_path: str = "",
    result_mime: str = "",
    notify_email: str = "",
    status: str = "running",
) -> str:
    init_job_store()
    job_id = f"job_{int(time.time())}_{hashlib.sha256(f'{file_name}:{file_size_bytes}:{time.time()}'.encode()).hexdigest()[:8]}"
    now = time.strftime("%Y-%m-%d %H:%M:%S")
    with sqlite3.connect(JOB_DB_PATH) as conn:
        conn.execute(
            """
            INSERT INTO translation_jobs (
                job_id, file_name, file_size_bytes, translation_mode, status,
                source_file_path, result_file_path, result_mime, total_blocks,
                translatable_blocks, completed_blocks, total_batches,
                completed_batches, notify_email, created_at, updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                job_id,
                file_name,
                file_size_bytes,
                translation_mode,
                status,
                source_file_path,
                result_file_path,
                result_mime,
                total_blocks,
                translatable_blocks,
                0,
                total_batches,
                0,
                notify_email,
                now,
                now,
            ),
        )
    return job_id


def update_translation_job(job_id: str, **fields) -> None:
    if not job_id or not fields:
        return
    init_job_store()
    allowed_fields = {
        "status",
        "completed_blocks",
        "completed_batches",
        "input_tokens",
        "output_tokens",
        "total_tokens",
        "result_file_name",
        "source_file_path",
        "result_file_path",
        "result_mime",
        "notify_email",
        "notification_status",
        "progress_message",
        "error_message",
        "finished_at",
    }
    updates = {key: value for key, value in fields.items() if key in allowed_fields}
    if not updates:
        return
    updates["updated_at"] = time.strftime("%Y-%m-%d %H:%M:%S")
    assignments = ", ".join(f"{key} = ?" for key in updates)
    values = list(updates.values()) + [job_id]
    with sqlite3.connect(JOB_DB_PATH) as conn:
        conn.execute(f"UPDATE translation_jobs SET {assignments} WHERE job_id = ?", values)


def recent_translation_jobs(limit: int = 10) -> pd.DataFrame:
    init_job_store()
    with sqlite3.connect(JOB_DB_PATH) as conn:
        return pd.read_sql_query(
            """
            SELECT
                job_id AS "Job ID",
                file_name AS "File",
                translation_mode AS "Mode",
                status AS "Status",
                completed_blocks || '/' || translatable_blocks AS "Blocks",
                completed_batches || '/' || total_batches AS "Batches",
                result_file_name AS "Result",
                updated_at AS "Updated"
            FROM translation_jobs
            ORDER BY created_at DESC
            LIMIT ?
            """,
            conn,
            params=(limit,),
        )


def translation_job_detail(job_id: str) -> pd.DataFrame:
    init_job_store()
    with sqlite3.connect(JOB_DB_PATH) as conn:
        return pd.read_sql_query(
            """
            SELECT
                job_id,
                file_name,
                file_size_bytes,
                translation_mode,
                source_file_path,
                result_file_path,
                result_mime,
                notify_email,
                notification_status,
                progress_message,
                status,
                total_blocks,
                translatable_blocks,
                completed_blocks,
                total_batches,
                completed_batches,
                input_tokens,
                output_tokens,
                total_tokens,
                result_file_name,
                error_message,
                created_at,
                updated_at,
                finished_at
            FROM translation_jobs
            WHERE job_id = ?
            """,
            conn,
            params=(job_id,),
        )


def is_safe_glossary_term(jp: str) -> bool:
    jp = clean_text(jp)
    if len(jp) < 2:
        return False
    if not has_japanese_text(jp):
        return False
    if PROTECTED_PATTERN.fullmatch(jp):
        return False
    return True


def empty_terms_dataframe() -> pd.DataFrame:
    return pd.DataFrame(columns=["JP", "EN", "Note", "Category", "Owner"])


def xlsx_to_dataframe(raw: bytes, sheet_index: int = 0) -> pd.DataFrame:
    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

    def column_index(cell_ref: str) -> int:
        letters = re.sub(r"[^A-Z]", "", cell_ref.upper())
        index = 0
        for letter in letters:
            index = index * 26 + (ord(letter) - ord("A") + 1)
        return max(index - 1, 0)

    with ZipFile(io.BytesIO(raw)) as archive:
        names = archive.namelist()
        shared_strings = []

        if "xl/sharedStrings.xml" in names:
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in root.findall("a:si", ns):
                shared_strings.append("".join(text.text or "" for text in item.findall(".//a:t", ns)))

        sheet_names = sorted(
            name for name in names if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )
        if not sheet_names:
            raise ValueError("No worksheet found in Excel file.")

        sheet_name = sheet_names[min(sheet_index, len(sheet_names) - 1)]
        root = ET.fromstring(archive.read(sheet_name))
        rows = []

        for row in root.findall("a:sheetData/a:row", ns):
            indexed_values = {}
            for cell in row.findall("a:c", ns):
                if cell.attrib.get("t") == "inlineStr":
                    value = "".join(text.text or "" for text in cell.findall(".//a:t", ns))
                else:
                    value_node = cell.find("a:v", ns)
                    value = "" if value_node is None else value_node.text or ""
                    if cell.attrib.get("t") == "s" and value.isdigit():
                        value = shared_strings[int(value)]
                ref = cell.attrib.get("r", "")
                indexed_values[column_index(ref)] = value
            values = [indexed_values.get(index, "") for index in range(max(indexed_values.keys(), default=-1) + 1)]
            if any(str(value).strip() for value in values):
                rows.append(values)

    if not rows:
        return pd.DataFrame(columns=["JP", "EN"])

    header = [str(value).strip() for value in rows[0]]
    width = len(header)
    normalized_rows = [(row + [""] * width)[:width] for row in rows[1:]]
    return pd.DataFrame(normalized_rows, columns=header)


def read_glossary(uploaded_file) -> pd.DataFrame:
    if uploaded_file is None:
        glossary_path = next((path for path in DEFAULT_GLOSSARY_PATHS if path.exists()), None)
        if glossary_path is None:
            expected = ", ".join(path.name for path in DEFAULT_GLOSSARY_PATHS)
            raise FileNotFoundError(f"No glossary file found. Expected one of: {expected}")
        raw = glossary_path.read_bytes()
        name = glossary_path.name
    else:
        raw = uploaded_file.getvalue()
        name = uploaded_file.name

    is_excel = raw[:2] == b"PK" or name.lower().endswith((".xlsx", ".xlsm", ".xls"))
    if is_excel:
        try:
            return pd.read_excel(io.BytesIO(raw), sheet_name=0)
        except ImportError:
            return xlsx_to_dataframe(raw)
        except BadZipFile as exc:
            raise ValueError("The glossary Excel file could not be opened.") from exc

    for encoding in ("utf-8-sig", "utf-8", "cp932", "shift_jis", "cp1252"):
        try:
            return pd.read_csv(io.BytesIO(raw), encoding=encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Could not read glossary file. Please use CSV UTF-8 or Excel format.")


def read_rules_file(path: Path) -> pd.DataFrame:
    raw = path.read_bytes()
    is_excel = raw[:2] == b"PK" or path.name.lower().endswith((".xlsx", ".xlsm", ".xls"))
    if is_excel:
        try:
            return pd.read_excel(io.BytesIO(raw), sheet_name=0)
        except ImportError:
            return xlsx_to_dataframe(raw)
        except BadZipFile as exc:
            raise ValueError(f"The rule Excel file could not be opened: {path.name}") from exc

    for encoding in ("utf-8-sig", "utf-8", "cp932", "shift_jis", "cp1252"):
        try:
            return pd.read_csv(io.BytesIO(raw), encoding=encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Could not read rule file: {path.name}")


def read_plc_rules() -> pd.DataFrame:
    rule_path = next((path for path in DEFAULT_PLC_RULE_PATHS if path.exists()), None)
    if rule_path is None:
        return empty_terms_dataframe()
    return read_rules_file(rule_path)


def normalize_glossary(df: pd.DataFrame) -> pd.DataFrame:
    aliases = {
        "jp": "JP",
        "japanese": "JP",
        "japanese term": "JP",
        "日本語": "JP",
        "en": "EN",
        "english": "EN",
        "english term": "EN",
        "英語": "EN",
        "note": "Note",
        "notes": "Note",
        "comment": "Note",
    }
    aliases.update(
        {
            "source japanese": "JP",
            "source": "JP",
            "japanese source": "JP",
            "japanese comment": "JP",
            "source term": "JP",
            "source text": "JP",
            "preferred english": "EN",
            "preferred abbreviation": "EN",
            "preferred en": "EN",
            "approved english": "EN",
            "approved abbreviation": "EN",
            "abbreviation": "EN",
            "standard english": "EN",
            "standard wording": "EN",
            "target": "EN",
            "target english": "EN",
            "do not use": "Note",
        }
    )

    renamed = {}
    for column in df.columns:
        key = str(column).strip().lower()
        renamed[column] = aliases.get(key, str(column).strip())

    glossary = df.rename(columns=renamed).fillna("")
    if "JP" not in glossary.columns or "EN" not in glossary.columns:
        raise ValueError("Glossary must include JP/Japanese and EN/English columns.")

    glossary = glossary[[column for column in ["JP", "EN", "Note", "Category", "Owner"] if column in glossary.columns]]
    glossary = glossary.copy()
    glossary["JP"] = glossary["JP"].astype(str).map(clean_text)
    glossary["EN"] = glossary["EN"].astype(str).map(clean_text)
    glossary = glossary[(glossary["JP"] != "") & (glossary["EN"] != "")]
    glossary = glossary[glossary["JP"].map(is_safe_glossary_term)]
    glossary = glossary.drop_duplicates(subset=["JP"], keep="first")
    glossary["term_length"] = glossary["JP"].str.len()
    return glossary.sort_values("term_length", ascending=False).drop(columns=["term_length"]).reset_index(drop=True)


def normalize_plc_rules(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty and not set(df.columns).intersection({"JP", "EN", "Japanese", "English"}):
        return empty_terms_dataframe()
    try:
        rules = normalize_glossary(df)
    except Exception:
        return empty_terms_dataframe()
    if rules.empty:
        return empty_terms_dataframe()
    rules = rules.copy()
    rules["Category"] = "PLC/SPLC Rule"
    return rules


def glossary_for_mode(glossary: pd.DataFrame, plc_rules: pd.DataFrame, translation_mode: str) -> pd.DataFrame:
    if translation_mode != PLC_TRANSLATION_MODE or plc_rules.empty:
        return glossary

    combined = pd.concat([plc_rules, glossary], ignore_index=True, sort=False).fillna("")
    combined = combined.drop_duplicates(subset=["JP"], keep="first")
    combined["term_length"] = combined["JP"].str.len()
    return combined.sort_values("term_length", ascending=False).drop(columns=["term_length"]).reset_index(drop=True)


def apply_glossary_to_source(text: str, glossary: pd.DataFrame) -> tuple[str, list[TermHit]]:
    translated_source = clean_text(text)
    hits = []

    for _, row in glossary.iterrows():
        jp = clean_text(row["JP"])
        en = clean_text(row["EN"])
        if not jp or jp not in translated_source:
            continue

        translated_source, count = re.subn(re.escape(jp), en, translated_source)
        if count:
            hits.append(TermHit(jp=jp, en=en, count=count))

    return translated_source, hits


def exact_controlled_term_match(text: str, glossary: pd.DataFrame) -> tuple[str | None, list[TermHit]]:
    source = clean_text(text)
    if not source:
        return None, []

    for _, row in glossary.iterrows():
        jp = clean_text(row["JP"])
        en = clean_text(row["EN"])
        if source == jp:
            return en, [TermHit(jp=jp, en=en, count=1)]
    return None, []


def find_protected_codes(text: str) -> list[str]:
    return sorted(set(PROTECTED_PATTERN.findall(text)))


def plc_mode_rules() -> str:
    return """
PLC/SPLC comment mode:
1. Treat the source as PLC/SPLC device comments, HMI labels, alarm labels, or control logic comments, not normal prose.
2. Output short engineering labels only. Avoid full sentences unless the source is clearly a sentence.
3. Do not string together multiple synonyms. Never output lists like "poor, defective, NG, inoperative".
4. Choose one stable plant-control term for each Japanese concept. Prefer concise PLC terms such as ON, OFF, OK, NG, Present, Absent, Complete, Confirm, Request, Command, Auto, Manual, Standby, Error.
5. Preserve PLC addresses, device IDs, robot names, station names, prefixes, symbols, arrows, brackets, and separators exactly.
6. Keep repeated Japanese patterns translated with repeated English patterns.
7. If a company glossary term is provided, it overrides the default PLC wording.
""".strip()


def general_mode_rules() -> str:
    return """
General plant translation mode:
1. Translate into clear, natural plant-floor engineering English for manufacturing users.
2. Use concise plant-floor English suitable for controls, seibi, production, and engineering users.
3. Preserve line breaks and list structure when useful.
""".strip()


def normalize_plc_translation_line(line: str) -> str:
    normalized = re.sub(r"\s+", " ", line).strip()
    for pattern, replacement in PLC_SYNONYM_CLEANUPS:
        normalized = pattern.sub(replacement, normalized)

    changed = True
    while changed:
        changed = False
        for word in PLC_DUPLICATE_STATUS_WORDS:
            pattern = re.compile(rf"\b({re.escape(word)})\s+\1\b", re.IGNORECASE)
            normalized, count = pattern.subn(r"\1", normalized)
            if count:
                changed = True

    return normalized


def post_process_translation(output_text: str, translation_mode: str) -> str:
    cleaned = output_text.strip()
    if translation_mode != PLC_TRANSLATION_MODE:
        return cleaned
    return "\n".join(
        normalize_plc_translation_line(line)
        for line in cleaned.splitlines()
    ).strip()


def build_prompt(source_text: str, hits: list[TermHit], protected_codes: list[str], translation_mode: str) -> str:
    terms = "\n".join(f"{hit.jp} = {hit.en}" for hit in hits)
    codes = ", ".join(protected_codes) if protected_codes else "None detected"
    mode_rules = plc_mode_rules() if translation_mode == PLC_TRANSLATION_MODE else general_mode_rules()

    return f"""
You are a professional Japanese-to-English translator for a battery manufacturing plant.

Translation mode: {translation_mode}

{mode_rules}

Mandatory rules:
1. Apply approved company glossary terms exactly as written. Do not paraphrase approved terms.
2. Preserve PLC addresses, device IDs, model names, station IDs, alarm codes, part numbers, and equipment codes exactly.
3. Do not invent missing information, causes, actions, measurements, or context that is not in the source.
4. If the source is ambiguous, translate only the meaning that is present and keep the wording neutral.
5. Output only the English translation. Do not add explanations, notes, or commentary.

Company terminology detected in the source:
{terms if terms else "None"}

Protected codes detected:
{codes}

Source Japanese text:
{source_text}
""".strip()


def openai_client() -> OpenAI:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError(
            "OPENAI_API_KEY was not found. Add it to app.env for local development "
            "or configure it as a secret/environment variable in the deployment environment."
        )
    base_url = os.getenv("OPENAI_BASE_URL")
    if base_url:
        client = OpenAI(api_key=api_key, base_url=base_url)
    else:
        client = OpenAI(api_key=api_key)

    tracing_enabled = os.getenv("LANGSMITH_TRACING", "").lower() in {"1", "true", "yes"}
    if tracing_enabled and wrap_openai is not None:
        return wrap_openai(client)
    return client


def openai_model() -> str:
    return os.getenv("OPENAI_MODEL", DEFAULT_MODEL)


def openai_timeout_seconds() -> float:
    try:
        return float(os.getenv("OPENAI_TIMEOUT_SECONDS", OPENAI_TIMEOUT_SECONDS))
    except ValueError:
        return float(OPENAI_TIMEOUT_SECONDS)


def max_parallel_batches() -> int:
    try:
        value = int(os.getenv("MAX_PARALLEL_BATCHES", MAX_PARALLEL_BATCHES))
    except ValueError:
        value = MAX_PARALLEL_BATCHES
    return max(value, 1)


def response_token_usage(response) -> TokenUsage:
    usage = getattr(response, "usage", None)
    if usage is None:
        return TokenUsage()

    def usage_value(*names: str) -> int:
        for name in names:
            value = getattr(usage, name, None)
            if value is not None:
                return int(value)
        if isinstance(usage, dict):
            for name in names:
                value = usage.get(name)
                if value is not None:
                    return int(value)
        return 0

    input_tokens = usage_value("input_tokens", "prompt_tokens")
    output_tokens = usage_value("output_tokens", "completion_tokens")
    total_tokens = usage_value("total_tokens")
    if total_tokens == 0:
        total_tokens = input_tokens + output_tokens
    return TokenUsage(
        input_tokens=input_tokens,
        output_tokens=output_tokens,
        total_tokens=total_tokens,
    )


def format_translation_error(exc: Exception) -> str:
    if isinstance(exc, APIConnectionError):
        return (
            "Connection error while calling the GPT API. Check whether OPENAI_BASE_URL is required "
            "for the company API, and confirm VPN/proxy/firewall access from this computer."
        )
    if isinstance(exc, AuthenticationError):
        return "Authentication failed. Check that OPENAI_API_KEY is correct and approved for this endpoint."
    if isinstance(exc, RateLimitError):
        return "Rate limit or quota exceeded. Check the API quota, rate limit, or billing/usage policy."
    if isinstance(exc, APIStatusError):
        return f"GPT API returned HTTP {exc.status_code}. {exc.message}"
    return str(exc)


def translate_text(source_text: str, hits: list[TermHit], protected_codes: list[str], translation_mode: str) -> tuple[str, TokenUsage]:
    client = openai_client()
    response = client.responses.create(
        model=openai_model(),
        input=build_prompt(source_text, hits, protected_codes, translation_mode),
        temperature=0.1,
        timeout=openai_timeout_seconds(),
    )
    return post_process_translation(response.output_text, translation_mode), response_token_usage(response)


def translate_block(text: str, glossary: pd.DataFrame, translation_mode: str) -> tuple[str, list[TermHit], TokenUsage]:
    exact_translation, exact_hits = exact_controlled_term_match(text, glossary)
    if exact_translation is not None:
        return post_process_translation(exact_translation, translation_mode), exact_hits, TokenUsage()

    glossary_applied_text, hits = apply_glossary_to_source(text, glossary)
    protected_codes = find_protected_codes(text)
    translation, token_usage = translate_text(glossary_applied_text, hits, protected_codes, translation_mode)
    return translation, hits, token_usage


def build_batch_prompt(items: list[tuple[int, str, list[TermHit], list[str]]], translation_mode: str) -> str:
    item_text = "\n\n".join(
        "\n".join(
            [
                f"[BLOCK {item_id}]",
                source_text,
                f"[/BLOCK {item_id}]",
            ]
        )
        for item_id, source_text, _, _ in items
    )
    terms = []
    codes = []
    for _, _, hits, protected_codes in items:
        terms.extend(f"{hit.jp} = {hit.en}" for hit in hits)
        codes.extend(protected_codes)

    unique_terms = "\n".join(sorted(set(terms))) or "None"
    unique_codes = ", ".join(sorted(set(codes))) if codes else "None detected"
    mode_rules = plc_mode_rules() if translation_mode == PLC_TRANSLATION_MODE else general_mode_rules()

    return f"""
You are a professional Japanese-to-English translator for a battery manufacturing plant.

Translation mode: {translation_mode}

{mode_rules}

Mandatory rules:
1. Apply approved company glossary terms exactly as written. Do not paraphrase approved terms.
2. Preserve PLC addresses, device IDs, model names, station IDs, alarm codes, part numbers, and equipment codes exactly.
3. Do not invent missing information, causes, actions, measurements, or context that is not in the source.
4. If the source is ambiguous, translate only the meaning that is present and keep the wording neutral.
5. Return each translated block using the same markers and do not add explanations, notes, or commentary:
[BLOCK 1]
English translation
[/BLOCK 1]

Company terminology detected:
{unique_terms}

Protected codes detected:
{unique_codes}

Source blocks:
{item_text}
""".strip()


def parse_batch_translation(output_text: str, item_ids: list[int]) -> dict[int, str]:
    translations = {}
    for item_id in item_ids:
        pattern = re.compile(
            rf"\[BLOCK {item_id}\]\s*(.*?)\s*\[/BLOCK {item_id}\]",
            re.DOTALL,
        )
        match = pattern.search(output_text)
        if match:
            translations[item_id] = match.group(1).strip()
    return translations


def translate_batch_chunk(chunk: list[TextBlock], glossary: pd.DataFrame, translation_mode: str) -> tuple[dict[str, str], list[TermHit], TokenUsage]:
    items = []
    chunk_hits = []
    direct_translations = {}

    for offset, block in enumerate(chunk, start=1):
        exact_translation, exact_hits = exact_controlled_term_match(block.text, glossary)
        if exact_translation is not None:
            direct_translations[block.location] = post_process_translation(exact_translation, translation_mode)
            chunk_hits.extend(exact_hits)
            continue

        glossary_applied_text, hits = apply_glossary_to_source(block.text, glossary)
        protected_codes = find_protected_codes(block.text)
        items.append((offset, glossary_applied_text, hits, protected_codes))
        chunk_hits.extend(hits)

    parsed = {}
    token_usage = TokenUsage()
    last_error = None
    if items:
        client = openai_client()
        for attempt in range(1, MAX_TRANSLATION_RETRIES + 1):
            try:
                response = client.responses.create(
                    model=openai_model(),
                    input=build_batch_prompt(items, translation_mode),
                    temperature=0.1,
                    timeout=openai_timeout_seconds(),
                )
                token_usage.add(response_token_usage(response))
                parsed = parse_batch_translation(response.output_text.strip(), [item[0] for item in items])
                missing_ids = [item[0] for item in items if item[0] not in parsed]
                if missing_ids:
                    raise ValueError(f"Translation response missed block marker(s): {missing_ids}")
                break
            except Exception as exc:
                last_error = exc
                if attempt == MAX_TRANSLATION_RETRIES:
                    raise
                time.sleep(5 * attempt)

    if items and not parsed and last_error:
        raise last_error

    chunk_translations = dict(direct_translations)
    for offset, block in enumerate(chunk, start=1):
        if block.location in chunk_translations:
            continue
        chunk_translations[block.location] = post_process_translation(parsed[offset], translation_mode)

    return chunk_translations, chunk_hits, token_usage


def translate_batch_chunk_resilient(
    chunk: list[TextBlock],
    glossary: pd.DataFrame,
    translation_mode: str,
) -> tuple[dict[str, str], list[TermHit], TokenUsage]:
    try:
        return translate_batch_chunk(chunk, glossary, translation_mode)
    except Exception:
        if len(chunk) <= 1:
            raise

    midpoint = max(len(chunk) // 2, 1)
    combined_translations = {}
    combined_hits = []
    combined_usage = TokenUsage()
    for part in (chunk[:midpoint], chunk[midpoint:]):
        part_translations, part_hits, part_usage = translate_batch_chunk_resilient(
            part,
            glossary,
            translation_mode,
        )
        combined_translations.update(part_translations)
        combined_hits.extend(part_hits)
        combined_usage.add(part_usage)
    return combined_translations, combined_hits, combined_usage


def translation_memory_key(text: str) -> str:
    return clean_text(text)


def translate_blocks_batch(
    blocks: list[TextBlock],
    glossary: pd.DataFrame,
    translation_mode: str,
    checkpoint_path=None,
    progress_callback=None,
) -> tuple[dict[str, str], list[TermHit], TokenUsage]:
    translations = load_checkpoint(checkpoint_path) if checkpoint_path else {}
    translatable_blocks = [block for block in blocks if should_translate(block.text)]
    source_memory = {}
    for block in translatable_blocks:
        if block.location in translations:
            source_memory.setdefault(translation_memory_key(block.text), translations[block.location])

    for block in translatable_blocks:
        key = translation_memory_key(block.text)
        if block.location not in translations and key in source_memory:
            translations[block.location] = source_memory[key]

    pending_by_key = {}
    duplicate_locations_by_key = {}
    for block in translatable_blocks:
        if block.location in translations:
            continue
        key = translation_memory_key(block.text)
        if key not in pending_by_key:
            pending_by_key[key] = block
            duplicate_locations_by_key[key] = []
        duplicate_locations_by_key[key].append(block.location)

    pending_blocks = list(pending_by_key.values())
    all_hits = []
    token_usage = TokenUsage()
    started_at = time.time()
    pending_location_count = sum(len(locations) for locations in duplicate_locations_by_key.values())
    completed_at_start = len(translatable_blocks) - pending_location_count
    total_batches = (len(pending_blocks) + DOCUMENT_BATCH_SIZE - 1) // DOCUMENT_BATCH_SIZE
    parallel_batches = min(max_parallel_batches(), max(total_batches, 1))

    if progress_callback:
        progress_callback(
            completed_at_start,
            len(translatable_blocks),
            0,
            total_batches,
            0,
            "Resuming from saved progress and translation memory." if completed_at_start else "Starting translation.",
        )

    chunks = [
        pending_blocks[start:start + DOCUMENT_BATCH_SIZE]
        for start in range(0, len(pending_blocks), DOCUMENT_BATCH_SIZE)
    ]
    completed_batches = 0
    completed_blocks = completed_at_start

    if chunks:
        with ThreadPoolExecutor(max_workers=parallel_batches) as executor:
            future_to_chunk = {
                executor.submit(translate_batch_chunk_resilient, chunk, glossary, translation_mode): chunk
                for chunk in chunks
            }

            for future in as_completed(future_to_chunk):
                chunk = future_to_chunk[future]
                chunk_translations, chunk_hits, chunk_token_usage = future.result()
                expanded_translations = {}
                for block in chunk:
                    key = translation_memory_key(block.text)
                    translated_text = chunk_translations[block.location]
                    source_memory[key] = translated_text
                    for location in duplicate_locations_by_key.get(key, [block.location]):
                        expanded_translations[location] = translated_text

                translations.update(expanded_translations)
                all_hits.extend(chunk_hits)
                token_usage.add(chunk_token_usage)
                completed_batches += 1
                completed_blocks += len(expanded_translations)

                if checkpoint_path:
                    save_checkpoint(checkpoint_path, translations)

                if progress_callback:
                    progress_callback(
                        min(completed_blocks, len(translatable_blocks)),
                        len(translatable_blocks),
                        completed_batches,
                        total_batches,
                        time.time() - started_at,
                        f"Saved batch progress. Running up to {parallel_batches} batch(es) in parallel.",
                    )

    return translations, all_hits, token_usage


def read_text_file(raw: bytes) -> str:
    return decode_document_text(raw)


def write_text_file(blocks: list[TextBlock], translations: dict[str, str]) -> bytes:
    lines = []
    for block in blocks:
        lines.append(translations.get(block.location, block.text))
    return "\n\n".join(lines).encode("utf-8-sig")


def sniff_csv_dialect(text: str):
    sample = "\n".join(text.splitlines()[:100])
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        return csv.excel


def split_leading_code_cell(value: str) -> list[str]:
    cleaned = clean_text(value)
    match = LEADING_CODE_PATTERN.match(cleaned)
    if not match:
        return [value]
    code, remainder = match.groups()
    remainder = clean_text(remainder)
    if not remainder:
        return [value]
    return [code, remainder]


def normalize_csv_structure(rows: list[list[str]]) -> list[list[str]]:
    normalized = []
    for row in rows:
        if len(row) == 1:
            normalized.append(split_leading_code_cell(row[0]))
        else:
            normalized.append(row)
    return normalized


def read_csv_rows(raw: bytes) -> list[list[str]]:
    text = decode_document_text(raw)
    return parse_csv_rows_lenient(text)


def parse_csv_rows_lenient(text: str) -> list[list[str]]:
    if "�" in text and not has_japanese_text(text):
        raise ValueError(
            "The CSV text could not be decoded into readable Japanese. "
            "Please save the CSV as UTF-8 CSV or Excel .xlsx, then upload again."
        )
    dialect = sniff_csv_dialect(text)
    try:
        return normalize_csv_structure([row for row in csv.reader(io.StringIO(text, newline=""), dialect)])
    except csv.Error:
        rows = []
        for line in text.splitlines():
            try:
                rows.append(next(csv.reader([line], dialect)))
            except csv.Error:
                rows.append([line])
        return normalize_csv_structure(rows)


def extract_csv_blocks(raw: bytes) -> list[TextBlock]:
    rows = read_csv_rows(raw)
    blocks = []
    for row_index, row in enumerate(rows):
        for column_index, cell in enumerate(row):
            value = clean_text(cell)
            if value:
                blocks.append(TextBlock(location=f"csv:{row_index}:{column_index}", text=value))
    return blocks


def build_translated_csv(raw: bytes, translations: dict[str, str]) -> bytes:
    rows = read_csv_rows(raw)
    for row_index, row in enumerate(rows):
        for column_index, _ in enumerate(row):
            key = f"csv:{row_index}:{column_index}"
            if key in translations:
                row[column_index] = translations[key]

    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")


def extract_text_blocks(raw: bytes, file_name: str) -> list[TextBlock]:
    lower_name = file_name.lower()
    if lower_name.endswith((".txt", ".as")):
        text = read_text_file(raw)
        return [TextBlock(location=f"text:{index}", text=part.strip()) for index, part in enumerate(text.split("\n\n")) if part.strip()]

    if lower_name.endswith(".csv"):
        return extract_csv_blocks(raw)

    if lower_name.endswith(".docx"):
        return extract_docx_blocks(raw)

    if lower_name.endswith((".xlsx", ".xlsm")):
        return extract_xlsx_blocks(raw)

    raise ValueError("Supported document types: CSV, TXT, AS, DOCX, XLSX, XLSM.")


def extract_docx_blocks(raw: bytes) -> list[TextBlock]:
    blocks = []
    with ZipFile(io.BytesIO(raw)) as archive:
        xml_names = [
            name for name in archive.namelist()
            if name == "word/document.xml" or name.startswith("word/header") or name.startswith("word/footer")
        ]
        for xml_name in xml_names:
            root = ET.fromstring(archive.read(xml_name))
            for index, paragraph in enumerate(root.findall(".//w:p", WORD_NS)):
                text = "".join(node.text or "" for node in paragraph.findall(".//w:t", WORD_NS)).strip()
                if text:
                    blocks.append(TextBlock(location=f"{xml_name}#{index}", text=text))
    return blocks


def extract_pptx_blocks(raw: bytes) -> list[TextBlock]:
    blocks = []
    with ZipFile(io.BytesIO(raw)) as archive:
        slide_names = sorted(
            name for name in archive.namelist()
            if name.startswith("ppt/slides/slide") and name.endswith(".xml")
        )
        for slide_name in slide_names:
            root = ET.fromstring(archive.read(slide_name))
            for index, paragraph in enumerate(root.findall(".//a:p", PPT_NS)):
                text = "".join(node.text or "" for node in paragraph.findall(".//a:t", PPT_NS)).strip()
                if text:
                    blocks.append(TextBlock(location=f"{slide_name}#{index}", text=text))
    return blocks


def read_shared_strings(archive: ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []

    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    strings = []
    for item in root.findall("a:si", EXCEL_NS):
        strings.append("".join(text.text or "" for text in item.findall(".//a:t", EXCEL_NS)))
    return strings


def cell_text(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(text.text or "" for text in cell.findall(".//a:t", EXCEL_NS)).strip()

    value_node = cell.find("a:v", EXCEL_NS)
    if value_node is None or value_node.text is None:
        return ""

    value = value_node.text
    if cell_type == "s" and value.isdigit():
        index = int(value)
        if 0 <= index < len(shared_strings):
            return shared_strings[index].strip()

    if cell_type in {"str", "e"}:
        return value.strip()

    return ""


def extract_xlsx_blocks(raw: bytes) -> list[TextBlock]:
    blocks = []
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=False)
        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            sheet_name = f"xl/worksheets/sheet{sheet_index}.xml"
            for row in worksheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.strip():
                        blocks.append(TextBlock(location=f"{sheet_name}#{cell.coordinate}", text=clean_text(value)))
        workbook.close()
        if blocks:
            return blocks
    except Exception:
        blocks = []

    with ZipFile(io.BytesIO(raw)) as archive:
        shared_strings = read_shared_strings(archive)
        sheet_names = sorted(
            name for name in archive.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )

        for sheet_name in sheet_names:
            root = ET.fromstring(archive.read(sheet_name))
            for cell in root.findall(".//a:c", EXCEL_NS):
                ref = cell.attrib.get("r", "")
                text = cell_text(cell, shared_strings)
                if text:
                    blocks.append(TextBlock(location=f"{sheet_name}#{ref}", text=text))
    return blocks


def replace_text_in_paragraph(paragraph: ET.Element, text_nodes: list[ET.Element], translated_text: str) -> None:
    if not text_nodes:
        return
    text_nodes[0].text = clean_office_xml_text(translated_text)
    for node in text_nodes[1:]:
        node.text = ""


def build_translated_document(raw: bytes, file_name: str, translations: dict[str, str], blocks: list[TextBlock]) -> bytes:
    lower_name = file_name.lower()
    if lower_name.endswith((".txt", ".as")):
        return write_text_file(blocks, translations)

    if lower_name.endswith(".csv"):
        return build_translated_csv(raw, translations)

    if lower_name.endswith(".docx"):
        return build_translated_docx(raw, translations)

    if lower_name.endswith((".xlsx", ".xlsm")):
        return build_translated_xlsx(raw, translations)

    raise ValueError("Supported document types: CSV, TXT, AS, DOCX, XLSX, XLSM.")


def build_translated_docx(raw: bytes, translations: dict[str, str]) -> bytes:
    source = io.BytesIO(raw)
    target = io.BytesIO()

    with ZipFile(source) as input_zip, ZipFile(target, "w", ZIP_DEFLATED) as output_zip:
        for item in input_zip.infolist():
            data = input_zip.read(item.filename)
            if item.filename == "word/document.xml" or item.filename.startswith("word/header") or item.filename.startswith("word/footer"):
                root = ET.fromstring(data)
                for index, paragraph in enumerate(root.findall(".//w:p", WORD_NS)):
                    key = f"{item.filename}#{index}"
                    if key in translations:
                        replace_text_in_paragraph(paragraph, paragraph.findall(".//w:t", WORD_NS), translations[key])
                data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            output_zip.writestr(item, data)

    return target.getvalue()


def build_translated_pptx(raw: bytes, translations: dict[str, str]) -> bytes:
    source = io.BytesIO(raw)
    target = io.BytesIO()

    with ZipFile(source) as input_zip, ZipFile(target, "w", ZIP_DEFLATED) as output_zip:
        for item in input_zip.infolist():
            data = input_zip.read(item.filename)
            if item.filename.startswith("ppt/slides/slide") and item.filename.endswith(".xml"):
                root = ET.fromstring(data)
                for index, paragraph in enumerate(root.findall(".//a:p", PPT_NS)):
                    key = f"{item.filename}#{index}"
                    if key in translations:
                        replace_text_in_paragraph(paragraph, paragraph.findall(".//a:t", PPT_NS), translations[key])
                data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            output_zip.writestr(item, data)

    return target.getvalue()


def replace_excel_cell_text(cell: ET.Element, translated_text: str) -> None:
    for child in list(cell):
        cell.remove(child)

    cell.attrib["t"] = "inlineStr"
    inline_string = ET.SubElement(cell, f"{{{EXCEL_NS['a']}}}is")
    text_node = ET.SubElement(inline_string, f"{{{EXCEL_NS['a']}}}t")
    text_node.text = clean_office_xml_text(translated_text)


def serialize_excel_xml(root: ET.Element) -> bytes:
    used_uris = set()
    for element in root.iter():
        if element.tag.startswith("{"):
            used_uris.add(element.tag[1:].split("}", 1)[0])
        for attr_name in element.attrib:
            if attr_name.startswith("{"):
                used_uris.add(attr_name[1:].split("}", 1)[0])

    ignorable_attr = "{http://schemas.openxmlformats.org/markup-compatibility/2006}Ignorable"
    if ignorable_attr in root.attrib:
        kept_prefixes = [
            prefix
            for prefix in root.attrib[ignorable_attr].split()
            if EXCEL_SERIALIZE_NAMESPACES.get(prefix) in used_uris
        ]
        if kept_prefixes:
            root.attrib[ignorable_attr] = " ".join(kept_prefixes)
        else:
            root.attrib.pop(ignorable_attr, None)

    for prefix, uri in EXCEL_SERIALIZE_NAMESPACES.items():
        ET.register_namespace(prefix, uri)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def build_translated_xlsx(raw: bytes, translations: dict[str, str]) -> bytes:
    source = io.BytesIO(raw)
    target = io.BytesIO()

    with ZipFile(source) as input_zip, ZipFile(target, "w", ZIP_DEFLATED) as output_zip:
        for item in input_zip.infolist():
            data = input_zip.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet") and item.filename.endswith(".xml"):
                root = ET.fromstring(data)
                for cell in root.findall(".//a:c", EXCEL_NS):
                    key = f"{item.filename}#{cell.attrib.get('r', '')}"
                    if key in translations:
                        replace_excel_cell_text(cell, translations[key])
                data = serialize_excel_xml(root)
            output_zip.writestr(item, data)

    return target.getvalue()


def output_file_name(file_name: str) -> str:
    path = Path(file_name)
    return f"Translated-{path.stem}{path.suffix or '.txt'}"


def mime_type(file_name: str) -> str:
    lower_name = file_name.lower()
    if lower_name.endswith(".csv"):
        return "text/csv"
    if lower_name.endswith(".as"):
        return "text/plain"
    if lower_name.endswith(".docx"):
        return "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    if lower_name.endswith(".xlsx"):
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if lower_name.endswith(".xlsm"):
        return "application/vnd.ms-excel.sheet.macroEnabled.12"
    return "text/plain"


def safe_storage_name(file_name: str) -> str:
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(file_name).name).strip("_")
    return safe_name or "document.txt"


def ensure_job_storage_dirs() -> None:
    JOB_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    JOB_RESULT_DIR.mkdir(parents=True, exist_ok=True)


def job_upload_path(job_id: str, file_name: str) -> Path:
    ensure_job_storage_dirs()
    return JOB_UPLOAD_DIR / f"{job_id}_{safe_storage_name(file_name)}"


def job_result_path(job_id: str, file_name: str) -> Path:
    ensure_job_storage_dirs()
    translated_name = output_file_name(file_name)
    return JOB_RESULT_DIR / f"{job_id}_{safe_storage_name(translated_name)}"


def is_valid_email(value: str) -> bool:
    return bool(re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", value.strip()))


def is_smtp_configured() -> bool:
    smtp_host = os.getenv("SMTP_HOST", "").strip()
    smtp_from = os.getenv("SMTP_FROM", os.getenv("SMTP_USERNAME", "")).strip()
    return bool(smtp_host and smtp_from)


def manual_email_status() -> str:
    return "Manual email draft ready. Download the translated file and attach it yourself."


def translation_mailto_link(to_email: str, file_name: str, result_file_name: str) -> str:
    subject = f"Term1 translation completed: {result_file_name}"
    body = "\n".join(
        [
            "Your Term1 translation job is complete.",
            "",
            f"Source file: {file_name}",
            f"Translated file: {result_file_name}",
            "",
            "Please attach the downloaded translated file before sending this email.",
        ]
    )
    return f"mailto:{quote(to_email)}?subject={quote(subject)}&body={quote(body)}"


def send_completed_translation_email(to_email: str, file_name: str, result_path: Path, result_file_name: str) -> str:
    if not to_email:
        return ""
    if not is_valid_email(to_email):
        return "Invalid email address."

    smtp_host = os.getenv("SMTP_HOST", "").strip()
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USERNAME", "").strip()
    smtp_password = os.getenv("SMTP_PASSWORD", "")
    smtp_from = os.getenv("SMTP_FROM", smtp_user).strip()
    smtp_tls = os.getenv("SMTP_TLS", "true").lower() in {"1", "true", "yes"}

    if not smtp_host or not smtp_from:
        return manual_email_status()
    if not result_path.exists():
        return "Email not sent: translated file was not found."
    if result_path.stat().st_size > MAX_EMAIL_ATTACHMENT_BYTES:
        return "Email not sent: translated file is larger than the email attachment limit. Download it from the app."

    message = EmailMessage()
    message["From"] = smtp_from
    message["To"] = to_email
    message["Subject"] = f"Term1 translation completed: {result_file_name}"
    message.set_content(
        "\n".join(
            [
                "Your Term1 translation job is complete.",
                "",
                f"Source file: {file_name}",
                f"Translated file: {result_file_name}",
                "",
                "The translated file is attached.",
            ]
        )
    )

    content_type, _ = mimetypes.guess_type(result_file_name)
    maintype, subtype = (content_type or "application/octet-stream").split("/", 1)
    message.add_attachment(
        result_path.read_bytes(),
        maintype=maintype,
        subtype=subtype,
        filename=result_file_name,
    )

    with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
        if smtp_tls:
            server.starttls()
        if smtp_user and smtp_password:
            server.login(smtp_user, smtp_password)
        server.send_message(message)

    return "Email sent."


@st.cache_resource
def background_job_executor() -> ThreadPoolExecutor:
    return ThreadPoolExecutor(max_workers=2)


def run_document_translation_job(
    job_id: str,
    raw_document: bytes,
    file_name: str,
    blocks: list[TextBlock],
    glossary: pd.DataFrame,
    translation_mode: str,
    checkpoint_path: Path,
    total_batches: int,
) -> None:
    started_at = time.time()

    def update_progress(done, total, done_batches, total_batches, elapsed, message):
        update_translation_job(
            job_id,
            status="running",
            completed_blocks=done,
            completed_batches=done_batches,
            progress_message=message,
        )

    try:
        update_translation_job(job_id, status="running", error_message="", progress_message="Background translation started.")
        translations, _, token_usage = translate_blocks_batch(
            blocks,
            glossary,
            translation_mode,
            checkpoint_path=checkpoint_path,
            progress_callback=update_progress,
        )
        translated_document = build_translated_document(raw_document, file_name, translations, blocks)
        translated_name = output_file_name(file_name)
        result_path = job_result_path(job_id, file_name)
        result_path.write_bytes(translated_document)
        notify_email = ""
        with sqlite3.connect(JOB_DB_PATH) as conn:
            row = conn.execute(
                "SELECT notify_email FROM translation_jobs WHERE job_id = ?",
                (job_id,),
            ).fetchone()
            notify_email = row[0] if row and row[0] else ""
        notification_status = ""
        if notify_email:
            try:
                notification_status = send_completed_translation_email(
                    notify_email,
                    file_name,
                    result_path,
                    translated_name,
                )
            except Exception as exc:
                notification_status = f"Email failed: {exc}"
        update_translation_job(
            job_id,
            status="completed",
            completed_blocks=sum(1 for block in blocks if should_translate(block.text)),
            completed_batches=total_batches,
            input_tokens=token_usage.input_tokens,
            output_tokens=token_usage.output_tokens,
            total_tokens=token_usage.total_tokens,
            result_file_name=translated_name,
            result_file_path=str(result_path),
            result_mime=mime_type(translated_name),
            notification_status=notification_status,
            progress_message=f"Completed in {format_duration(time.time() - started_at)}.",
            finished_at=time.strftime("%Y-%m-%d %H:%M:%S"),
        )
    except Exception as exc:
        update_translation_job(
            job_id,
            status="failed",
            error_message=format_translation_error(exc),
            finished_at=time.strftime("%Y-%m-%d %H:%M:%S"),
        )


def terminology_report(hits: list[TermHit]) -> pd.DataFrame:
    return pd.DataFrame(
        [{"Japanese": hit.jp, "Required English": hit.en, "Count": hit.count} for hit in hits]
    )


def glossary_version_text() -> str:
    glossary_path = next((path for path in DEFAULT_GLOSSARY_PATHS if path.exists()), None)
    if glossary_path is None:
        return "Glossary file was not found."

    raw = glossary_path.read_bytes()
    digest = hashlib.sha256(raw).hexdigest()[:12]
    modified = glossary_path.stat().st_mtime
    modified_text = datetime.fromtimestamp(modified, ZoneInfo("America/New_York")).strftime(
        "%Y-%m-%d %I:%M %p ET"
    )
    try:
        term_count = max(len(read_glossary(None)) - 1, 0)
    except Exception:
        term_count = "Unavailable"

    return "\n".join(
        [
            "Glossary file: glossary.xlsx",
            f"Last updated: {modified_text}",
            "Updated by: Aoi Minamoto",
            f"Glossary terms: {term_count}",
            f"Version hash: {digest}",
        ]
    )


def plc_rules_version_text() -> str:
    rule_path = next((path for path in DEFAULT_PLC_RULE_PATHS if path.exists()), None)
    if rule_path is None:
        expected = ", ".join(path.name for path in DEFAULT_PLC_RULE_PATHS)
        return f"PLC rule file was not found. Expected one of: {expected}"

    raw = rule_path.read_bytes()
    digest = hashlib.sha256(raw).hexdigest()[:12]
    modified = rule_path.stat().st_mtime
    modified_text = datetime.fromtimestamp(modified, ZoneInfo("America/New_York")).strftime(
        "%Y-%m-%d %I:%M %p ET"
    )
    try:
        rule_count = len(normalize_plc_rules(read_plc_rules()))
    except Exception:
        rule_count = "Unavailable"

    return "\n".join(
        [
            f"PLC rules file: {rule_path.name}",
            f"Last updated: {modified_text}",
            "Updated by: Aoi Minamoto",
            f"PLC rules: {rule_count}",
            f"Version hash: {digest}",
        ]
    )


def apply_compact_style() -> None:
    st.markdown(
        """
        <style>
        div.stButton > button,
        div.stDownloadButton > button {
            width: auto;
            min-width: 190px;
            border-radius: 4px;
            padding: 0.42rem 0.9rem;
            font-weight: 500;
            box-shadow: none;
        }

        div.stButton > button[kind="primary"],
        div.stDownloadButton > button[kind="primary"] {
            background-color: #ff4b4b;
            border: 1px solid #d63b3b;
        }

        div.stButton > button:hover,
        div.stDownloadButton > button:hover {
            box-shadow: 0 1px 2px rgba(15, 23, 42, 0.16);
        }

        div[data-testid="stHorizontalBlock"] {
            align-items: center;
        }

        div[data-testid="stProgress"] > div > div > div {
            height: 14px;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_translation_result(translated_text: str) -> None:
    escaped_text = html.escape(translated_text)
    js_text = json.dumps(translated_text)
    line_count = max(translated_text.count("\n") + 1, 4)
    height = min(max(210, line_count * 28 + 92), 520)
    components.html(
        f"""
        <div style="font-family: Arial, sans-serif;">
          <div style="display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:8px;">
            <div style="font-size:18px;font-weight:700;color:#111827;">English Translation</div>
            <button
              id="copyTranslation"
              style="border:1px solid #cbd5e1;background:#ffffff;color:#111827;border-radius:4px;padding:7px 12px;font-size:14px;cursor:pointer;"
            >
              Copy
            </button>
          </div>
          <div
            style="white-space:pre-wrap;border:1px solid #d0d7de;border-radius:6px;background:#ffffff;color:#111827;padding:18px;font-size:20px;line-height:1.65;min-height:150px;max-height:380px;overflow-y:auto;"
          >{escaped_text}</div>
        </div>
        <script>
          const copyButton = document.getElementById("copyTranslation");
          const translationText = {js_text};
          copyButton.addEventListener("click", async () => {{
            try {{
              if (navigator.clipboard && window.isSecureContext) {{
                await navigator.clipboard.writeText(translationText);
              }} else {{
                const textArea = document.createElement("textarea");
                textArea.value = translationText;
                textArea.style.position = "fixed";
                textArea.style.left = "-9999px";
                document.body.appendChild(textArea);
                textArea.focus();
                textArea.select();
                document.execCommand("copy");
                document.body.removeChild(textArea);
              }}
              copyButton.textContent = "Copied";
              setTimeout(() => copyButton.textContent = "Copy", 1400);
            }} catch (error) {{
              copyButton.textContent = "Copy failed";
              setTimeout(() => copyButton.textContent = "Copy", 1800);
            }}
          }});
        </script>
        """,
        height=height,
    )


def render_text_translation(glossary: pd.DataFrame, plc_rules: pd.DataFrame) -> None:
    translation_mode = st.radio(
        "Translation mode",
        TRANSLATION_MODES,
        horizontal=True,
        help="PLC/SPLC mode standardizes short control comments. General mode is for normal plant text.",
        key="text_translation_mode",
    )
    jp_text = st.text_area(
        "Paste Japanese text",
        height=220,
        placeholder="例: 稼働モニで設備異常を確認してください。",
    )

    if st.button("Translate Text", type="primary"):
        if not jp_text.strip():
            st.warning("Please paste Japanese text first.")
            return

        progress = st.progress(0)
        status = st.empty()
        active_glossary = glossary_for_mode(glossary, plc_rules, translation_mode)

        try:
            status.write("Preparing glossary terms and protected codes...")
            progress.progress(0.2)
            translated_text, hits, token_usage = translate_block(jp_text, active_glossary, translation_mode)
            progress.progress(1.0)
            status.success("Translation complete.")
            if token_usage.total_tokens == 0 and hits:
                st.success("Translated by controlled rule. OpenAI API was not called.")
            render_translation_result(translated_text)

            st.download_button(
                "Download translation",
                data=translated_text.encode("utf-8-sig"),
                file_name="Translated-text.txt",
                mime="text/plain",
            )

            if hits:
                st.subheader("Detected Terminology")
                st.dataframe(terminology_report(hits), use_container_width=True, hide_index=True)
            else:
                st.info("No glossary terms were detected in this text.")
        except Exception as exc:
            status.error("Translation failed.")
            st.error(f"Translation failed: {format_translation_error(exc)}")


def render_document_translation(glossary: pd.DataFrame, plc_rules: pd.DataFrame) -> None:
    translation_mode = st.radio(
        "Translation mode",
        TRANSLATION_MODES,
        horizontal=True,
        key="document_translation_mode",
    )

    active_job_id = st.session_state.get("active_document_job_id")
    if active_job_id:
        detail = translation_job_detail(active_job_id)
        if not detail.empty:
            active_job = detail.iloc[0].to_dict()
            result_path_text = str(active_job.get("result_file_path") or "")
            result_path = Path(result_path_text) if result_path_text else None
            if active_job["status"] == "completed" and result_path is not None and result_path.exists():
                result_file_name = active_job["result_file_name"] or output_file_name(active_job["file_name"])
                st.success("Large file complete. Download ready.")
                st.download_button(
                    "Download Translated File",
                    data=result_path.read_bytes(),
                    file_name=result_file_name,
                    mime=active_job["result_mime"] or mime_type(active_job["file_name"]),
                    type="primary",
                    key=f"active_download_{active_job_id}",
                )
                if active_job["notify_email"]:
                    st.link_button(
                        "Open Email Draft",
                        translation_mailto_link(active_job["notify_email"], active_job["file_name"], result_file_name),
                    )
            elif active_job["status"] == "failed":
                st.error(f"Large file failed: {active_job['error_message'] or 'No error detail.'}")
            else:
                active_done = int(active_job["completed_blocks"] or 0)
                active_total = int(active_job["translatable_blocks"] or 0)
                active_elapsed = elapsed_since_timestamp(active_job.get("created_at", ""))
                st.info(
                    f"Large file running: {active_done}/{active_total} JP blocks · "
                    f"{progress_text(active_done, active_total, active_elapsed)}"
                )
                components.html("<script>setTimeout(() => window.parent.location.reload(), 15000);</script>", height=0)

    uploaded_document = st.file_uploader(
        "Upload Japanese document (Max 100 MB. Files over 5 MB or 1,000 text blocks may run in background)",
        type=["csv", "txt", "as", "docx", "xlsx", "xlsm"],
    )

    if uploaded_document is None:
        return

    raw_document = uploaded_document.getvalue()
    if len(raw_document) > MAX_UPLOAD_BYTES:
        st.warning("This document is larger than the 100 MB safety limit. Please split the file or test with a smaller copy.")
        return

    document_key = f"{document_fingerprint(uploaded_document.name, raw_document)}::{translation_mode}"
    progress_path = checkpoint_path_for(uploaded_document.name, raw_document, translation_mode)
    if st.session_state.get("translated_document_key") != document_key:
        st.session_state.pop("translated_document_bytes", None)
        st.session_state.pop("translated_document_name", None)
        st.session_state.pop("translated_document_mime", None)
        st.session_state.pop("translated_document_preview", None)
        st.session_state.pop("translated_document_terms", None)
        st.session_state.pop("active_document_job_id", None)
        st.session_state["translated_document_key"] = document_key

    try:
        blocks = extract_text_blocks(raw_document, uploaded_document.name)
    except Exception as exc:
        st.error(f"Document error: {exc}")
        return

    translatable_blocks = [block for block in blocks if should_translate(block.text)]
    unique_translatable_count = len({translation_memory_key(block.text) for block in translatable_blocks})
    saved_translations = load_checkpoint(progress_path)
    saved_memory_keys = {
        translation_memory_key(block.text)
        for block in translatable_blocks
        if block.location in saved_translations
    }
    saved_count = sum(1 for block in translatable_blocks if block.location in saved_translations)
    pending_unique_count = len(
        {
            translation_memory_key(block.text)
            for block in translatable_blocks
            if block.location not in saved_translations and translation_memory_key(block.text) not in saved_memory_keys
        }
    )
    batch_count = (pending_unique_count + DOCUMENT_BATCH_SIZE - 1) // DOCUMENT_BATCH_SIZE

    size_or_block_recommendation = (
        len(raw_document) >= EMAIL_DRAFT_RECOMMENDED_BYTES
        or len(blocks) >= EMAIL_DRAFT_RECOMMENDED_BLOCKS
    )
    stat_cols = st.columns(4)
    stat_cols[0].metric("Size", format_file_size(len(raw_document)))
    stat_cols[1].metric("Text blocks", len(blocks))
    stat_cols[2].metric("JP blocks", len(translatable_blocks))
    stat_cols[3].metric("Batches", batch_count)

    smtp_configured = is_smtp_configured()
    delivery_options = ["Download only", "Auto email" if smtp_configured else "Email draft"]
    delivery_choice = st.radio(
        "Delivery",
        delivery_options,
        index=1 if size_or_block_recommendation else 0,
        horizontal=True,
    )
    notify_email = ""
    if delivery_choice != "Download only":
        notify_email = st.text_input("Email", placeholder="name@example.com").strip()
    if size_or_block_recommendation:
        st.caption("Large file: background job recommended.")

    if len(blocks) >= 1000:
        st.info("Large file mode")

    initial_ratio = 1.0 if not translatable_blocks else min(saved_count / len(translatable_blocks), 1.0)
    progress_col, progress_text_col, _ = st.columns([0.34, 0.22, 0.44])
    with progress_col:
        progress = st.progress(initial_ratio)
    progress_info = progress_text_col.empty()
    status = st.empty()
    metrics = st.empty()
    progress_info.write(progress_text(saved_count, len(translatable_blocks)))
    if saved_count:
        status.write("Saved progress found.")
    metrics.write(
        f"{saved_count}/{len(translatable_blocks)} saved. {batch_count} batch(es)."
    )

    action_col, clear_col, _ = st.columns([0.28, 0.22, 0.5])
    with action_col:
        translate_clicked = st.button("Translate Document", type="primary")
    with clear_col:
        clear_clicked = st.button("Clear Saved Progress")

    if clear_clicked:
        if progress_path.exists():
            progress_path.unlink()
        st.session_state.pop("translated_document_bytes", None)
        st.session_state.pop("translated_document_name", None)
        st.session_state.pop("translated_document_mime", None)
        st.session_state.pop("translated_document_preview", None)
        st.session_state.pop("translated_document_terms", None)
        st.success("Saved progress cleared for this document.")
        rerun_app()

    if translate_clicked:
        job_id = ""
        active_glossary = glossary_for_mode(glossary, plc_rules, translation_mode)
        if notify_email and not is_valid_email(notify_email):
            st.warning("Please enter a valid email address, or leave the email field blank.")
            return
        if not blocks:
            st.warning(
                "No translatable text was found in this document. "
                "Please upload a CSV, TXT, AS, DOCX, XLSX, or XLSM file that contains selectable text, not scanned images. "
                "For Excel, save old .xls files as .xlsx first."
            )
            return
        if not translatable_blocks:
            st.info("No Japanese text was found. Downloading the original document is enough.")
            translated_document = build_translated_document(raw_document, uploaded_document.name, {}, blocks)
            translated_name = output_file_name(uploaded_document.name)
            job_id = create_translation_job(
                uploaded_document.name,
                len(raw_document),
                len(blocks),
                0,
                0,
                translation_mode,
                notify_email=notify_email,
            )
            notification_status = ""
            if notify_email:
                result_path = job_result_path(job_id, uploaded_document.name)
                result_path.write_bytes(translated_document)
                try:
                    notification_status = send_completed_translation_email(
                        notify_email,
                        uploaded_document.name,
                        result_path,
                        translated_name,
                    )
                except Exception as exc:
                    notification_status = f"Email failed: {exc}"
            update_translation_job(
                job_id,
                status="completed",
                result_file_name=translated_name,
                notification_status=notification_status,
                finished_at=time.strftime("%Y-%m-%d %H:%M:%S"),
            )
            st.session_state["translated_document_bytes"] = translated_document
            st.session_state["translated_document_name"] = translated_name
            st.session_state["translated_document_mime"] = mime_type(translated_name)
            st.session_state["translated_document_preview"] = []
            st.session_state["translated_document_terms"] = []
            progress.progress(1.0)
            progress_info.write("100%")
            status.success("Download ready.")
        elif size_or_block_recommendation:
            job_id = create_translation_job(
                uploaded_document.name,
                len(raw_document),
                len(blocks),
                len(translatable_blocks),
                batch_count,
                translation_mode,
                notify_email=notify_email,
                status="pending",
            )
            source_path = job_upload_path(job_id, uploaded_document.name)
            source_path.write_bytes(raw_document)
            update_translation_job(job_id, source_file_path=str(source_path))
            background_job_executor().submit(
                run_document_translation_job,
                job_id,
                raw_document,
                uploaded_document.name,
                blocks,
                active_glossary,
                translation_mode,
                progress_path,
                batch_count,
            )
            st.session_state["active_document_job_id"] = job_id
            status.success(f"Large file job started: {job_id}")
            metrics.write("This page checks every 15 seconds. Download appears here when complete.")
        else:
            job_id = create_translation_job(
                uploaded_document.name,
                len(raw_document),
                len(blocks),
                len(translatable_blocks),
                batch_count,
                translation_mode,
                notify_email=notify_email,
                status="running",
            )
            source_path = job_upload_path(job_id, uploaded_document.name)
            source_path.write_bytes(raw_document)
            update_translation_job(job_id, source_file_path=str(source_path))

            def update_foreground_progress(done, total, done_batches, total_batches, elapsed, message):
                ratio = 1.0 if total == 0 else min(done / total, 1.0)
                progress.progress(ratio)
                progress_info.write(progress_text(done, total, elapsed))
                status.write(message)
                metrics.write(f"{done}/{total} JP blocks. {done_batches}/{total_batches} batch(es).")
                update_translation_job(
                    job_id,
                    status="running",
                    completed_blocks=done,
                    completed_batches=done_batches,
                    progress_message=message,
                )

            started_at = time.time()
            try:
                translations, all_hits, token_usage = translate_blocks_batch(
                    blocks,
                    active_glossary,
                    translation_mode,
                    checkpoint_path=progress_path,
                    progress_callback=update_foreground_progress,
                )
                translated_document = build_translated_document(raw_document, uploaded_document.name, translations, blocks)
                translated_name = output_file_name(uploaded_document.name)
                result_path = job_result_path(job_id, uploaded_document.name)
                result_path.write_bytes(translated_document)
                notification_status = ""
                if notify_email:
                    try:
                        notification_status = send_completed_translation_email(
                            notify_email,
                            uploaded_document.name,
                            result_path,
                            translated_name,
                        )
                    except Exception as exc:
                        notification_status = f"Email failed: {exc}"
                update_translation_job(
                    job_id,
                    status="completed",
                    completed_blocks=len(translatable_blocks),
                    completed_batches=batch_count,
                    input_tokens=token_usage.input_tokens,
                    output_tokens=token_usage.output_tokens,
                    total_tokens=token_usage.total_tokens,
                    result_file_name=translated_name,
                    result_file_path=str(result_path),
                    result_mime=mime_type(translated_name),
                    notification_status=notification_status,
                    progress_message=f"Completed in {format_duration(time.time() - started_at)}.",
                    finished_at=time.strftime("%Y-%m-%d %H:%M:%S"),
                )
                st.session_state["translated_document_bytes"] = translated_document
                st.session_state["translated_document_name"] = translated_name
                st.session_state["translated_document_mime"] = mime_type(translated_name)
                st.session_state["translated_document_preview"] = []
                st.session_state["translated_document_terms"] = []
                progress.progress(1.0)
                progress_info.write("100%")
                status.success("Download ready.")
                metrics.write("Translation complete.")
            except Exception as exc:
                update_translation_job(
                    job_id,
                    status="failed",
                    error_message=format_translation_error(exc),
                    finished_at=time.strftime("%Y-%m-%d %H:%M:%S"),
                )
                status.error("Translation failed.")
                st.error(f"Translation failed: {format_translation_error(exc)}")

    if st.session_state.get("translated_document_bytes"):
        st.success("Download ready. Click the button below to save the translated file.")
        st.download_button(
            "Download Translated File",
            data=st.session_state["translated_document_bytes"],
            file_name=st.session_state["translated_document_name"],
            mime=st.session_state["translated_document_mime"],
            type="primary",
        )



load_env()

st.set_page_config(page_title="JP to EN Translator", layout="wide")
apply_compact_style()
usage_count = increment_usage_count_once()
st.title("JP to EN Plant Translator")
st.caption("Type Japanese text or upload Word, Excel, or TXT files. The app applies company terminology first, then translates with OpenAI API.")
st.warning(
    "Data notice: Text entered or uploaded in this app is sent to OpenAI API for translation. "
    "Do not upload confidential or restricted information unless approved by company policy.",
)

with st.sidebar:
    st.metric("App use times", usage_count)
    st.header("Knowledge Base")
    st.success("Internal glossary is already loaded.")
    st.code(glossary_version_text(), language="text")
    st.success("PLC/SPLC rules are controlled by the app owner.")
    st.code(plc_rules_version_text(), language="text")

try:
    glossary = normalize_glossary(read_glossary(None))
except Exception as exc:
    st.error(f"Glossary error: {exc}")
    st.stop()

plc_rules_error = ""
try:
    plc_rules = normalize_plc_rules(read_plc_rules())
except Exception as exc:
    plc_rules_error = str(exc)
    plc_rules = empty_terms_dataframe()

if plc_rules_error:
    st.warning(f"PLC rule file could not be loaded, so the app will continue without PLC abbreviation rules: {plc_rules_error}")

text_tab, document_tab = st.tabs(["Text Translation", "Document Translation"])

with text_tab:
    render_text_translation(glossary, plc_rules)

with document_tab:
    render_document_translation(glossary, plc_rules)
